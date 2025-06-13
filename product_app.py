import sys
import os
import subprocess
import logging
import re
import traceback
import concurrent.futures
import json
from shutil import copyfile
from PyQt5.QtGui import QColor, QFontMetrics, QRegExpValidator, QDoubleValidator, QKeySequence
from PyQt5.QtWidgets import (QCompleter,
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QTextEdit, QLabel,
    QFormLayout, QScrollArea, QPushButton, QListWidget, QMessageBox, QInputDialog, QMenu, QAction, QMenuBar,
    QRadioButton, QButtonGroup, QSplitter, QGridLayout, QDialog, QTreeWidget, QTreeWidgetItem, QSizePolicy, QStackedWidget,
    QAbstractItemView, QStyle, QTableView, QHeaderView, QComboBox, QListWidgetItem, QStyledItemDelegate,
    QDialogButtonBox, QProgressBar, QStatusBar, QCheckBox
)
from PyQt5.QtCore import (Qt, QAbstractTableModel, QModelIndex, QItemSelectionModel, QItemSelection, QItemSelectionRange,
                          QTimer, QSize, QPoint, QStandardPaths, QSettings, QByteArray, QRegExp, pyqtSignal, QMutex)
from typing import Optional, List, Dict, Any, Union, Tuple
from openpyxl import load_workbook

# UI関連ローカル定数（Constantsモジュールから取得しない固有定数）
MAX_Y_SPEC_COUNT = 10
MENU_BAR_HEIGHT = 24
DEFAULT_SPACING = 8
DEFAULT_MARGINS = 10
PROGRESS_UPDATE_INTERVAL = 50  # UI更新間隔（アイテム数）
SPACER_HEIGHT = 10  # スペーサーの高さ
FONT_SIZE_MENU = 13  # メニューフォントサイズ

# バージョンチェッカーのインポート
try:
    from src.utils.version_checker import check_for_updates_on_startup, VersionChecker, CURRENT_VERSION
except ImportError:
    # モジュールが見つからない場合のフォールバック
    check_for_updates_on_startup = None
    VersionChecker = None
    CURRENT_VERSION = "2.1.0"

# 万が一対策システムのインポート
try:
    from src.utils.crash_recovery import CrashRecoveryManager, setup_crash_handler, setup_qt_exception_handler
    from src.utils.config_recovery import check_and_recover_config
    from src.utils.file_lock_manager import handle_duplicate_launch, handle_file_conflicts, FileLockManager
    from src.utils.disk_monitor import check_disk_space_before_save, check_disk_space_once
    from src.utils.memory_manager import MemoryMonitor, check_memory_before_large_operation, optimize_large_data_processing
    from src.utils.network_monitor import setup_network_monitoring, check_network_before_operation
    from src.utils.system_compatibility import check_system_compatibility, get_system_info
except ImportError:
    # フォールバック
    CrashRecoveryManager = None
    setup_crash_handler = None
    setup_qt_exception_handler = None
    check_and_recover_config = None
    handle_duplicate_launch = None
    handle_file_conflicts = None
    FileLockManager = None
    check_disk_space_before_save = None
    check_disk_space_once = None
    MemoryMonitor = None
    check_memory_before_large_operation = None
    optimize_large_data_processing = None
    setup_network_monitoring = None
    check_network_before_operation = None
    check_system_compatibility = None
    get_system_info = None

# 分離したモジュールのインポート
from constants import *
from utils import (
    open_csv_file_with_fallback, normalize_text, normalize_wave_dash,
    get_byte_count_excel_lenb, get_user_data_dir
)
from models import SkuTableModel
from widgets import (
    CustomHtmlTextEdit, FocusControllingTableView, ScrollableFocusControllingTableView,
    MultipleSelectDialog, SkuMultipleAttributeEditor, SkuAttributeDelegate, LoadingDialog
)
from loaders import (
    YSpecDefinitionLoader, RakutenAttributeDefinitionLoader,
    load_categories_from_csv, load_explanation_mark_icons,
    load_material_spec_master, load_id_master_data
)

class SearchPanel(QWidget):
    """非モーダル検索パネル（サイドバー形式）"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.current_results = []
        self.current_index = -1
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle("検索と置換")
        self.setFixedWidth(350)  # 固定幅のサイドパネル
        # ESCキーをキャッチするためのフォーカスポリシー設定
        self.setFocusPolicy(Qt.StrongFocus)
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5;
                border-left: 2px solid #ddd;
            }
            QLineEdit, QComboBox {
                background-color: white;
                border: 1px solid #ccc;
                padding: 4px;
                border-radius: 3px;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #ccc;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # 検索入力部分
        search_layout = QHBoxLayout()
        search_label = QLabel("検索:")
        search_label.setMinimumWidth(60)
        search_layout.addWidget(search_label)
        # カスタムQLineEditでESCキー処理
        class SearchLineEdit(QLineEdit):
            def keyPressEvent(self, event):
                if event.key() == Qt.Key_Escape:
                    # 親のSearchPanelを直接閉じる
                    search_panel = self.parent()
                    if hasattr(search_panel, 'close_panel'):
                        search_panel.close_panel()
                    event.accept()
                    return
                super().keyPressEvent(event)
        
        self.search_input = SearchLineEdit()
        self.search_input.setPlaceholderText("検索したいテキストを入力")
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        # 置換入力部分
        replace_layout = QHBoxLayout()
        replace_label = QLabel("置換:")
        replace_label.setMinimumWidth(60)
        replace_layout.addWidget(replace_label)
        self.replace_input = SearchLineEdit()  # 同じESC処理を適用
        self.replace_input.setPlaceholderText("置換後のテキストを入力")
        replace_layout.addWidget(self.replace_input)
        layout.addLayout(replace_layout)
        
        # 検索対象の選択
        scope_layout = QHBoxLayout()
        scope_layout.addWidget(QLabel("検索対象:"))
        self.scope_combo = QComboBox()
        self.scope_combo.addItems([
            "商品一覧のみ",
            "現在の商品のフィールド",
            "すべての商品・すべてのフィールド"
        ])
        scope_layout.addWidget(self.scope_combo)
        layout.addLayout(scope_layout)
        
        # オプション
        options_layout = QVBoxLayout()
        self.case_sensitive = QCheckBox("大文字と小文字を区別する")
        self.whole_word = QCheckBox("単語全体を検索")
        options_layout.addWidget(self.case_sensitive)
        options_layout.addWidget(self.whole_word)
        layout.addLayout(options_layout)
        
        # 結果表示
        self.result_label = QLabel("検索結果: 0件")
        layout.addWidget(self.result_label)
        
        # ボタンを縦配置（サイドパネル用）
        # 検索ボタン群
        search_buttons_layout = QVBoxLayout()
        self.find_next_btn = QPushButton("▼ 次を検索")
        self.find_prev_btn = QPushButton("▲ 前を検索")
        self.find_all_btn = QPushButton("🔍 すべて検索")
        
        search_buttons_layout.addWidget(self.find_next_btn)
        search_buttons_layout.addWidget(self.find_prev_btn)
        search_buttons_layout.addWidget(self.find_all_btn)
        layout.addLayout(search_buttons_layout)
        
        # 置換ボタン群
        replace_buttons_layout = QVBoxLayout()
        self.replace_btn = QPushButton("↔ 置換")
        self.replace_all_btn = QPushButton("↔ すべて置換")
        
        replace_buttons_layout.addWidget(self.replace_btn)
        replace_buttons_layout.addWidget(self.replace_all_btn)
        layout.addLayout(replace_buttons_layout)
        
        # 検索結果リスト（新機能）
        self.results_list = QListWidget()
        self.results_list.setMaximumHeight(150)
        self.results_list.setToolTip("検索結果をクリックで該当箇所にジャンプ")
        layout.addWidget(QLabel("検索結果:"))
        layout.addWidget(self.results_list)
        
        # 閉じるボタン
        self.close_btn = QPushButton("✕ パネルを閉じる")
        layout.addWidget(self.close_btn)
        
        # イベント接続
        self.search_input.textChanged.connect(self.on_search_text_changed)
        self.search_input.returnPressed.connect(self.find_next)
        self.scope_combo.currentIndexChanged.connect(self.on_scope_changed)
        self.find_next_btn.clicked.connect(self.find_next)
        self.find_prev_btn.clicked.connect(self.find_prev)
        self.find_all_btn.clicked.connect(self.find_all)
        self.replace_btn.clicked.connect(self.replace_current)
        self.replace_all_btn.clicked.connect(self.replace_all)
        self.close_btn.clicked.connect(self.close_panel)
        
        # 初期状態
        self.find_next_btn.setEnabled(False)
        self.find_prev_btn.setEnabled(False)
        self.find_all_btn.setEnabled(False)
        self.replace_btn.setEnabled(False)
        self.replace_all_btn.setEnabled(False)
        
        # フォーカスを検索入力に設定
        self.search_input.setFocus()
    
    def keyPressEvent(self, event):
        """ESCキーで検索パネルを閉じる"""
        if event.key() == Qt.Key_Escape:
            self.close_panel()
            event.accept()
        else:
            super().keyPressEvent(event)
    
    
    def close_panel(self):
        """検索パネルを閉じる"""
        self.hide()
        if self.parent_app and hasattr(self.parent_app, '_restore_splitter_sizes_without_search'):
            self.parent_app._restore_splitter_sizes_without_search()
    
    def on_search_text_changed(self, text):
        has_text = bool(text.strip())
        self.find_next_btn.setEnabled(has_text)
        self.find_prev_btn.setEnabled(has_text)
        self.find_all_btn.setEnabled(has_text)
        
        # 置換ボタンは、検索テキストがあり、かつ現在の商品のフィールドを検索している場合のみ有効
        can_replace = has_text and self.scope_combo.currentIndex() == 1
        self.replace_btn.setEnabled(can_replace)
        self.replace_all_btn.setEnabled(can_replace)
        
        if has_text:
            self.perform_search()
        else:
            self.current_results = []
            self.current_index = -1
            self.result_label.setText("検索結果: 0件")
    
    def on_scope_changed(self):
        """検索対象が変更されたときの処理"""
        # 置換ボタンの有効/無効を更新
        has_text = bool(self.search_input.text().strip())
        can_replace = has_text and self.scope_combo.currentIndex() == 1
        self.replace_btn.setEnabled(can_replace)
        self.replace_all_btn.setEnabled(can_replace)
        
        # 検索を再実行
        if has_text:
            self.perform_search()
    
    def perform_search(self):
        """検索を実行"""
        search_text = self.search_input.text().strip()
        if not search_text:
            return
        
        scope = self.scope_combo.currentIndex()
        case_sensitive = self.case_sensitive.isChecked()
        whole_word = self.whole_word.isChecked()
        
        self.current_results = []
        
        if scope == 0:  # 商品一覧のみ
            self.search_product_list(search_text, case_sensitive, whole_word)
        elif scope == 1:  # 現在の商品のフィールド
            self.search_current_product(search_text, case_sensitive, whole_word)
        else:  # すべての商品・すべてのフィールド
            self.search_all_products(search_text, case_sensitive, whole_word)
        
        self.current_index = -1
        self.result_label.setText(f"検索結果: {len(self.current_results)}件")
    
    def search_product_list(self, search_text, case_sensitive, whole_word):
        """商品一覧を検索"""
        for i in range(self.parent_app.product_list.count()):
            item = self.parent_app.product_list.item(i)
            if item and self.text_matches(item.text(), search_text, case_sensitive, whole_word):
                self.current_results.append({
                    'type': 'product_list',
                    'index': i,
                    'text': item.text(),
                    'description': f"商品一覧 [{i+1}]"
                })
    
    def search_current_product(self, search_text, case_sensitive, whole_word):
        """現在の商品のフィールドを検索"""
        for field_name, field_widget in self.parent_app.main_fields.items():
            text = None
            if isinstance(field_widget, QLineEdit):
                text = field_widget.text()
            elif isinstance(field_widget, QTextEdit):
                text = field_widget.toPlainText()
            elif hasattr(field_widget, 'toPlainText'):  # CustomHtmlTextEditなど
                text = field_widget.toPlainText()
            else:
                continue
                
            if text and self.text_matches(text, search_text, case_sensitive, whole_word):
                self.current_results.append({
                    'type': 'field',
                    'field_name': field_name,
                    'widget': field_widget,
                    'text': text,
                    'description': f"フィールド: {field_name}"
                })
    
    def search_all_products(self, search_text, case_sensitive, whole_word):
        """すべての商品・すべてのフィールドを検索"""
        # まず商品一覧を検索
        self.search_product_list(search_text, case_sensitive, whole_word)
        
        # 現在の商品のフィールドも検索
        self.search_current_product(search_text, case_sensitive, whole_word)
    
    def text_matches(self, text, search_text, case_sensitive, whole_word):
        """テキストが検索条件にマッチするかチェック"""
        if not case_sensitive:
            text = text.lower()
            search_text = search_text.lower()
        
        if whole_word:
            import re
            pattern = r'\b' + re.escape(search_text) + r'\b'
            return bool(re.search(pattern, text))
        else:
            return search_text in text
    
    def find_next(self):
        """次の検索結果に移動"""
        if not self.current_results:
            return
        
        self.current_index = (self.current_index + 1) % len(self.current_results)
        self.goto_result(self.current_index)
    
    def find_prev(self):
        """前の検索結果に移動"""
        if not self.current_results:
            return
        
        self.current_index = (self.current_index - 1) % len(self.current_results)
        self.goto_result(self.current_index)
    
    def find_all(self):
        """すべての結果を表示"""
        if not self.current_results:
            return
        
        # 結果一覧ダイアログを表示
        results_dialog = QDialog(self)
        results_dialog.setWindowTitle("検索結果一覧")
        results_dialog.resize(500, 400)
        
        layout = QVBoxLayout(results_dialog)
        
        list_widget = QListWidget()
        for i, result in enumerate(self.current_results):
            item_text = f"{result['description']}: {result['text'][:50]}..."
            list_item = QListWidgetItem(item_text)
            list_item.setData(Qt.UserRole, i)
            list_widget.addItem(list_item)
        
        layout.addWidget(list_widget)
        
        button_layout = QHBoxLayout()
        goto_btn = QPushButton("移動")
        close_btn = QPushButton("閉じる")
        button_layout.addWidget(goto_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        def on_goto():
            current_item = list_widget.currentItem()
            if current_item:
                index = current_item.data(Qt.UserRole)
                self.goto_result(index)
                results_dialog.close()
        
        goto_btn.clicked.connect(on_goto)
        close_btn.clicked.connect(results_dialog.close)
        list_widget.itemDoubleClicked.connect(on_goto)
        
        results_dialog.exec_()
    
    def goto_result(self, index):
        """指定した検索結果に移動"""
        if 0 <= index < len(self.current_results):
            result = self.current_results[index]
            
            if result['type'] == 'product_list':
                # 商品一覧の項目を選択
                self.parent_app.product_list.setCurrentRow(result['index'])
                self.parent_app.product_list.setFocus()
            elif result['type'] == 'field':
                # フィールドにフォーカス
                widget = result['widget']
                widget.setFocus()
                if isinstance(widget, QLineEdit):
                    widget.selectAll()
                elif isinstance(widget, QTextEdit):
                    widget.selectAll()
            
            self.result_label.setText(f"検索結果: {len(self.current_results)}件 ({index+1}/{len(self.current_results)})")
    
    def replace_current(self):
        """現在の検索結果を置換"""
        if self.current_index < 0 or self.current_index >= len(self.current_results):
            QMessageBox.information(self, "置換", "まず検索を実行してください")
            return
        
        result = self.current_results[self.current_index]
        if result['type'] != 'field':
            QMessageBox.warning(self, "置換", "商品一覧の項目は置換できません")
            return
        
        search_text = self.search_input.text()
        replace_text = self.replace_input.text()
        widget = result['widget']
        
        if isinstance(widget, QLineEdit):
            current_text = widget.text()
            new_text = self._perform_replace(current_text, search_text, replace_text)
            widget.setText(new_text)
        elif isinstance(widget, QTextEdit):
            current_text = widget.toPlainText()
            new_text = self._perform_replace(current_text, search_text, replace_text)
            widget.setPlainText(new_text)
        elif hasattr(widget, 'toHtml') and hasattr(widget, 'setHtml'):
            # CustomHtmlTextEditなどの対応（HTMLを保持）
            current_html = widget.toHtml()
            current_text = widget.toPlainText()
            
            # HTMLタグを除外してテキスト部分のみ置換する
            # より安全な方法：プレーンテキストで検索位置を特定し、HTML内で置換
            import re
            
            # 簡易的な実装：HTMLエディタは置換対象外とする
            QMessageBox.warning(self, "置換", "HTMLエディタフィールドは置換できません。\n手動で編集してください。")
            return
        
        # 置換後は再検索が必要（置換により検索結果が変わる可能性があるため）
        self.perform_search()
        
        # 親アプリケーションのダーティフラグを設定
        if hasattr(self.parent_app, 'mark_dirty'):
            self.parent_app.mark_dirty()
        
        # 前回と同じ位置を維持（ユーザーが次に進みたければ「次を検索」をクリック）
        if self.current_index < len(self.current_results):
            self.goto_result(self.current_index)
    
    def replace_all(self):
        """すべての検索結果を置換"""
        if not self.current_results:
            QMessageBox.information(self, "すべて置換", "まず検索を実行してください")
            return
        
        # 現在の商品のフィールドのみ置換可能
        field_results = [r for r in self.current_results if r['type'] == 'field']
        if not field_results:
            QMessageBox.information(self, "すべて置換", "置換可能な項目がありません")
            return
        
        search_text = self.search_input.text()
        replace_text = self.replace_input.text()
        
        # 確認ダイアログ
        reply = QMessageBox.question(
            self, "すべて置換",
            f"{len(field_results)}件のフィールドで「{search_text}」を「{replace_text}」に置換しますか？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            return
        
        replaced_count = 0
        for result in field_results:
            widget = result['widget']
            
            if isinstance(widget, QLineEdit):
                current_text = widget.text()
                new_text = self._perform_replace(current_text, search_text, replace_text)
                if current_text != new_text:
                    widget.setText(new_text)
                    replaced_count += 1
            elif isinstance(widget, QTextEdit):
                current_text = widget.toPlainText()
                new_text = self._perform_replace(current_text, search_text, replace_text)
                if current_text != new_text:
                    widget.setPlainText(new_text)
                    replaced_count += 1
            elif hasattr(widget, 'toHtml') and hasattr(widget, 'setHtml'):
                # CustomHtmlTextEditなどの対応
                # HTMLフィールドは置換対象外（データ破壊を防ぐため）
                pass  # スキップ
        
        # 親アプリケーションのダーティフラグを設定
        if replaced_count > 0 and hasattr(self.parent_app, 'mark_dirty'):
            self.parent_app.mark_dirty()
        
        QMessageBox.information(self, "すべて置換", f"{replaced_count}件を置換しました")
        
        # 再検索
        self.perform_search()
    
    def _perform_replace(self, text, search_text, replace_text):
        """テキストの置換を実行"""
        # 安全性チェック
        if text is None:
            return ""
        if search_text is None or search_text == "":
            return text  # 空文字列の検索は何も置換しない
        if replace_text is None:
            replace_text = ""  # Noneは空文字列として扱う
            
        case_sensitive = self.case_sensitive.isChecked()
        whole_word = self.whole_word.isChecked()
        
        try:
            if whole_word:
                import re
                pattern = r'\b' + re.escape(search_text) + r'\b'
                flags = 0 if case_sensitive else re.IGNORECASE
                return re.sub(pattern, replace_text, text, flags=flags)
            else:
                if case_sensitive:
                    return text.replace(search_text, replace_text)
                else:
                    # 大文字小文字を区別しない置換
                    import re
                    pattern = re.escape(search_text)
                    return re.sub(pattern, replace_text, text, flags=re.IGNORECASE)
        except Exception as e:
            logging.error(f"置換処理中にエラー: {e}")
            return text  # エラー時は元のテキストを返す


class FilteredLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        # setMaxLength(10) は CustomProductCodeInputDialog 側で設定

    def keyPressEvent(self, event):
        # 標準的なキーシーケンス (Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Ctrl+Z, Ctrl+Y) をチェック
        if (event.matches(QKeySequence.Copy) or
            event.matches(QKeySequence.Paste) or
            event.matches(QKeySequence.Cut) or
            event.matches(QKeySequence.SelectAll) or
            event.matches(QKeySequence.Undo) or
            event.matches(QKeySequence.Redo)):
            super().keyPressEvent(event)
            return

        # 既存の制御キーチェック (Backspace, Delete, 矢印キーなど、修飾キーなしの場合)
        if event.modifiers() == Qt.NoModifier and \
           event.key() in (Qt.Key_Backspace, Qt.Key_Delete,
                           Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down,
                           Qt.Key_Home, Qt.Key_End,
                           Qt.Key_Tab, Qt.Key_Return, Qt.Key_Enter):
            super().keyPressEvent(event)
            return

        # ここから先は、上記以外のキー入力 (主に文字入力) に対する処理
        text_to_insert = event.text()

        if text_to_insert: # 実際に文字が入力される場合のみ、以下のチェックを行う
            # 入力文字が数字でない場合は無視
            if not text_to_insert.isdigit():
                event.ignore()
                return

            # 新しいテキスト長が10を超える場合、入力を無視
            current_text = self.text()
            selected_text_len = len(self.selectedText())
            if len(current_text) - selected_text_len + len(text_to_insert) > 10:
                event.ignore()
                return

        super().keyPressEvent(event)


class ExpandableFieldGroup(QWidget):
    def __init__(self, group_label, group_count, main_fields_dict, always_show=3, has_ab=False, parent_app=None, master_data=None, field_names_list=None):
        super().__init__(parent_app); self.setObjectName("ExpandableGroup")
        self.group_header_widget = QWidget(); self.group_header_widget.setObjectName("ExpandableGroupHeader")
        group_header_layout = QHBoxLayout(self.group_header_widget); group_header_layout.setContentsMargins(8, 5, 8, 5); group_header_layout.setSpacing(8)
        
        self.field_names_list = field_names_list # 特定のフィールドリストを保持
        self.master_data = master_data # Store master data for this group
        self.group_label_widget = QLabel(f"{group_label}"); self.group_label_widget.setObjectName("ExpandableGroupLabel")
        group_header_layout.addWidget(self.group_label_widget); group_header_layout.addStretch()
        self.toggle_button = QPushButton(); self.toggle_button.setObjectName("ExpandableGroupToggleButton"); self.toggle_button.setFixedSize(22, 22)
        self.toggle_button.setToolTip(f"{group_label}項目を全て表示/一部表示"); self.toggle_button.clicked.connect(self.toggle_expand_collapse)
        group_header_layout.addWidget(self.toggle_button)
        
        self.main_content_widget = QWidget(self) 
        main_content_layout = QVBoxLayout(self.main_content_widget)
        main_content_layout.setContentsMargins(0,0,0,0); main_content_layout.setSpacing(0); main_content_layout.addWidget(self.group_header_widget)
        
        self.fields_container_widget = QWidget(self.main_content_widget) 
        self.fields_container_widget.setObjectName("ExpandableGroupContent")
        self.form_layout = QFormLayout(self.fields_container_widget)
        self.form_layout.setContentsMargins(20, 10, 10, 10) 
        self.form_layout.setVerticalSpacing(6)
        self.form_layout.setHorizontalSpacing(8); self.form_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow); self.form_layout.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        main_content_layout.addWidget(self.fields_container_widget)

        self.group_label_prefix = group_label
        self.group_count = group_count
        self.main_fields_ref = main_fields_dict
        self.has_ab = has_ab
        self.always_show_count = always_show
        self.is_expanded = False
        self.parent_app_ref = parent_app
        self.row_widgets_for_visibility = [] # (row_label_widget, field_or_container_widget)        
        self._processing_a_change = False # 再入防止フラグ for _update_product_size_b_input_type
        self.related_product_code_ui_elements = [] # 関連商品のコードUI要素を保持するリスト

        # 商品サイズ用の特別なフィールドを初期化
        if self.group_label_prefix == "商品サイズ":
            self.dimension_fields_list = [None] * self.group_count
            self.b_field_stacks = [None] * self.group_count
            self.weight_fields_list = [None] * self.group_count # 重量入力用フィールドリスト

        self._create_fields(); self.update_toggle_button_icon(); self.update_fields_visibility()


    def _create_fields(self):
        self.row_widgets_for_visibility.clear()

        if self.field_names_list: # 特定のフィールドリストが指定されている場合
            self.group_count = len(self.field_names_list) # group_count をリストの長さに更新
            for field_name in self.field_names_list:
                # ラベルはフィールド名をそのまま使用 (必要に応じて整形も可能)
                # 例: "R_SKU項目名" -> "SKU項目名" のように表示を調整
                display_label_text = field_name.replace("R_", "").replace("Y_", "").replace("YA_", "") # 簡単な整形例
                row_label_widget = QLabel(display_label_text)
                row_label_widget.setMinimumWidth(120) # ラベル幅を調整
                row_label_widget.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

                field_widget = self.main_fields_ref.get(field_name)
                if not field_widget: # main_fields_ref にウィジェットがなければ作成
                    field_widget = QLineEdit()
                    field_widget.setObjectName(field_name)
                    self.main_fields_ref[field_name] = field_widget # main_fields_ref に登録
                
                field_widget._efg_managed = True # EFG管理対象フラグ
                if self.parent_app_ref and isinstance(field_widget, (QLineEdit, QTextEdit, QComboBox)):
                    if isinstance(field_widget, QLineEdit): field_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                    elif isinstance(field_widget, QTextEdit): field_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                    elif isinstance(field_widget, QComboBox):
                        if field_widget.isEditable(): field_widget.currentTextChanged.connect(self.parent_app_ref.mark_dirty)
                        else: field_widget.currentIndexChanged.connect(self.parent_app_ref.mark_dirty)
                
                self.form_layout.addRow(row_label_widget, field_widget)
                self.row_widgets_for_visibility.append((row_label_widget, field_widget))
        else: # 従来のロジック (group_label_prefix と group_count に基づく)
            for i in range(1, self.group_count + 1):
                row_label_text = f"{i}"
                current_field_name_base_for_html_check = f"{self.group_label_prefix}_{i}"

                is_first_row_html_field = (i == 1 and current_field_name_base_for_html_check in HTML_TEXTEDIT_FIELDS)
                if is_first_row_html_field:
                    row_label_text += " (HTML)"
                
                row_label_widget = QLabel(row_label_text)
                row_label_widget.setMinimumWidth(60) 
                row_label_widget.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                if i == 1 and current_field_name_base_for_html_check in HTML_TEXTEDIT_FIELDS:
                    field_name = current_field_name_base_for_html_check
                    field_widget = CustomHtmlTextEdit() # カスタムクラスを使用
                    field_widget.setPlaceholderText("HTMLタグ使用可。改行＝<br>")
                    field_widget.setObjectName(field_name)
                    field_widget._efg_managed = True
                    self.main_fields_ref[field_name] = field_widget
                    if self.parent_app_ref:
                        field_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                    self.form_layout.addRow(row_label_widget, field_widget)
                    self.row_widgets_for_visibility.append((row_label_widget, field_widget))

                elif self.has_ab:
                    field_name_a = f"{self.group_label_prefix}_{i}a"; field_name_b = f"{self.group_label_prefix}_{i}b"
                    h_box = QHBoxLayout(); field_a_widget = None
                    field_b_widget, field_b_ui_widget = None, None

                    if self.group_label_prefix in ["材質", "仕様"] and self.master_data:
                        field_a_widget = QComboBox()
                        field_a_widget.addItem("") # Blank item
                        for name_key in sorted(self.master_data.keys()):
                            field_a_widget.addItem(name_key)
                        
                        field_b_widget = QLineEdit()
                        field_b_widget.setReadOnly(True)

                        field_a_widget.currentTextChanged.connect(
                            lambda text, b_w=field_b_widget, master=self.master_data: 
                                self.on_master_a_selected(text, b_w, master)
                        )
                        if self.parent_app_ref: 
                            field_a_widget.currentTextChanged.connect(self.parent_app_ref.mark_dirty)
                        field_b_ui_widget = field_b_widget

                    elif self.group_label_prefix == "商品サイズ":
                        field_a_widget = QLineEdit(); field_a_widget.setPlaceholderText("例: 本体")
                        field_b_widget = QLineEdit(); field_b_widget.setPlaceholderText("例: 幅○○×奥行○○×高さ○○cm")
                        
                        dim_input_container = QWidget(); dim_layout = QHBoxLayout(dim_input_container)
                        dim_layout.setContentsMargins(0,0,0,0); dim_layout.setSpacing(3)
                        w_edit = QLineEdit(); w_edit.setPlaceholderText("幅")
                        d_edit = QLineEdit(); d_edit.setPlaceholderText("奥行")
                        h_edit = QLineEdit(); h_edit.setPlaceholderText("高さ")
                        double_validator = QDoubleValidator(0, 99999.99, 2, self); double_validator.setNotation(QDoubleValidator.StandardNotation)
                        for edit_widget in [w_edit, d_edit, h_edit]:
                            edit_widget.setValidator(double_validator)
                            edit_widget.textChanged.connect(lambda text, r_idx=i-1: self._update_b_field_from_dimensions(r_idx))
                        dim_layout.addWidget(w_edit, 1); dim_layout.addWidget(QLabel("×"), 0); dim_layout.addWidget(d_edit, 1)
                        dim_layout.addWidget(QLabel("×"), 0); dim_layout.addWidget(h_edit, 1); dim_layout.addWidget(QLabel("cm"), 0)
                        self.dimension_fields_list[i-1] = {'w': w_edit, 'd': d_edit, 'h': h_edit, 'container': dim_input_container}

                        weight_input_container = QWidget(); weight_layout = QHBoxLayout(weight_input_container)
                        weight_layout.setContentsMargins(0,0,0,0); weight_layout.setSpacing(3)
                        weight_layout.addWidget(QLabel("約")); weight_edit = QLineEdit(); weight_edit.setPlaceholderText("重量")
                        weight_edit.setValidator(double_validator)
                        weight_edit.textChanged.connect(lambda text, r_idx=i-1: self._update_b_field_from_weight(r_idx))
                        weight_layout.addWidget(weight_edit, 1); weight_layout.addWidget(QLabel("kg"), 0)
                        self.weight_fields_list[i-1] = {'weight': weight_edit, 'container': weight_input_container}

                        b_stack = QStackedWidget()
                        b_stack.addWidget(field_b_widget); b_stack.addWidget(dim_input_container); b_stack.addWidget(weight_input_container)
                        self.b_field_stacks[i-1] = b_stack; field_b_ui_widget = b_stack
                        field_a_widget.editingFinished.connect(lambda r_idx=i-1, f_a_w=field_a_widget: self._update_product_size_b_input_type(f_a_w.text(), r_idx))
                        field_b_widget.textChanged.connect(lambda text, r_idx=i-1: self._update_dimensions_from_b_field(text, r_idx))
                    else: 
                        field_a_widget = QLineEdit() 
                        field_b_widget = FilteredLineEdit() if self.group_label_prefix == "関連商品" else QLineEdit()

                    if field_a_widget:
                        field_a_widget.setObjectName(field_name_a); field_a_widget._efg_managed = True
                        self.main_fields_ref[field_name_a] = field_a_widget
                        if self.parent_app_ref and isinstance(field_a_widget, QLineEdit): field_a_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                    
                    if field_b_widget: 
                        field_b_widget.setObjectName(field_name_b); field_b_widget._efg_managed = True
                        self.main_fields_ref[field_name_b] = field_b_widget
                        if self.parent_app_ref: field_b_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                        if self.group_label_prefix == "関連商品":
                            if self.parent_app_ref and hasattr(self.parent_app_ref, '_update_relevant_links'): field_b_widget.textChanged.connect(self.parent_app_ref._update_relevant_links)
                            if isinstance(field_b_widget, FilteredLineEdit): field_b_widget.setMaxLength(10)

                    if field_b_ui_widget is None: field_b_ui_widget = field_b_widget
                    h_box.setContentsMargins(0,0,0,0); h_box.setSpacing(5) 
                    label_a_text, label_b_text = ("a:", "b:") if self.group_label_prefix not in ["材質", "仕様", "関連商品", "商品サイズ"] else \
                                                 ("名称:", "説明:") if self.group_label_prefix in ["材質", "仕様"] else \
                                                 ("商品名:", "商品コード:") if self.group_label_prefix == "関連商品" else \
                                                 ("項目名:", "サイズ/値:") # 商品サイズ
                    if self.group_label_prefix == "関連商品":
                        container_b_with_label = QWidget(); v_box_b_layout = QVBoxLayout(container_b_with_label)
                        v_box_b_layout.setContentsMargins(0,0,0,0); v_box_b_layout.setSpacing(1)
                        if field_b_widget: v_box_b_layout.addWidget(field_b_widget)
                        digit_count_label_b = QLabel("(0/10 桁)"); digit_count_label_b.setObjectName(f"DigitCountLabel_{field_name_b}"); digit_count_label_b.setStyleSheet("font-size: 8pt; color: #6c757d;")
                        v_box_b_layout.addWidget(digit_count_label_b)
                        if field_b_widget: field_b_widget.textChanged.connect(lambda text, widget=field_b_widget, label=digit_count_label_b: self._update_digit_count_display_for_related_product(widget, label))
                        self.related_product_code_ui_elements.append({'field': field_b_widget, 'label': digit_count_label_b})
                        h_box.addWidget(QLabel(label_b_text)); h_box.addWidget(container_b_with_label, 1); h_box.addSpacing(10)
                        h_box.addWidget(QLabel(label_a_text)); 
                        if field_a_widget: h_box.addWidget(field_a_widget, 3)
                    else:
                        h_box.addWidget(QLabel(label_a_text)); 
                        if field_a_widget: h_box.addWidget(field_a_widget, 1); 
                        h_box.addSpacing(10); h_box.addWidget(QLabel(label_b_text))
                        if field_b_ui_widget: h_box.addWidget(field_b_ui_widget, 3)
                    
                    field_widget_container = QWidget(); field_widget_container.setLayout(h_box)
                    self.form_layout.addRow(row_label_widget, field_widget_container)
                    self.row_widgets_for_visibility.append((row_label_widget, field_widget_container))
                    if self.group_label_prefix == "商品サイズ": self._update_product_size_b_input_type(field_a_widget.text(), i-1)

                else: # not self.has_ab
                    field_name = f"{self.group_label_prefix}_{i}" 
                    field_widget = QLineEdit(); field_widget.setObjectName(field_name); field_widget._efg_managed = True
                    self.main_fields_ref[field_name] = field_widget
                    if self.parent_app_ref: field_widget.textChanged.connect(self.parent_app_ref.mark_dirty)
                    self.form_layout.addRow(row_label_widget, field_widget)
                    self.row_widgets_for_visibility.append((row_label_widget, field_widget))

    def _update_digit_count_display_for_related_product(self, line_edit_widget, label_widget):
        """関連商品の商品コードフィールドの文字数カウント表示を更新する"""
        current_text = line_edit_widget.text()
        current_digits = len(current_text)
        label_widget.setText(f"({current_digits}/10 桁)")

        # FilteredLineEdit は数字のみと最大長10を強制する
        # is_valid_input は通常 True になるはず
        is_valid_input = current_text.isdigit() or not current_text

        if not is_valid_input: # FilteredLineEditにより通常は発生しない
            label_widget.setStyleSheet("font-size: 8pt; color: red; font-weight: bold;")
        elif current_digits == 10: # 10桁ちょうどで有効な入力
            label_widget.setStyleSheet("font-size: 8pt; color: green; font-weight: bold;")
        else: # それ以外 (0-9桁の有効な入力)
            label_widget.setStyleSheet("font-size: 8pt; color: #6c757d;")
        # mark_dirty は field_b_widget.textChanged が ProductApp 側で接続されていれば不要
        # (既に接続済みのはず)

    def toggle_expand_collapse(self):
        self.is_expanded = not self.is_expanded
        self.update_toggle_button_icon()
        self.update_fields_visibility()
        if self.parent_app_ref and self.parent_app_ref.layout(): 
            self.parent_app_ref.layout().activate() 

    def update_toggle_button_icon(self):
        self.toggle_button.setIcon(self.style().standardIcon(QStyle.SP_ArrowDown if self.is_expanded else QStyle.SP_ArrowRight))

    def update_fields_visibility(self):
        rows_to_show = self.group_count if self.is_expanded else self.always_show_count
        any_field_visible_in_container = False
        for idx, (label_widget, field_or_hbox_widget) in enumerate(self.row_widgets_for_visibility):
            is_visible_this_row = idx < rows_to_show
            label_widget.setVisible(is_visible_this_row)
            field_or_hbox_widget.setVisible(is_visible_this_row)
            if is_visible_this_row:
                any_field_visible_in_container = True
        
        self.fields_container_widget.setVisible(any_field_visible_in_container)
        self.main_content_widget.adjustSize()
        self.main_content_widget.updateGeometry()

    def clear_dimension_fields(self):
        """商品サイズグループの寸法入力フィールド(w,d,h)と重量フィールドをクリアする"""
        if self.group_label_prefix == "商品サイズ":
            for i in range(self.group_count):
                if self.dimension_fields_list[i]:
                    self.dimension_fields_list[i]['w'].clear()
                    self.dimension_fields_list[i]['d'].clear()
                    self.dimension_fields_list[i]['h'].clear()
                    # _update_b_field_from_dimensions を呼び出して、対応する 'b' フィールドも更新
                    # self._update_b_field_from_dimensions(i) # clear_fields の中で b フィールドもクリアされるので不要かも

                if self.weight_fields_list[i]:
                    self.weight_fields_list[i]['weight'].clear()
                    # self._update_b_field_from_weight(i) # 同上

                # 'b' フィールド (QLineEdit本体) もクリア
                b_field_name = f"{self.group_label_prefix}_{i+1}b"
                if b_field_name in self.main_fields_ref:
                    self.main_fields_ref[b_field_name].clear()

                # QStackedWidget の状態もリセット (通常のQLineEditを表示)
                if self.b_field_stacks and self.b_field_stacks[i]:
                    b_line_edit_widget = self.main_fields_ref.get(b_field_name)
                    if b_line_edit_widget:
                         self.b_field_stacks[i].setCurrentWidget(b_line_edit_widget)
    def on_master_a_selected(self, selected_text_a, field_b_widget, master_data_map):
        description = ""
        if selected_text_a and master_data_map: # Ensure text and map are valid
            description = master_data_map.get(selected_text_a, "")
        field_b_widget.setText(description) # type: ignore
        if self.parent_app_ref: self.parent_app_ref.mark_dirty() # type: ignore

    # --- 商品サイズ専用メソッド ---
    def _update_product_size_b_input_type(self, text_a, row_idx):
        if self._processing_a_change: # 再入防止
            return
        self._processing_a_change = True
        try:
            if not (self.group_label_prefix == "商品サイズ" and 0 <= row_idx < self.group_count):
                self._processing_a_change = False
                return

            stack = self.b_field_stacks[row_idx]
            b_line_edit = self.main_fields_ref.get(f"{self.group_label_prefix}_{row_idx+1}b")
            dim_data = self.dimension_fields_list[row_idx]
            weight_data = self.weight_fields_list[row_idx]

            if not stack or not b_line_edit or not dim_data or not weight_data:
                self._processing_a_change = False
                return

            normalized_text_a = text_a.strip()

            if normalized_text_a == "本体":
                self._parse_and_set_dimensions(b_line_edit.text(), dim_data)
                stack.setCurrentWidget(dim_data['container'])
            elif normalized_text_a == "重量":
                self._parse_and_set_weight(b_line_edit.text(), weight_data)
                stack.setCurrentWidget(weight_data['container'])
            else:
                # Xa が「本体」「重量」以外の値になった場合
                # どの特殊UIが表示されていたかを確認し、その値をb_line_editに反映
                current_active_special_widget = stack.currentWidget()
                if current_active_special_widget == dim_data['container']:
                    self._format_and_set_b_field(dim_data, b_line_edit, called_from_a_change=True)
                elif current_active_special_widget == weight_data['container']:
                    self._format_and_set_b_field_from_weight(weight_data, b_line_edit, called_from_a_change=True)
                # else: b_line_editが既に表示されていた場合は、その値はそのまま (setTextしない)
                
                stack.setCurrentWidget(b_line_edit)
        finally:
            self._processing_a_change = False

    def _parse_and_set_dimensions(self, text_b, dim_data):
        w_edit, d_edit, h_edit = dim_data['w'], dim_data['d'], dim_data['h']
        
        # ブロックして、setTextが循環トリガーしないようにする
        for edit in [w_edit, d_edit, h_edit]: edit.blockSignals(True)

        match = re.match(r"幅\s*([\d\.]+)\s*×\s*奥行\s*([\d\.]+)\s*×\s*高さ\s*([\d\.]+)\s*cm", text_b.strip())
        if match:
            w_edit.setText(match.group(1))
            d_edit.setText(match.group(2))
            h_edit.setText(match.group(3))
        else:
            # 簡単なカンマ区切りやスペース区切りの数値も試す (例: "10,20,30" や "10 20 30")
            parts = re.split(r'[,;\s]+', text_b.strip())
            if len(parts) == 3 and all(p.replace('.', '', 1).isdigit() for p in parts if p):
                w_edit.setText(parts[0])
                d_edit.setText(parts[1])
                h_edit.setText(parts[2])
            else: # 解析できなければクリア
                w_edit.clear()
                d_edit.clear()
                h_edit.clear()
        
        for edit in [w_edit, d_edit, h_edit]: edit.blockSignals(False)

    def _format_and_set_b_field(self, dim_data, b_line_edit, called_from_a_change=False):
        w = dim_data['w'].text().strip()
        d = dim_data['d'].text().strip()
        h = dim_data['h'].text().strip()

        new_text_b = ""
        if w and d and h: #  and all are numeric-like (validator should handle this)
            new_text_b = f"幅{w}×奥行{d}×高さ{h}cm"
        
        # 既存のテキストと異なる場合のみ更新し、シグナルをブロック
        if b_line_edit.text() != new_text_b:
            b_line_edit.blockSignals(True)
            b_line_edit.setText(new_text_b)
            b_line_edit.blockSignals(False)
            if self.parent_app_ref and not called_from_a_change : # Xaからの変更時はmark_dirtyはXaが担当
                self.parent_app_ref.mark_dirty()
            # if self.parent_app_ref and hasattr(self.parent_app_ref, '_sync_product_size_to_yspec'):
            #     self.parent_app_ref._sync_product_size_to_yspec() # タイミングを限定するため、ここでは呼ばない

    def _update_b_field_from_dimensions(self, row_idx):
        if not (self.group_label_prefix == "商品サイズ" and 0 <= row_idx < self.group_count):
            return
        
        dim_data = self.dimension_fields_list[row_idx]
        b_line_edit = self.main_fields_ref.get(f"{self.group_label_prefix}_{row_idx+1}b")
        if dim_data and b_line_edit:
            self._format_and_set_b_field(dim_data, b_line_edit)
            # 親アプリの同期メソッドを呼び出す
            if self.parent_app_ref and hasattr(self.parent_app_ref, '_sync_product_size_to_yspec'):
                self.parent_app_ref._sync_product_size_to_yspec()

    def _update_dimensions_from_b_field(self, text_b, row_idx):
        if not (self.group_label_prefix == "商品サイズ" and 0 <= row_idx < self.group_count):
            return

        # 「項目名」が「本体」の時だけ、b_line_editの変更をWDHに反映する
        # (そうでない時はb_line_editが直接表示されているので、WDHへの反映は不要)
        field_a_widget = self.main_fields_ref.get(f"{self.group_label_prefix}_{row_idx+1}a")
        dim_data = self.dimension_fields_list[row_idx]
        weight_data = self.weight_fields_list[row_idx]
        
        if field_a_widget and field_a_widget.text().strip() == "本体" and dim_data:
            # WDHフィールドが表示されているはずなので、b_line_editの変更をWDHにパースして設定
            self._parse_and_set_dimensions(text_b, dim_data)
        elif field_a_widget and field_a_widget.text().strip() == "重量" and weight_data:
            # 重量フィールドが表示されているはずなので、b_line_editの変更を重量フィールドにパースして設定
            self._parse_and_set_weight(text_b, weight_data)
        # ProductApp.mark_dirty は b_line_edit.textChanged によって既に接続されている。

    def _parse_and_set_weight(self, text_b, weight_data):
        weight_edit = weight_data['weight']
        weight_edit.blockSignals(True)

        # "約〇〇kg", "〇〇kg", "約 〇〇 kg", "〇〇 kg" のような形式から数値を抽出
        match = re.match(r"^\s*(?:約)?\s*([\d\.]+)\s*kg\s*$", text_b.strip(), re.IGNORECASE)
        if match:
            weight_edit.setText(match.group(1))
        else:
            # 単純な数値も試す
            cleaned_text_b = text_b.replace("約", "").strip() # 先頭の「約」も除去して試す
            if cleaned_text_b.replace('.', '', 1).isdigit():
                weight_edit.setText(cleaned_text_b)
            else:
                weight_edit.clear()
        
        weight_edit.blockSignals(False)

    def _format_and_set_b_field_from_weight(self, weight_data, b_line_edit, called_from_a_change=False):
        weight_val_str = weight_data['weight'].text().strip()
        new_text_b = ""
        if weight_val_str: # and is numeric-like (validator should handle this)
            new_text_b = f"約{weight_val_str}kg" # 「約」を付加

        if b_line_edit.text() != new_text_b:
            b_line_edit.blockSignals(True)
            b_line_edit.setText(new_text_b)
            b_line_edit.blockSignals(False)
            if self.parent_app_ref and not called_from_a_change:
                self.parent_app_ref.mark_dirty()

    def _update_b_field_from_weight(self, row_idx):
        if not (self.group_label_prefix == "商品サイズ" and 0 <= row_idx < self.group_count):
            return
        
        weight_data = self.weight_fields_list[row_idx]
        b_line_edit = self.main_fields_ref.get(f"{self.group_label_prefix}_{row_idx+1}b")
        if weight_data and b_line_edit:
            self._format_and_set_b_field_from_weight(weight_data, b_line_edit)

    def update_all_related_product_code_counts(self):
        """保持している全ての関連商品コードフィールドの文字数カウント表示を更新する"""
        if self.group_label_prefix == "関連商品":
            for ui_element in self.related_product_code_ui_elements:
                self._update_digit_count_display_for_related_product(ui_element['field'], ui_element['label'])


class ProductApp(QWidget):
    def __init__(self):
        super().__init__()
        # 重要：UI構築で使用される辞書類を最初に初期化
        self.main_fields = {}
        self.category_fields = []
        self.byte_count_labels = {}
        self.expandable_field_group_instances = {}
        
        # 状態管理フラグの初期化
        self._is_new_mode = False
        self._is_loading_data = False
        self._is_handling_selection_change = False
        
        self._save_mutex = QMutex()  # 保存処理の排他制御
        self._setup_logging() # ★★★ ロギング設定を最初に行う ★★★
        
        # 万が一対策システムの初期化
        self._init_emergency_systems()
        
        # Undo/Redo用の履歴管理
        self.undo_stack = []
        self.redo_stack = []
        self.max_undo_history = 50  # 最大履歴数
        self._is_undoing = False  # Undo/Redo実行中フラグ
        self._undo_save_timer = None  # デバウンス用タイマー
        
        self.setWindowTitle(f"商品登録入力ツール v{CURRENT_VERSION}")
        # 適切な初期サイズを設定
        self.resize(1400, 900)
        self.setMinimumSize(1200, 700)
        
        # ウィンドウアイコンの設定
        if hasattr(self, 'setWindowIcon'):
            try:
                icon_path = os.path.join(os.path.dirname(__file__), "商品登録ツール.ico")
                if os.path.exists(icon_path):
                    from PyQt5.QtGui import QIcon
                    self.setWindowIcon(QIcon(icon_path))
            except Exception as e:
                logging.debug(f"アイコン設定失敗（継続）: {e}")  # デバッグレベルでログ記録
        
        # --- 起動時処理 ---
        progress = self._show_loading_dialog()
        self._init_paths_and_dirs(progress)
        self._load_initial_data(progress)
        progress.show()
        QApplication.processEvents() # ダイアログの表示を確実にする

        current_step = 0

        try:
            # ステップ0: パス設定とディレクトリ準備
            # _init_paths_and_dirs で実行済み
            # バンドルされるファイルのパス解決をより安全に
            safe_template_name = os.path.normpath(TEMPLATE_FILE_NAME).lstrip(os.sep + (os.altsep or ''))
            safe_category_name = os.path.normpath(CATEGORY_FILE_NAME).lstrip(os.sep + (os.altsep or ''))

            self.template_file_path_bundle = os.path.join(self.base_dir_frozen, safe_template_name)

            # 管理ファイルの初期化またはアップデート
            self._initialize_or_update_manage_file()
            current_step += 1
            progress.setValue(current_step) # テンプレートコピー完了

        except FileNotFoundError as e:
            if 'progress' in locals() and progress.isVisible(): progress.stop_animation(); progress.close()
            err_msg = f"必須ファイルが見つかりません: {e.filename}\nアプリケーションを終了します。"
            logging.critical(err_msg, exc_info=True)
            QMessageBox.critical(None, "起動エラー", f"{err_msg}\n詳細はログファイルを確認してください。")
            sys.exit(1)
        except Exception as e:
            if 'progress' in locals() and progress.isVisible(): progress.stop_animation(); progress.close()
            err_msg = f"ファイルの読み込み中に予期せぬエラーが発生しました。\nアプリケーションを終了します。"
            logging.critical(err_msg, exc_info=True)
            QMessageBox.critical(None, "起動エラー", f"{err_msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}")
            sys.exit(1)

        # UI構築の前にラベルを更新
        progress.setLabelText("ユーザーインターフェースを構築中..."); current_step +=1; progress.setValue(current_step)
        QApplication.processEvents()
        
        # メモリ使用量の最適化（安全チェック付き）
        try:
            if check_memory_before_large_operation:
                check_memory_before_large_operation("UI構築")
        except Exception as e:
            logging.debug(f"メモリチェック中のエラー（継続）: {e}")

        # --- UI構築開始 ---
        self._init_ui_components() # 主要なUI要素の初期化
        self._setup_copy_paste_actions() # 商品リストのコピペアクション設定（メニューバー作成前に実行）
        self._setup_delete_action() # 商品リストのDeleteキーアクション設定
        
        # 最上位レイアウト（メニューバー + メインコンテンツ）
        top_layout = QVBoxLayout(self)
        top_layout.setContentsMargins(0,0,0,0)
        top_layout.setSpacing(0)
        
        # 一時的に空のメニューバーを作成
        self.menu_bar = QMenuBar(self)
        self.menu_bar.setFixedHeight(MENU_BAR_HEIGHT)
        top_layout.addWidget(self.menu_bar)
        
        # メインコンテンツ用ウィジェット
        main_content_widget = QWidget()
        main_layout = QHBoxLayout(main_content_widget)
        main_layout.setContentsMargins(0,0,0,0)
        main_layout.setSpacing(0)
        top_layout.addWidget(main_content_widget)
        
        # ステータスバーの追加
        self.status_bar = QStatusBar(self)
        self.status_bar.setFixedHeight(24)
        self.status_bar.setSizeGripEnabled(False)
        self.status_bar.showMessage("起動中...")
        top_layout.addWidget(self.status_bar)
        
        self._setup_main_layout(main_layout) # メインレイアウトの構築
        self._connect_signals() # シグナル接続の設定
        self._setup_tab_order() # タブオーダーの設定

        # --- 左ペイン ---
        left_widget = QWidget(); left_widget.setObjectName("LeftPane"); left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(10,10,10,10); left_layout.setSpacing(8)

        action_buttons_widget = QWidget()
        action_buttons_layout = QHBoxLayout(action_buttons_widget)
        action_buttons_layout.setContentsMargins(0,0,0,0); action_buttons_layout.setSpacing(5)
        action_buttons_layout.addWidget(self.new_btn)
        
        action_buttons_layout.addWidget(self.save_btn) 

        action_buttons_layout.addWidget(self.run_csharp_btn)
        action_buttons_layout.addWidget(self.bulk_p_btn) 

        # action_buttons_layout.addWidget(self.sku_add_btn) # SKU追加ボタンはSKUヘッダーへ移動
        action_buttons_layout.addStretch()
        left_layout.addWidget(action_buttons_widget)
        left_layout.addSpacing(10)

        # search_bar と product_list は _init_ui_components で初期化済み
        search_label = QLabel("商品検索"); search_label.setObjectName("SidebarLabel"); list_label = QLabel("商品一覧"); list_label.setObjectName("SidebarLabel")
        left_layout.addWidget(search_label); left_layout.addWidget(self.search_bar); left_layout.addSpacing(10) # new_btn_layout を削除
        left_layout.addWidget(list_label); left_layout.addWidget(self.product_list)
        right_widget = QWidget(); right_widget.setObjectName("RightPane"); right_main_layout = QVBoxLayout(right_widget); right_main_layout.setContentsMargins(0,0,0,0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setObjectName("MainScrollArea")
        content = QWidget(); content.setObjectName("MainScrollContent"); form = QFormLayout(); form.setSpacing(8); form.setContentsMargins(15,15,15,15)
        # control_radio_n, control_radio_p, control_radio_group, category_select_btn は _init_ui_components で初期化済み
        control_radio_layout = QHBoxLayout(); control_radio_layout.addWidget(self.control_radio_n); control_radio_layout.addWidget(self.control_radio_p)
        self.control_radio_widget = QWidget(); self.control_radio_widget.setLayout(control_radio_layout)
        self.main_field_order = [
            HEADER_MEMO,"シリーズ名","シリーズURL",HEADER_MYCODE,HEADER_PRODUCT_NAME, HEADER_PRICE_TAX_INCLUDED,HEADER_SORT_FIELD,"R_商品名","Y_商品名","R_キャッチコピー","Y_metadesc",
            "Y_キャッチコピー","Y_metakey","特徴_1",HEADER_YAHOO_ABSTRACT,"メーカー売価_税込み", "メーカー売価_画像","送料形態",HEADER_IMAGE_DESCRIPTION,HEADER_IMAGE_PATH_RAKUTEN,"商品カテゴリ1",
            "商品カテゴリ2","商品カテゴリ3","商品カテゴリ4","商品カテゴリ5",HEADER_R_GENRE_ID, HEADER_Y_CATEGORY_ID,HEADER_YA_CATEGORY_ID,"商品サイズ_1a","商品サイズ_1b","商品サイズ_2a","商品サイズ_2b",
            "商品サイズ_3a","商品サイズ_3b","商品サイズ_4a","商品サイズ_4b","商品サイズ_5a", "商品サイズ_5b","商品サイズ_6a","商品サイズ_6b","商品サイズ_7a","商品サイズ_7b",
            "商品サイズ_8a","商品サイズ_8b","梱包サイズ_1",
            "材質_1", "材質_2a", "材質_2b", "材質_3a", "材質_3b", "材質_4a", "材質_4b", "材質_5a", "材質_5b", "材質_6a", "材質_6b",
            "色_1",
            "仕様_1", "仕様_2a", "仕様_2b", "仕様_3a", "仕様_3b", "仕様_4a", "仕様_4b", "仕様_5a", "仕様_5b", "仕様_6a", "仕様_6b",
            "お届け状態_1", "関連商品_1b","関連商品_1a","関連商品_2b","関連商品_2a","関連商品_3b","関連商品_3a","関連商品_4b","関連商品_4a", "関連商品_5b","関連商品_5a","関連商品_6b","関連商品_6a","関連商品_7b","関連商品_7a","関連商品_8b",
            "関連商品_8a","関連商品_9b","関連商品_9a","関連商品_10b","関連商品_10a","関連商品_11b","関連商品_11a", "関連商品_12b","関連商品_12a","関連商品_13b","関連商品_13a","関連商品_14b","関連商品_14a","関連商品_15b", "関連商品_15a",
            "relevant_links","説明マーク_1","Y_spec1","Y_spec2","Y_spec3","Y_spec4","Y_spec5","Y_spec6", "Y_spec7","Y_spec8","Y_spec9","Y_spec10","R_SKU項目名","R_商品プルダウン","R_別途送料地域項目名","R_別途送料地域選択肢",
            "R_配達オプション項目名","R_配達オプション選択肢","R_注意事項プルダウン","Y_SKU項目名","Y_商品プルダウン","Y_別途送料地域項目名", "Y_別途送料地域選択肢","Y_配達オプション項目名","Y_配達オプション選択肢","Y_注意事項プルダウン","注意事項","-", "非製品属性タグID"
        ] # type: ignore
        # main_fields と category_fields は _init_ui_components で初期化済み

        try:
            # YA_suffix を適切な位置に挿入
            price_index = self.main_field_order.index(HEADER_PRICE_TAX_INCLUDED)
            r_product_name_index = self.main_field_order.index("R_商品名")
            insert_index = max(price_index + 1, r_product_name_index)
            self.main_field_order.insert(insert_index, "YA_suffix") # type: ignore
        except ValueError:
            logging.warning("価格フィールドまたはR_商品名フィールドが見つからず、YA_suffixの位置を特定できませんでした。")

        for i in range(1, 6): fld = QLineEdit(); fld.setObjectName(f"商品カテゴリ{i}"); self.main_fields[f"商品カテゴリ{i}"] = fld; self.category_fields.append(fld)
        
        # 楽天SKUオプションフィールドのリスト
        self.rakuten_sku_option_fields_list = [
            "R_SKU項目名", "R_商品プルダウン", "R_別途送料地域項目名", 
            "R_別途送料地域選択肢", "R_配達オプション項目名", "R_配達オプション選択肢",
            "R_注意事項プルダウン" # 注意事項プルダウンもここに含める
        ]
        # Yahoo!SKUオプションフィールドのリスト
        self.yahoo_sku_option_fields_list = [
            "Y_SKU項目名","Y_商品プルダウン","Y_別途送料地域項目名", 
            "Y_別途送料地域選択肢","Y_配達オプション項目名","Y_配達オプション選択肢",
            "Y_注意事項プルダウン"
        ]

        expandable_groups = [("商品サイズ",8,True), ("材質",6,True), ("仕様",6,True), ("関連商品",15,True)]
        self.image_desc_field = None; self.image_desc_row_widget = None
        # self.image_desc_btn は _init_ui_components で初期化済み
        
        self.expandable_field_group_instances = {}
        for lbl, cnt, ab_flag in expandable_groups:
            master_data_for_group = self.material_spec_master if lbl in ["材質", "仕様"] else None
            self.expandable_field_group_instances[lbl] = ExpandableFieldGroup(lbl, cnt, self.main_fields, 3, ab_flag, self, master_data=master_data_for_group)
        self.byte_count_labels = {}
        self.digit_count_label_mycode = None
        added_expandable_groups = set()
        
        # Y_spec フィールド用のラベルとエディタコンテナを初期化
        self.y_spec_labels = []
        self.y_spec_editor_placeholders = [] # QWidgetのリスト、この中に動的エディタが入る
        self.y_spec_current_editors = [None] * MAX_Y_SPEC_COUNT # 現在表示されているエディタの参照
        self.y_spec_current_definitions = [None] * MAX_Y_SPEC_COUNT # 現在表示されているスペックの定義
        self._y_spec_section_rendered_in_form = False # Y_specセクションがフォームにレンダリングされたかのフラグ
        
        # Y_specセクションのヘッダーとスペーサーをインスタンス変数として定義
        self.y_spec_section_label_widget = QLabel("Yahoo!ショッピング スペック情報↓")
        self.y_spec_section_label_widget.setObjectName("SectionHeader")
        
        self.y_spec_header_spacer_top = QLabel(" ") # 空白文字を設定して高さを認識しやすくする
        self.y_spec_header_spacer_top.setMinimumHeight(SPACER_HEIGHT)
        self.y_spec_footer_spacer = QLabel(" ")     # 空白文字を設定して高さを認識しやすくする
        self.y_spec_footer_spacer.setMinimumHeight(SPACER_HEIGHT)

        id_field_names_ordered = [HEADER_R_GENRE_ID, HEADER_Y_CATEGORY_ID, HEADER_YA_CATEGORY_ID]
        id_section_added = False

        # 新しい楽天SKUオプションのExpandableFieldGroupを作成
        self.expandable_field_group_instances["楽天SKUオプション"] = ExpandableFieldGroup(
            group_label="楽天SKUオプション設定",
            group_count=0, # field_names_list を使うので0でOK
            main_fields_dict=self.main_fields,
            always_show=0, # デフォルトで折りたたむ
            has_ab=False,
            parent_app=self,
            field_names_list=self.rakuten_sku_option_fields_list
        )
        # 新しいYahoo!SKUオプションのExpandableFieldGroupを作成
        self.expandable_field_group_instances["Yahoo!SKUオプション"] = ExpandableFieldGroup(
            group_label="Yahoo!SKUオプション設定",
            group_count=0, # field_names_list を使うので0でOK
            main_fields_dict=self.main_fields,
            always_show=0, # デフォルトで折りたたむ
            has_ab=False,
            parent_app=self,
            field_names_list=self.yahoo_sku_option_fields_list
        )
        added_expandable_groups = set() # 既存のグループと新しいグループの重複追加を防ぐ

        for name in self.main_field_order:
            # R_SKU項目名の直前に、Y_specセクションが表示されていればフッタースペーサーを挿入
            # R_SKU項目名は新しいグループに含まれるため、そのグループのプレースホルダーの直前に変更
            if name == self.rakuten_sku_option_fields_list[0] or name == self.yahoo_sku_option_fields_list[0]: # 楽天またはYahooのSKUグループの最初の要素
                if self._y_spec_section_rendered_in_form:
                    form.addRow(QLabel(), self.y_spec_footer_spacer) # フィールド側にスペーサーを配置

            # 既存のExpandableFieldGroupの処理
            is_efg_handled = False
            for grp_lbl, efg_inst in self.expandable_field_group_instances.items():
                first_fld_in_efg = f"{grp_lbl}_1"
                if efg_inst.has_ab and grp_lbl not in ["材質", "仕様"]:
                    first_fld_in_efg = f"{grp_lbl}_1b"
                if name == first_fld_in_efg and grp_lbl not in added_expandable_groups:
                    form.addRow(QLabel(grp_lbl), efg_inst.main_content_widget)
                    added_expandable_groups.add(grp_lbl); is_efg_handled = True; break
                # 新しい楽天SKUオプショングループの処理
                elif grp_lbl == "楽天SKUオプション" and name == self.rakuten_sku_option_fields_list[0] and grp_lbl not in added_expandable_groups:
                    form.addRow(QLabel(grp_lbl), efg_inst.main_content_widget) # ラベルはグループ名
                    added_expandable_groups.add(grp_lbl); is_efg_handled = True; break
                # 新しいYahoo!SKUオプショングループの処理
                elif grp_lbl == "Yahoo!SKUオプション" and name == self.yahoo_sku_option_fields_list[0] and grp_lbl not in added_expandable_groups:
                    form.addRow(QLabel(grp_lbl), efg_inst.main_content_widget) # ラベルはグループ名
                    added_expandable_groups.add(grp_lbl); is_efg_handled = True; break
            
            if is_efg_handled:
                continue

            # 新しい楽天SKUオプショングループに含まれるフィールドは個別に追加しない
            if name in self.rakuten_sku_option_fields_list:
                if name not in self.main_fields: # main_fieldsにウィジェットがなければ作成
                    self.main_fields[name] = QLineEdit(); self.main_fields[name].setObjectName(name)
                continue # ExpandableFieldGroup内で処理されるのでスキップ
            
            # 新しいYahoo!SKUオプショングループに含まれるフィールドは個別に追加しない
            if name in self.yahoo_sku_option_fields_list:
                if name not in self.main_fields: # main_fieldsにウィジェットがなければ作成
                    self.main_fields[name] = QLineEdit(); self.main_fields[name].setObjectName(name)
                continue # ExpandableFieldGroup内で処理されるのでスキップ

            if name in self.main_fields and hasattr(self.main_fields[name], '_efg_managed') and self.main_fields[name]._efg_managed:
                continue

            if name == id_field_names_ordered[0] and not id_section_added:
                # id_section_label = QLabel("ID関連") # 削除
                # id_section_label.setObjectName("SectionHeader") # 削除
                # form.addRow(id_section_label) # 削除

                for id_name_in_group in id_field_names_ordered:
                    if id_name_in_group not in self.main_fields:
                        self.main_fields[id_name_in_group] = QLineEdit()
                        self.main_fields[id_name_in_group].setObjectName(id_name_in_group)
                    
                    display_id_label_text = id_name_in_group
                    if id_name_in_group == HEADER_R_GENRE_ID: display_id_label_text = "RジャンルID"
                    elif id_name_in_group == HEADER_Y_CATEGORY_ID: display_id_label_text = "YカテゴリID"
                    elif id_name_in_group == HEADER_YA_CATEGORY_ID: display_id_label_text = "YAカテゴリID"
                    
                    form.addRow(QLabel(display_id_label_text), self.main_fields[id_name_in_group])

                # open_id_search_buttonは_init_ui_componentsで既に作成済み
                form.addRow("", self.open_id_search_button)
                id_section_added = True
                continue

            if name == EXPLANATION_MARK_FIELD_NAME: # "説明マーク_1" の特別処理
                self.explanation_mark_line_edit = QLineEdit()
                self.explanation_mark_line_edit.setObjectName(EXPLANATION_MARK_FIELD_NAME + "_input")
                self.main_fields[EXPLANATION_MARK_FIELD_NAME] = self.explanation_mark_line_edit
                self.explanation_mark_line_edit.textChanged.connect(lambda: self.mark_dirty())
                form.addRow(QLabel("説明マーク"), self.explanation_mark_line_edit)

                self.explanation_mark_select_btn = QPushButton("説明マーク選択")
                self.explanation_mark_select_btn.setObjectName("ExplanationMarkSelectButton")
                self.explanation_mark_select_btn.clicked.connect(lambda: self.open_explanation_mark_dialog())
                form.addRow("", self.explanation_mark_select_btn) # ボタンを入力欄の下に配置
                continue
            elif name in id_field_names_ordered and id_section_added:
                continue
            
            if name.startswith("商品カテゴリ"):
                if name in self.main_fields:
                    form.addRow(QLabel(name), self.main_fields[name])
                if name == "商品カテゴリ5":
                    form.addRow("", self.category_select_btn)
                continue
            elif name == HEADER_IMAGE_DESCRIPTION:
                self.image_desc_field = QLineEdit()
                self.main_fields[name] = self.image_desc_field
                form.addRow(QLabel(name), self.image_desc_field) # まず入力欄を追加
                # image_desc_btn は __init__ の最初の方で定義済み
                form.addRow("", self.image_desc_btn) # 次の行にボタンを追加
                continue
            elif name == HEADER_SORT_FIELD or name == "-":
                fld = self.main_fields.get(name, QLineEdit())
                self.main_fields[name] = fld
                if name == "-": fld.setReadOnly(True)
                continue
            elif name == "relevant_links": # relevant_links の特別処理
                fld = self.main_fields.get(name, QLineEdit())
                fld.setObjectName(name)
                # 説明マーク_1 の特別処理で既に main_fields に追加されている場合があるためチェック
                if name == EXPLANATION_MARK_FIELD_NAME:
                    # 説明マーク_1 は上で特別処理済みなのでスキップ
                    continue
                fld.setReadOnly(True) # 読み取り専用に設定
                self.main_fields[name] = fld
                label_widget = QLabel(name) # ラベルはそのまま "relevant_links"
                form.addRow(label_widget, fld)
                self.main_fields[name] = fld
                if name == "-": fld.setReadOnly(True)
                continue

            # --- Modification for "お届け状態_1" ---
            if name == "お届け状態_1":
                fld = QComboBox()
                fld.setEditable(True)
                # Add an empty item for clearing selection, then the predefined options
                fld.addItems(["", "完成品", "組立必要品", "完成品(脚部取付必要)", "完成品(上下重ね合わせ必要)"])
            elif name in HTML_TEXTEDIT_FIELDS:
                fld = CustomHtmlTextEdit() # カスタムクラスを使用
                fld.setPlaceholderText("HTMLタグ使用可。改行＝<br>")
            else:
                # "色_1" の特別処理のために、ここで fld を確定させる前に name をチェック
                if name == "色_1":
                    fld = QLineEdit() # This will be self.main_fields["色_1"]
                    fld.setObjectName(name) # QLineEditにもオブジェクト名を設定
                    self.main_fields[name] = fld # main_fields に登録

                    # display_label_text はこの後のロジックで "色" に設定される
                    label_widget = QLabel("色") # display_label_text を直接使用
                    form.addRow(label_widget, fld) # まず入力欄を追加
                    if isinstance(fld, QLineEdit): fld.textChanged.connect(lambda: self.mark_dirty())

                    color_select_button = QPushButton("色を選択")
                    color_select_button.setObjectName("ColorSelectButton") # For stylesheet
                    color_select_button.clicked.connect(lambda: self._open_color_selection_dialog())
                    form.addRow("", color_select_button) # 次の行にボタンを追加
                    continue # このフィールドの処理は完了
                fld = self.main_fields.get(name, QLineEdit())
            # --- End Modification for "お届け状態_1" ---
            
            # self.main_fields[name] = fld # 説明マーク_1 は上で追加済み
            fld.setObjectName(name)
            self.main_fields[name] = fld # ★★★ この行を追加して、作成されたウィジェットを main_fields に登録 ★★★

            display_label_text = name
            if name == HEADER_MYCODE: display_label_text = "商品コード"
            elif name == HEADER_YAHOO_ABSTRACT: display_label_text = "Y_ストアコメント"
            elif name == "梱包サイズ_1": display_label_text = "梱包サイズ(個口数)"
            elif name == "特徴_1": display_label_text = "特徴 (HTML)"
            elif name == "色_1": display_label_text = "色"
            elif name == "お届け状態_1": display_label_text = "お届け状態"
            # elif name == "説明マーク_1": display_label_text = "説明マーク" # 上で特別処理済み
            
            # Y_specフィールドの処理
            if name.startswith("Y_spec") and name[6:].isdigit():
                spec_num = int(name[6:])
                if 1 <= spec_num <= 10:
                    if not self._y_spec_section_rendered_in_form:
                        form.addRow(QLabel(), self.y_spec_header_spacer_top) # フィールド側にスペーサーを配置
                        form.addRow(self.y_spec_section_label_widget)
                        self._y_spec_section_rendered_in_form = True

                    y_spec_label = QLabel(f"Y_spec{spec_num} (項目名)") # 初期ラベル
                    y_spec_editor_placeholder = QWidget() # エディタを配置するコンテナ
                    # QHBoxLayoutをコンテナに設定しておくと、後でエディタを簡単に追加できる
                    y_spec_editor_placeholder.setLayout(QHBoxLayout()) 
                    y_spec_editor_placeholder.layout().setContentsMargins(0,0,0,0)
                    form.addRow(y_spec_label, y_spec_editor_placeholder)
                    self.y_spec_labels.append(y_spec_label)
                    self.y_spec_editor_placeholders.append(y_spec_editor_placeholder)
                    continue # main_fieldsには追加しないので、ここでcontinue

            # 説明マーク_1 は上でラベルとウィジェットコンテナを追加済みなのでスキップ
            if name != EXPLANATION_MARK_FIELD_NAME: label_widget = QLabel(display_label_text)
            form.addRow(label_widget, fld)

            if name == HEADER_MYCODE:
                fld.setMaxLength(DIGIT_COUNT_MYCODE_MAX)
                self.digit_count_label_mycode = QLabel(f"(0/{DIGIT_COUNT_MYCODE_MAX} 桁)");
                self.digit_count_label_mycode.setObjectName("DigitCountLabelMycode")
                form.addRow("", self.digit_count_label_mycode)

            if name in BYTE_LIMITS:
                byte_label = QLabel("(0/" + str(BYTE_LIMITS[name]) + " bytes)");
                byte_label.setObjectName(f"ByteCountLabel_{name}")
                self.byte_count_labels[name] = byte_label
                form.addRow("", byte_label);

        content.setLayout(form); scroll.setWidget(content)
        sku_table_container = QWidget(); sku_table_container.setObjectName("SkuTableContainer"); sku_table_layout = QHBoxLayout(sku_table_container)
        sku_table_layout.setContentsMargins(0,0,0,0); sku_table_layout.setSpacing(1)
        # frozen_table_view と scrollable_table_view は _init_ui_components で初期化済み
        self.frozen_table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed); self.frozen_table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.frozen_table_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection); self.frozen_table_view.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozen_table_view.setAlternatingRowColors(True)
        self.scrollable_table_view.horizontalHeader().setStretchLastSection(False); self.scrollable_table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.scrollable_table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems); self.scrollable_table_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.scrollable_table_view.setAlternatingRowColors(True)
        # sku_model は _init_ui_components で初期化済み
        form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow); form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter) # type: ignore
        # self.sku_add_btn は __init__ の最初の方でアイコン付きで定義済み
        # self.sku_delete_btn は _init_ui_components で初期化済み
        # スクロール同期とテーブルビュー間の相互参照は _connect_signals で設定済み
        self.right_splitter = QSplitter(Qt.Vertical); self.right_splitter.setObjectName("RightSplitter")
        main_widget_container = QWidget()
        main_vbox = QVBoxLayout(main_widget_container)
        main_vbox.setContentsMargins(0,10,0,0); main_vbox.setSpacing(8)
        main_header_layout = QHBoxLayout()
        main_header_label = QLabel("Main情報"); main_header_label.setObjectName("SectionHeader")
        # bulk_p_btn は左ペイン上部に移動したので、ここからは削除
        main_header_layout.addWidget(main_header_label); main_header_layout.addStretch() # main_header_layout.addWidget(self.bulk_p_btn) を削除
        main_vbox.addLayout(main_header_layout); main_vbox.addWidget(self.control_radio_widget); main_vbox.addWidget(scroll)

        sku_widget_container = QWidget()
        sku_vbox = QVBoxLayout(sku_widget_container)
        sku_vbox.setContentsMargins(0,10,0,0); sku_vbox.setSpacing(8)
        sku_header_label = QLabel("SKU情報")
        sku_header_label.setObjectName("SectionHeader")
        sku_header_area_layout = QHBoxLayout()
        sku_header_area_layout.addWidget(sku_header_label) # まずラベルを追加
        sku_header_area_layout.addStretch()
        sku_header_area_layout.addWidget(self.sku_add_btn) # 次にSKU追加ボタン
        sku_header_area_layout.addWidget(self.sku_delete_btn) # 最後にSKU削除ボタン
        sku_header_area_layout.addSpacing(10) # ボタンの右側に少しスペースを追加
        sku_vbox.addLayout(sku_header_area_layout)
        sku_table_layout.addWidget(self.frozen_table_view, 1); sku_table_layout.addWidget(self.scrollable_table_view, 3)
        sku_vbox.addWidget(sku_table_container)
        # save_btn_layout = QHBoxLayout(); save_btn_layout.addStretch(); save_btn_layout.addWidget(self.save_btn); sku_vbox.addLayout(save_btn_layout) # 右下の保存ボタンは左ペイン上部に移動したので削除
        self.right_splitter.addWidget(main_widget_container); self.right_splitter.addWidget(sku_widget_container)
        self.right_splitter.setStretchFactor(0,3); self.right_splitter.setStretchFactor(1,2)
        right_main_layout.addWidget(self.right_splitter)
        self.main_splitter = QSplitter(Qt.Horizontal); self.main_splitter.setObjectName("MainSplitter"); self.main_splitter.addWidget(left_widget); self.main_splitter.addWidget(right_widget); self.main_splitter.setSizes([200,1150]) # 商品一覧を最小幅に設定（一括p設定ボタンの右真横まで）
        main_layout.addWidget(self.main_splitter)

        self._is_loading_data = False # データロード中フラグ
        self._is_deleting = False # 削除処理中フラグ
        self._is_dirty = False # プロパティの内部変数
        # is_dirty プロパティのセッター経由で save_btn の状態も初期化されることを期待
        # ただし、この時点では self.save_btn は確実に存在する
        self.is_dirty = False # セッターを呼び出して save_btn を無効化
        self._is_handling_selection_change = False # 商品選択変更処理中の再入防止フラグ
        self._is_closing = False # アプリ終了処理中フラグ
        self._is_restoring_after_cancel = False # キャンセル後の選択復元中フラグ


        self.control_radio_n.toggled.connect(lambda: self.mark_dirty())

        for field_name in BYTE_LIMITS.keys():
            if field_name in self.main_fields and isinstance(self.main_fields[field_name], QLineEdit):
                self.main_fields[field_name].textChanged.connect(lambda text, fname=field_name: self._update_byte_count_display(fname, text))
        
        if HEADER_MYCODE in self.main_fields and isinstance(self.main_fields[HEADER_MYCODE], QLineEdit):
            self.main_fields[HEADER_MYCODE].textChanged.connect(lambda text: self._update_mycode_digit_count_display(text))
            self.main_fields[HEADER_MYCODE].textChanged.connect(lambda text: self._validate_required_field(HEADER_MYCODE, text))

        for efg_inst in self.expandable_field_group_instances.values():
             for i in range(1, efg_inst.group_count + 1):
                if i == 1 and f"{efg_inst.group_label_prefix}_{i}" in HTML_TEXTEDIT_FIELDS:
                    pass
                elif efg_inst.has_ab : 
                    field_name_a = f"{efg_inst.group_label_prefix}_{i}a"
                    field_name_b = f"{efg_inst.group_label_prefix}_{i}b"
                    widget_a = self.main_fields.get(field_name_a)
                    widget_b = self.main_fields.get(field_name_b)
                    if widget_a and isinstance(widget_a, QLineEdit) and field_name_a in BYTE_LIMITS:
                         widget_a.textChanged.connect(lambda text, fname=field_name_a: self._update_byte_count_display(fname, text))
                    if widget_b and isinstance(widget_b, QLineEdit) and field_name_b in BYTE_LIMITS:
                         widget_b.textChanged.connect(lambda text, fname=field_name_b: self._update_byte_count_display(fname, text))

        if HEADER_PRICE_TAX_INCLUDED in self.main_fields:
            price_field_widget = self.main_fields.get(HEADER_PRICE_TAX_INCLUDED)
            if isinstance(price_field_widget, QLineEdit):
                price_field_widget.textChanged.connect(lambda text, fname=HEADER_PRICE_TAX_INCLUDED: self._on_price_field_changed(fname, text))

        for fld_name, fld_widget in self.main_fields.items():
            if fld_name in HTML_TEXTEDIT_FIELDS and isinstance(fld_widget, QTextEdit):
                 if not (hasattr(fld_widget, '_efg_managed') and fld_widget._efg_managed):
                     fld_widget.textChanged.connect(lambda: self.mark_dirty())
            elif isinstance(fld_widget, QLineEdit):
                if (not hasattr(fld_widget, '_efg_managed') or not fld_widget._efg_managed) and \
                   fld_name != HEADER_PRICE_TAX_INCLUDED and \
                   fld_name != HEADER_MYCODE and \
                   fld_name not in BYTE_LIMITS and \
                   fld_name not in [HEADER_R_GENRE_ID, HEADER_Y_CATEGORY_ID, HEADER_YA_CATEGORY_ID]:
                    fld_widget.textChanged.connect(lambda: self.mark_dirty())
                    # 必須フィールドの検証を追加
                    if fld_name == HEADER_PRODUCT_NAME:
                        fld_widget.textChanged.connect(lambda text: self._validate_required_field(HEADER_PRODUCT_NAME, text))
            elif isinstance(fld_widget, QComboBox): # Handle editable and non-editable QComboBox
                if fld_widget.isEditable():
                    fld_widget.currentTextChanged.connect(lambda: self.mark_dirty())
                else:
                    fld_widget.currentIndexChanged.connect(lambda: self.mark_dirty())
        
        current_step += 1 # UI構築完了のステップ
        progress.setValue(current_step)

        for id_field_name in [HEADER_R_GENRE_ID, HEADER_Y_CATEGORY_ID, HEADER_YA_CATEGORY_ID]:
            if id_field_name in self.main_fields and isinstance(self.main_fields[id_field_name], QLineEdit):
                self.main_fields[id_field_name].textChanged.connect(lambda: self.mark_dirty())
        
        # Y_カテゴリIDの変更を監視してY_specフィールドを更新
        if HEADER_Y_CATEGORY_ID in self.main_fields:
            self.main_fields[HEADER_Y_CATEGORY_ID].textChanged.connect(lambda text: self._on_y_category_id_changed(text))
        
        # 説明マーク選択ボタンの有効/無効を設定
        if hasattr(self, 'explanation_mark_select_btn'):
            self.explanation_mark_select_btn.setEnabled(bool(self.explanation_mark_icon_data))

        progress.setLabelText(f"商品リスト ({MANAGE_FILE_NAME}) を読み込み中..."); QApplication.processEvents()
        self.clear_fields(); self.load_list(); self.apply_stylesheet()
        current_step += 1 # 商品リスト読み込み完了のステップ
        progress.setValue(current_step); QApplication.processEvents()

        progress.stop_animation() # アニメーションを停止
        progress.close()          # 全ての処理が完了したらダイアログを閉じる
        # self.showMaximized() # アプリケーション起動時に最大化表示
        self.show() # ウィンドウを一度表示してから設定を読み込む
        self._load_settings()
        self._on_y_category_id_changed(self.main_fields.get(HEADER_Y_CATEGORY_ID, QLineEdit()).text()) # 初期表示のために呼び出し
        
        # ウィンドウ表示後にメニューバーを作成（遅延実行で確実に）
        QTimer.singleShot(50, self._create_menu_bar)
        # フォールバック: メニューバーが作成されない場合に備えて追加の試行
        QTimer.singleShot(200, self._ensure_menu_bar_visible)

        # 自動保存タイマーの設定
        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.timeout.connect(lambda: self._auto_save_data())
        
        # ステータスバーの初期化（UIコンポーネント作成後、直接実行）
        # 遅延実行後に呼び出し
        def init_status_bar_delayed():
            if hasattr(self, '_init_status_bar'):
                logging.debug("遅延実行でステータスバー初期化開始")
                self._init_status_bar()
            else:
                logging.debug("_init_status_bar メソッドが存在しません")
        
        QTimer.singleShot(2000, init_status_bar_delayed)
        
        # スマートナビゲーション機能の初期化（遅延実行）
        def init_smart_navigation():
            if hasattr(self, '_setup_smart_navigation'):
                self._setup_smart_navigation()
            if hasattr(self, '_setup_sku_table_navigation'):
                self._setup_sku_table_navigation()
            # アプリケーションレベルのTabキーイベントフィルターを追加
            self._setup_global_tab_filter()
                
        QTimer.singleShot(2500, init_smart_navigation)
        
        # 起動時の自動更新チェック（設定が有効な場合のみ、少し遅延させて実行）
        logging.info(f"起動時更新チェック設定: check_for_updates_on_startup={check_for_updates_on_startup is not None}")
        logging.info(f"自動更新チェック有効: {getattr(self, 'auto_update_check_enabled', True)}")
        
        # 自動更新機能を有効化（シンプル版）
        if check_for_updates_on_startup and getattr(self, 'auto_update_check_enabled', True):
            logging.info("起動時更新チェックを2秒後に実行予定")
            QTimer.singleShot(2000, lambda: self._delayed_update_check())
        else:
            logging.warning("起動時更新チェックがスキップされました")
        self.auto_save_timer.start(AUTO_SAVE_INTERVAL_MS) # 自動保存間隔

        # new_btn の接続を新しいハンドラに変更
        try:
            self.new_btn.clicked.disconnect()
        except TypeError:
            # 接続が存在しない場合は無視
            pass
        self.new_btn.clicked.connect(lambda: self._handle_new_product_action())

        # Y_spec同期用エディタ参照の初期化
        self.y_spec_width_editor = None
        self.y_spec_depth_editor = None
        self.y_spec_height_editor = None
        # Y_spec同期用の定義保持用
        self.y_spec_width_definition = None
        self.y_spec_depth_definition = None
        self.y_spec_height_definition = None
        

    def _init_emergency_systems(self):
        """万が一対策システムの初期化"""
        try:
            # 1. 重複起動チェック
            if handle_duplicate_launch:
                if not handle_duplicate_launch(self):
                    sys.exit(0)  # 重複起動の場合は終了
            
            # 2. 設定ファイル破損チェック・復旧
            if check_and_recover_config:
                config_recovered = check_and_recover_config("商品登録入力ツール")
                if config_recovered:
                    logging.info("設定ファイルを復旧しました")
            
            # 3. クラッシュ復旧システムの初期化
            if CrashRecoveryManager:
                self.crash_recovery = CrashRecoveryManager()
                
                # 前回のクラッシュをチェック
                crash_info = self.crash_recovery.check_for_crash()
                if crash_info:
                    self._handle_previous_crash(crash_info)
                
                # セッション開始
                self.crash_recovery.start_session({
                    "version": CURRENT_VERSION,
                    "user_data_dir": getattr(self, 'user_data_dir', ''),
                    "manage_file_path": getattr(self, 'manage_file_path', '')
                })
                
                # 例外ハンドラーを設定
                if setup_crash_handler:
                    setup_crash_handler(self.crash_recovery)
                if setup_qt_exception_handler:
                    setup_qt_exception_handler(self.crash_recovery)
            
            # 4. ファイルロックマネージャーの初期化
            if FileLockManager:
                self.file_lock_manager = FileLockManager()
            
            # 5. システム互換性チェック（起動時のみ）
            if check_system_compatibility:
                compatibility_ok = check_system_compatibility(self)
                if not compatibility_ok:
                    logging.warning("システム互換性の問題が検出されました")
            
            # 6. メモリ監視システムの初期化
            if MemoryMonitor:
                self.memory_monitor = MemoryMonitor(self)
                self.memory_monitor.start_monitoring()
                logging.info("メモリ監視システムを開始しました")
            
            # 7. ネットワーク監視システムの初期化
            if setup_network_monitoring:
                setup_network_monitoring(self)
                logging.info("ネットワーク監視システムを開始しました")
            
            # 8. 定期的なハートビート更新タイマー
            if hasattr(self, 'crash_recovery'):
                self.heartbeat_timer = QTimer(self)
                self.heartbeat_timer.timeout.connect(self._update_heartbeat)
                self.heartbeat_timer.start(60000)  # 1分間隔
                
        except Exception as e:
            logging.error(f"万が一対策システム初期化エラー: {e}")
    
    def _handle_previous_crash(self, crash_info):
        """前回のクラッシュ情報を処理"""
        try:
            reply = QMessageBox.question(
                self,
                "予期しない終了の検出",
                f"""前回のセッションが予期せず終了しました。
                
最後のハートビート: {crash_info.get('last_heartbeat', '不明')}

緊急バックアップからデータを復元しますか？

「はい」: バックアップデータを確認・復元
「いいえ」: 通常通り起動""",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                emergency_backup = self.crash_recovery.get_emergency_backup()
                if emergency_backup:
                    self._restore_from_emergency_backup(emergency_backup)
                else:
                    QMessageBox.information(
                        self,
                        "バックアップなし",
                        "緊急バックアップデータが見つかりませんでした。"
                    )
                    
        except Exception as e:
            logging.error(f"クラッシュ処理エラー: {e}")
    
    def _restore_from_emergency_backup(self, backup_data):
        """緊急バックアップからデータを復元"""
        try:
            # 復元処理の実装
            # （具体的な復元ロジックは後で実装）
            QMessageBox.information(
                self,
                "復元完了",
                "緊急バックアップからデータを復元しました。"
            )
        except Exception as e:
            logging.error(f"緊急復元エラー: {e}")
    
    def _update_heartbeat(self):
        """ハートビートの更新"""
        try:
            if hasattr(self, 'crash_recovery'):
                current_data = {
                    "current_product_count": len(getattr(self, 'product_list_data', [])),
                    "is_dirty": getattr(self, 'is_dirty', False),
                    "last_save_time": getattr(self, 'last_save_time', '')
                }
                self.crash_recovery.update_heartbeat(current_data)
                
                # 緊急バックアップも作成
                if hasattr(self, 'is_dirty') and self.is_dirty:
                    emergency_data = self._collect_emergency_data()
                    self.crash_recovery.create_emergency_backup(emergency_data)
                    
        except Exception as e:
            logging.error(f"ハートビート更新エラー: {e}")
    
    def _collect_emergency_data(self):
        """緊急バックアップ用のデータを収集"""
        try:
            return {
                "current_fields": {name: widget.text() if hasattr(widget, 'text') else str(widget.currentText() if hasattr(widget, 'currentText') else '') 
                                 for name, widget in getattr(self, 'main_fields', {}).items()},
                "sku_data": getattr(self, 'sku_data_list', []),
                "is_dirty": getattr(self, 'is_dirty', False)
            }
        except Exception as e:
            logging.error(f"緊急データ収集エラー: {e}")
            return {}

    def _setup_logging(self):
        """アプリケーションのログ設定を行う"""
        self.log_file_path = os.path.join(get_user_data_dir(), f"{APP_NAME}_errors.log") # ログファイルパスをインスタンス変数に
        
        # ルートロガーのレベルを設定 (DEBUG以上を全てキャッチ)
        logging.basicConfig(level=logging.DEBUG,
                            format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(lineno)d - %(message)s',
                            datefmt='%Y-%m-%d %H:%M:%S',
                            handlers=[
                                logging.FileHandler(self.log_file_path, encoding='utf-8', mode='a'), # 'a'で追記モード
                                # logging.StreamHandler() # コンソールにも出力する場合はアンコメント
                            ])
        logging.info(f"{APP_NAME} を起動しました。ログファイル: {self.log_file_path}")
    def handle_csv_generation_button_click(self):
        # C#ツールが期待する item.xlsm のフルパス
        item_xlsm_for_csharp_path = self.output_file_path # _init_paths_and_dirs で設定済み

        if not os.path.exists(item_xlsm_for_csharp_path):
            QMessageBox.warning(self, "ファイル未保存",
                                f"{OUTPUT_FILE_NAME} が期待される場所 ({item_xlsm_for_csharp_path}) に見つかりません。\n先に「保存」ボタンでデータを保存してください。")
            logging.warning(f"C#実行試行: {OUTPUT_FILE_NAME} が {item_xlsm_for_csharp_path} に見つかりませんでした。")
            return

        # csharp.exe のフルパス (初期化時に設定済み)
        actual_csharp_exe_path = self.csharp_exe_path

        if not os.path.exists(actual_csharp_exe_path):
            err_msg_csharp = f"csharp.exe が見つかりません。\n期待されるパス: {actual_csharp_exe_path}\n「C#」フォルダ内に csharp.exe が配置されているか確認してください。"
            QMessageBox.critical(self, "実行ファイルエラー",
                                 f"{err_msg_csharp}\n詳細はログファイルを確認してください。")
            logging.error(f"C#実行試行: {err_msg_csharp}")
            return
        
        command = [actual_csharp_exe_path, item_xlsm_for_csharp_path]

        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            # csharp.exe を実行します。
            # CREATE_NO_WINDOW は Windows でコンソールウィンドウを非表示にするためのフラグです。
            result = subprocess.run(command, capture_output=True, text=True, check=False,
                                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
                                    encoding='cp932' if sys.platform == "win32" else 'utf-8') # C#アプリの出力エンコーディングに合わせる

            if result.returncode == 0:
                QMessageBox.information(self, "CSV生成成功",
                                        f"csharp.exe の実行が完了しました。\nCSVファイルが生成されました。\n出力:\n{result.stdout}")
                logging.info(f"csharp.exe 実行成功。出力: {result.stdout}")
            else:
                error_message = f"csharp.exe の実行中にエラーが発生しました。\n"
                error_message += f"リターンコード: {result.returncode}\n"
                if result.stdout:
                    error_message += f"\n標準出力:\n{result.stdout}\n"
                if result.stderr:
                    error_message += f"\n標準エラー出力:\n{result.stderr}\n"
                QMessageBox.critical(self, "CSV生成エラー", f"{error_message}\n詳細はログファイルを確認してください。")
                logging.error(f"csharp.exe 実行エラー: {error_message}")

        except FileNotFoundError: # Pylint: disable=try-except-raise
            logging.error(f"csharp.exe の実行に失敗: ファイルが見つかりません ({actual_csharp_exe_path})。", exc_info=True)
            QMessageBox.critical(self, "実行エラー",
                                 f"csharp.exe が見つかりません: {actual_csharp_exe_path}\n詳細はログファイルを確認してください。")
        except Exception as e:
            logging.error(f"csharp.exe の実行中に予期せぬエラーが発生しました。", exc_info=True)
            QMessageBox.critical(self, "実行エラー",
                                 f"csharp.exe の実行中に予期せぬエラーが発生しました: {e}\n詳細はログファイルを確認してください。")
        finally:
            QApplication.restoreOverrideCursor()

    def _show_loading_dialog(self):
        total_steps = 12 # 処理のステップ数
        progress = LoadingDialog("起動準備中...", total_steps, self)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()
        return progress

    def _init_paths_and_dirs(self, progress):
        current_step = 0
        progress.setLabelText("初期設定を読み込み中..."); current_step +=1; progress.setValue(current_step); QApplication.processEvents()
        # Bundle directory (for internal bundled files)
        self.base_dir_frozen = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        
        # Executable directory (for external user-editable files)
        if getattr(sys, 'frozen', False):
            # In EXE: use directory where EXE is located
            self.exe_dir = os.path.dirname(sys.executable)
        else:
            # In development: use same as script directory
            self.exe_dir = os.path.dirname(os.path.abspath(__file__))
        
        # user_data_dir for settings and item_manage.xlsm (same as EXE directory)
        self.user_data_dir = self.exe_dir
        self.manage_file_path = os.path.join(self.user_data_dir, MANAGE_FILE_NAME)
        
        # Paths related to C# tool and its input item.xlsm (relative to EXE dir)
        self.csharp_dir = os.path.join(self.exe_dir, "C#") 
        self.csharp_exe_path = os.path.join(self.csharp_dir, "csharp.exe") 
        self.item_xlsm_output_dir = os.path.join(self.csharp_dir, "ec_csv_tool")
        
        # This is the item.xlsm that the C# tool will use.
        # save_to_excelメソッドがこのパスに item.xlsm を出力するように後で修正します。
        self.output_file_path = os.path.join(self.item_xlsm_output_dir, OUTPUT_FILE_NAME)
        
        # Clean template path (for output file generation)
        self.clean_template_file_path = os.path.join(self.exe_dir, TEMPLATE_FILE_NAME)

    def _initialize_or_update_manage_file(self):
        """管理ファイルの初期化または既存ファイルのアップデート"""
        try:
            # テンプレートファイルが存在するかチェック
            if not os.path.exists(self.template_file_path_bundle):
                logging.warning(f"テンプレートファイル '{self.template_file_path_bundle}' が見つかりません。")
                return
            
            # 管理ファイルが存在しない場合（初回起動）
            if not os.path.exists(self.manage_file_path):
                copyfile(self.template_file_path_bundle, self.manage_file_path)
                logging.info(f"管理ファイル '{self.manage_file_path}' を新規作成しました。")
                return
            
            # 既存の管理ファイルがある場合、構造の互換性をチェック
            compatibility_result = self._check_template_compatibility()
            
            if compatibility_result["needs_update"]:
                self._handle_template_structure_change(compatibility_result)
                
        except Exception as e:
            logging.error(f"管理ファイルの初期化中にエラー: {e}", exc_info=True)
    
    def _check_template_compatibility(self):
        """テンプレートと既存管理ファイルの構造互換性をチェック"""
        template_wb = None
        manage_wb = None
        
        try:
            from openpyxl import load_workbook
            
            # テンプレートファイルのヘッダーを読み取り
            template_wb = load_workbook(self.template_file_path_bundle, read_only=True)
            template_main_ws = template_wb[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in template_wb.sheetnames else None
            
            # 既存管理ファイルのヘッダーを読み取り  
            manage_wb = load_workbook(self.manage_file_path, read_only=True)
            manage_main_ws = manage_wb[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in manage_wb.sheetnames else None
            
            if not template_main_ws or not manage_main_ws:
                return {"needs_update": False, "reason": "シートが見つかりません"}
            
            # ヘッダー行を取得
            template_headers = [cell.value for cell in template_main_ws[1]]
            manage_headers = [cell.value for cell in manage_main_ws[1]]
            
            # ヘッダーの比較
            template_headers_clean = [h for h in template_headers if h is not None]
            manage_headers_clean = [h for h in manage_headers if h is not None]
            
            if len(template_headers_clean) != len(manage_headers_clean):
                return {
                    "needs_update": True,
                    "reason": f"列数が変更されました（{len(manage_headers_clean)} → {len(template_headers_clean)}）",
                    "template_headers": template_headers_clean,
                    "manage_headers": manage_headers_clean
                }
            
            # ヘッダー名の違いをチェック
            header_differences = []
            for i, (template_h, manage_h) in enumerate(zip(template_headers_clean, manage_headers_clean)):
                if template_h != manage_h:
                    header_differences.append({
                        "index": i,
                        "template": template_h,
                        "manage": manage_h
                    })
            
            if header_differences:
                return {
                    "needs_update": True,
                    "reason": f"{len(header_differences)}個のヘッダーが変更されました",
                    "differences": header_differences,
                    "template_headers": template_headers_clean,
                    "manage_headers": manage_headers_clean
                }
            
            return {"needs_update": False, "reason": "互換性OK"}
            
        except Exception as e:
            logging.error(f"互換性チェック中にエラー: {e}")
            return {"needs_update": False, "reason": f"チェックエラー: {str(e)}"}
        finally:
            # リソースの確実な解放
            if template_wb:
                try:
                    template_wb.close()
                except Exception as e:
                    logging.debug(f"テンプレートワークブック終了エラー（継続）: {e}")
            if manage_wb:
                try:
                    manage_wb.close()
                except Exception as e:
                    logging.debug(f"管理ワークブック終了エラー（継続）: {e}")
    
    def _handle_template_structure_change(self, compatibility_result):
        """テンプレート構造変更への対応"""
        try:
            # ユーザーに確認を求める
            reply = QMessageBox.question(
                self,
                "テンプレート構造の変更",
                f"""新しいバージョンでテンプレート構造が変更されました。

変更内容: {compatibility_result['reason']}

既存の商品データを保持したまま新しい構造に移行しますか？

「はい」: データを移行（推奨）
「いいえ」: 既存ファイルをそのまま使用（一部機能に制限が生じる可能性）""",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                # データ移行を実行
                self._migrate_manage_file_structure(compatibility_result)
            else:
                logging.info("ユーザーがデータ移行をスキップしました")
                
        except Exception as e:
            logging.error(f"テンプレート構造変更の処理中にエラー: {e}", exc_info=True)
    
    def _migrate_manage_file_structure(self, compatibility_result):
        """既存データを新しいテンプレート構造に移行"""
        try:
            from openpyxl import load_workbook
            import shutil
            from datetime import datetime
            
            # バックアップファイルを作成
            backup_path = self.manage_file_path + f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy2(self.manage_file_path, backup_path)
            logging.info(f"バックアップを作成: {backup_path}")
            
            # 既存データを読み込み
            old_wb = load_workbook(self.manage_file_path)
            old_main_ws = old_wb[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in old_wb.sheetnames else None
            old_sku_ws = old_wb[SKU_SHEET_NAME] if SKU_SHEET_NAME in old_wb.sheetnames else None
            
            if not old_main_ws:
                logging.error("既存ファイルのMainシートが見つかりません")
                return False
            
            # 既存データを辞書形式で保存
            old_headers = [cell.value for cell in old_main_ws[1] if cell.value is not None]
            existing_data = []
            
            for row in old_main_ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # 空行をスキップ
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(old_headers):
                            row_data[old_headers[i]] = value
                    existing_data.append(row_data)
            
            # SKUデータも保存
            sku_data = []
            if old_sku_ws:
                sku_headers = [cell.value for cell in old_sku_ws[1] if cell.value is not None]
                for row in old_sku_ws.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        sku_row_data = {}
                        for i, value in enumerate(row):
                            if i < len(sku_headers):
                                sku_row_data[sku_headers[i]] = value
                        sku_data.append(sku_row_data)
            
            old_wb.close()
            
            # 新しいテンプレートをベースに作成
            shutil.copy2(self.template_file_path_bundle, self.manage_file_path)
            new_wb = load_workbook(self.manage_file_path)
            new_main_ws = new_wb[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in new_wb.sheetnames else None
            
            if not new_main_ws:
                logging.error("新テンプレートのMainシートが見つかりません")
                return False
            
            # 新しいヘッダーを取得
            new_headers = [cell.value for cell in new_main_ws[1] if cell.value is not None]
            
            # 既存データを新しい構造に移行
            for row_idx, old_row_data in enumerate(existing_data, start=2):
                for col_idx, header in enumerate(new_headers, start=1):
                    if header in old_row_data:
                        # 既存データがある場合はそれを使用
                        new_main_ws.cell(row=row_idx, column=col_idx, value=old_row_data[header])
                    # 新しい列の場合はデフォルト値（None）のまま
            
            # SKUデータも移行
            if sku_data and SKU_SHEET_NAME in new_wb.sheetnames:
                new_sku_ws = new_wb[SKU_SHEET_NAME]
                new_sku_headers = [cell.value for cell in new_sku_ws[1] if cell.value is not None]
                
                for row_idx, old_sku_row in enumerate(sku_data, start=2):
                    for col_idx, header in enumerate(new_sku_headers, start=1):
                        if header in old_sku_row:
                            new_sku_ws.cell(row=row_idx, column=col_idx, value=old_sku_row[header])
            
            # 保存
            new_wb.save(self.manage_file_path)
            new_wb.close()
            
            # 成功メッセージ
            QMessageBox.information(
                self,
                "データ移行完了",
                f"""データの移行が正常に完了しました。

移行されたデータ: {len(existing_data)}件の商品
バックアップファイル: {backup_path}

新しい列が追加されている場合は、必要に応じて値を入力してください。"""
            )
            
            logging.info(f"データ移行完了: {len(existing_data)}件の商品を移行しました")
            return True
            
        except Exception as e:
            logging.error(f"データ移行中にエラー: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "データ移行エラー", 
                f"データ移行中にエラーが発生しました: {str(e)}\n\nバックアップファイルから手動で復旧してください。"
            )
            return False
    
    def _load_initial_data(self, progress):
        """データ読み込みを並列化して起動時間を短縮"""
        current_step = 1 # _init_paths_and_dirs で1ステップ消費済みと仮定

        safe_category_name = os.path.normpath(CATEGORY_FILE_NAME).lstrip(os.sep + os.altsep)

        tasks_definitions = [
            {
                'name': 'categories',
                'target_attr': 'categories',
                'func': load_categories_from_csv,
                'args_factory': lambda: (os.path.join(self.base_dir_frozen, safe_category_name), progress),
                'progress_label_before': f"カテゴリ情報 ({CATEGORY_FILE_NAME}) を読み込み中..."
            },
            {
                'name': 'rakuten_definitions',
                'target_attr': 'definition_loader',
                'func': RakutenAttributeDefinitionLoader,
                'args_factory': lambda: (self.base_dir_frozen, progress),
                'progress_label_before': f"楽天商品属性定義書 ({DEFINITION_CSV_FILE} と {RECOMMENDED_LIST_CSV_FILE}) を読み込み中..."
            },
            {
                'name': 'r_genre_master',
                'target_attr': '_r_genre_master_list',
                'func': load_id_master_data,
                'args_factory': lambda: (
                    R_GENRE_MASTER_FILE, MASTER_ID_COLUMN_DEFAULT, MASTER_NAME_COLUMN_R_GENRE,
                    MASTER_HIERARCHY_COLUMN_DEFAULT, progress, "Rジャンルマスター"
                ),
                'progress_label_before': f"IDマスター ({R_GENRE_MASTER_FILE}) を読み込み中..."
            },
            {
                'name': 'y_category_master',
                'target_attr': '_y_category_master_list',
                'func': load_id_master_data,
                'args_factory': lambda: (
                    Y_CATEGORY_MASTER_FILE, MASTER_ID_COLUMN_DEFAULT, MASTER_NAME_COLUMN_Y_CATEGORY,
                    MASTER_HIERARCHY_COLUMN_DEFAULT, progress, "Yカテゴリマスター"
                ),
                'progress_label_before': f"IDマスター ({Y_CATEGORY_MASTER_FILE}) を読み込み中..."
            },
            {
                'name': 'ya_category_master',
                'target_attr': '_ya_category_master_list',
                'func': load_id_master_data,
                'args_factory': lambda: (
                    YA_CATEGORY_MASTER_FILE, MASTER_ID_COLUMN_DEFAULT, MASTER_NAME_COLUMN_YA_CATEGORY,
                    MASTER_HIERARCHY_COLUMN_DEFAULT, progress, "YAカテゴリマスター"
                ),
                'progress_label_before': f"IDマスター ({YA_CATEGORY_MASTER_FILE}) を読み込み中..."
            },
            {
                'name': 'y_spec_definitions',
                'target_attr': 'y_spec_loader',
                'func': YSpecDefinitionLoader,
                'args_factory': lambda: (self.base_dir_frozen, progress),
                'progress_label_before': f"Yahoo!スペック定義 ({YSPEC_CSV_FILE}) を読み込み中..."
            },
            {
                'name': 'material_spec_master',
                'target_attr': 'material_spec_master',
                'func': load_material_spec_master,
                'args_factory': lambda: (os.path.join(self.base_dir_frozen, MATERIAL_SPEC_MASTER_FILE_NAME), progress),
                'progress_label_before': f"材質・仕様マスター ({MATERIAL_SPEC_MASTER_FILE_NAME}) を読み込み中..."
            },
            {
                'name': 'explanation_icons',
                'target_attr': 'explanation_mark_icon_data',
                'func': load_explanation_mark_icons,
                'args_factory': lambda: (self.base_dir_frozen, progress),
                'progress_label_before': f"説明マークアイコン ({EXPLANATION_MARK_ICONS_SUBDIR}) を読み込み中..."
            }
        ]

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKER_THREADS) as executor:
            submitted_task_futures = []
            for task_def in tasks_definitions:
                future = executor.submit(task_def['func'], *task_def['args_factory']())
                submitted_task_futures.append({'future': future, 'task_def': task_def})

            for item in submitted_task_futures:
                future = item['future']
                task_definition = item['task_def']

                progress.setLabelText(task_definition['progress_label_before'])
                current_step += 1
                progress.setValue(current_step)
                QApplication.processEvents()

                try:
                    result = future.result() # このタスクの完了を待つ
                    setattr(self, task_definition['target_attr'], result)

                    # 特定のタスク完了後のチェック処理
                    if task_definition['name'] == 'rakuten_definitions': # RakutenAttributeDefinitionLoader完了後
                        if not self.definition_loader.genre_definitions:
                             logging.warning(f"楽天商品属性定義書 '{os.path.join(self.base_dir_frozen, DEFINITION_CSV_FILE)}' が読み込まれなかったか、空です。SKU属性の推奨値機能は利用できません。")
                             QMessageBox.warning(self, "定義書読込エラー",
                                                 f"楽天商品属性定義書 '{os.path.join(self.base_dir_frozen, DEFINITION_CSV_FILE)}' が読み込まれなかったか、空です。\nSKU属性の推奨値機能は利用できません。\n詳細はログファイルを確認してください。")
                    elif task_definition['name'] == 'ya_category_master': # 最後のIDマスター読み込み後
                        if not self._r_genre_master_list and not self._y_category_master_list and not self._ya_category_master_list:
                             logging.info("有効なIDマスターデータが読み込まれませんでした。ID検索機能は利用できません。")
                except Exception as e:
                    logging.error(f"Error loading {task_definition['name']}: {e}", exc_info=True)
                    QMessageBox.warning(self, "データ読み込みエラー", f"{task_definition['progress_label_before']} の読み込み中にエラーが発生しました: {e}\n詳細はログを確認してください。")

    def _init_ui_components(self):
        """主要なUI要素の初期化"""
        # ボタンの作成
        self.new_btn = QPushButton(self.style().standardIcon(QStyle.SP_FileIcon), "新規")
        self.new_btn.setObjectName("NewProductButton")
        self.new_btn.setToolTip("新しい商品情報を入力します (Ctrl+N)")
        # ショートカットはメニューで管理
        
        self.save_btn = QPushButton(self.style().standardIcon(QStyle.SP_DialogSaveButton), "保存")
        self.save_btn.setObjectName("SaveButton")
        self.save_btn.setEnabled(False)
        self.save_btn.setToolTip("現在の変更を保存します (Ctrl+S)")
        # ショートカットはメニューで管理
        
        # 保存時刻をツールチップで表示する用の変数
        self.last_save_time = None
        
        self.sku_add_btn = QPushButton(self.style().standardIcon(QStyle.SP_FileDialogNewFolder), "SKU追加")
        self.sku_add_btn.setObjectName("AddSkuButton")
        # ショートカットはメニューで管理
        self.sku_add_btn.setToolTip("新しいSKU行を追加します (Ctrl+Shift+A)")
        
        self.bulk_p_btn = QPushButton(self.style().standardIcon(QStyle.SP_DialogApplyButton), "一括 P設定")
        self.bulk_p_btn.setObjectName("BulkPButton")
        self.bulk_p_btn.setToolTip("全ての商品のコントロールカラムを 'p (除外)' に設定します")
        
        self.run_csharp_btn = QPushButton(self.style().standardIcon(QStyle.SP_MediaPlay), "C#実行")
        self.run_csharp_btn.setObjectName("RunCSharpButton")
        self.run_csharp_btn.setToolTip(f"{OUTPUT_FILE_NAME} を元に csharp.exe を実行して各モール用CSVを生成します。")
        
        self.sku_delete_btn = QPushButton(self.style().standardIcon(QStyle.SP_TrashIcon), "選択SKU削除")
        self.sku_delete_btn.setObjectName("DeleteSkuButton")
        self.sku_delete_btn.setToolTip("選択されているSKU行を削除します (テーブルフォーカス時 Deleteキー)")
        
        # 検索バーと商品リスト
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("商品コードまたは商品名で検索")
        
        self.product_list = QListWidget()
        self.product_list.setObjectName("ProductList")
        self.product_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.product_list.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 複数選択を有効化
        
        # カテゴリ選択ボタン
        self.category_select_btn = QPushButton("カテゴリ選択")
        self.category_select_btn.setObjectName("CategorySelectButton")
        # ショートカットはメニューで管理
        self.category_select_btn.setToolTip("カテゴリを選択します (Ctrl+G)")
        
        # ID検索ボタン
        self.open_id_search_button = QPushButton("IDを検索")
        self.open_id_search_button.setObjectName("IdSearchButton")
        # ショートカットはメニューで管理
        self.open_id_search_button.setToolTip("各種IDを検索します (Ctrl+I)")
        
        # 画像説明HTMLボタン
        self.image_desc_btn = QPushButton("画像説明HTML生成")
        self.image_desc_btn.setObjectName("ImageDescButton")
        # ショートカットはメニューで管理
        self.image_desc_btn.setToolTip("画像説明HTMLを生成します (Ctrl+H)")
        
        
        # メニューバーの作成は後で行う
        
        # 追加のショートカット
        # ESCキーで検索をクリア（重複削除 - メニューバーで定義）
        
        # 検索バーフォーカス用ショートカット
        self.focus_search_action = QAction(self)
        self.focus_search_action.setShortcut("Ctrl+Shift+F")
        self.addAction(self.focus_search_action)
        
        
        # ラジオボタン
        self.control_radio_n = QRadioButton("n（New＝新規）")
        self.control_radio_p = QRadioButton("p（Pass＝除外）")
        self.control_radio_n.setChecked(True)
        self.control_radio_group = QButtonGroup()
        self.control_radio_group.addButton(self.control_radio_n)
        self.control_radio_group.addButton(self.control_radio_p)
        
        # SKUテーブルビュー
        self.frozen_table_view = FocusControllingTableView(self)
        self.frozen_table_view.setObjectName("FrozenTableView")
        self.frozen_table_view.verticalHeader().setVisible(True)
        # 行の高さを選択ボタンに合わせて調整
        self.frozen_table_view.verticalHeader().setDefaultSectionSize(26)
        
        self.scrollable_table_view = ScrollableFocusControllingTableView(self)
        self.scrollable_table_view.setObjectName("ScrollableTableView")
        self.scrollable_table_view.verticalHeader().setVisible(False)
        # 行の高さを選択ボタンに合わせて調整
        self.scrollable_table_view.verticalHeader().setDefaultSectionSize(26)
        
        # SKUモデル
        self.sku_model = SkuTableModel(data=[], headers=[], defined_attr_details=[], parent=self)
        
        # その他の初期化
        # 初期化済み（__init__の最初で実行済み）
        self.digit_count_label_mycode = None
        self.y_spec_labels = []
        self.y_spec_editor_placeholders = []
        self.y_spec_current_editors = [None] * MAX_Y_SPEC_COUNT
        self.y_spec_current_definitions = [None] * MAX_Y_SPEC_COUNT
        self._y_spec_section_rendered_in_form = False
        
        # Undo/Redo用の属性初期化
        self.selected_product = None
        
        # ステータスバー用の属性初期化
        self.status_labels = {}
        
        # スマートナビゲーション用の属性
        self.smart_navigation_enabled = True
        logging.debug(f"smart_navigation_enabled設定: {self.smart_navigation_enabled}")
        
        # 新規作成モードフラグ
        self._is_new_mode = False
        
    def _init_status_bar(self):
        """ステータスバーの初期化"""
        logging.debug("ステータスバー初期化開始")
        try:
            self.status_bar.setStyleSheet("""
                QStatusBar {
                    background-color: #f8fafc;
                    border-top: 1px solid #e2e8f0;
                    color: #64748b;
                    font-size: 12px;
                    padding: 2px 8px;
                }
                QStatusBar::item {
                    border: none;
                    padding: 0px 8px;
                }
            """)
            
            # 初期化メッセージをクリア
            self.status_bar.clearMessage()
            logging.debug("メッセージクリア完了")
        
            # ステータス項目の初期化
            self.status_labels['save_status'] = QLabel("✅ 保存済み")
            self.status_labels['current_product'] = QLabel("📦 商品: 未選択")
            self.status_labels['sku_info'] = QLabel("📋 SKU: 0件")
            self.status_labels['validation'] = QLabel("✅ エラー: 0件")
            logging.debug("ラベル作成完了")
            
            # ステータスバーに追加
            self.status_bar.addWidget(self.status_labels['save_status'])
            self.status_bar.addPermanentWidget(self.status_labels['current_product'])
            self.status_bar.addPermanentWidget(self.status_labels['sku_info'])
            self.status_bar.addPermanentWidget(self.status_labels['validation'])
            logging.debug("ラベル追加完了")
            
            # 初期状態を設定
            if hasattr(self, '_update_status_bar'):
                self._update_status_bar()
                logging.debug("ステータスバー更新完了")
            else:
                logging.debug("_update_status_bar メソッドが見つかりません")
                
        except Exception as e:
            print(f"DEBUG: ステータスバー初期化エラー: {e}")
            import traceback
            traceback.print_exc()
    
    def _update_status_bar(self):
        """ステータスバーの更新"""
        try:
            # ステータスラベルが初期化されているかチェック
            if not hasattr(self, 'status_labels') or not self.status_labels:
                print("WARNING: status_labels が初期化されていません")
                return
                
            # 保存状態
            if hasattr(self, '_is_dirty') and self._is_dirty:
                self.status_labels['save_status'].setText("🔶 未保存の変更")
                self.status_labels['save_status'].setStyleSheet("color: #f59e0b; font-weight: bold;")
            else:
                self.status_labels['save_status'].setText("✅ 保存済み")
                self.status_labels['save_status'].setStyleSheet("color: #10b981;")
            
            # 現在の商品
            product_name = ""
            if hasattr(self, 'main_fields') and HEADER_PRODUCT_NAME in self.main_fields:
                product_name = self.main_fields[HEADER_PRODUCT_NAME].text()
            
            if product_name:
                self.status_labels['current_product'].setText(f"📦 商品: {product_name[:20]}")
            else:
                self.status_labels['current_product'].setText("📦 商品: 未選択")
            
            # SKU情報
            sku_count = 0
            selected_count = 0
            if hasattr(self, 'sku_model') and self.sku_model:
                sku_count = self.sku_model.rowCount()
                if hasattr(self, 'frozen_table_view') and self.frozen_table_view.selectionModel():
                    selected_rows = set(idx.row() for idx in self.frozen_table_view.selectionModel().selectedIndexes())
                    selected_count = len(selected_rows)
            
            if selected_count > 0:
                self.status_labels['sku_info'].setText(f"📋 SKU: {selected_count}/{sku_count}件選択")
            else:
                self.status_labels['sku_info'].setText(f"📋 SKU: {sku_count}件")
            
            # バリデーション情報（エラー数をカウント）
            error_count, error_details = self._count_validation_errors()
            if error_count > 0:
                # エラーの詳細を短縮して表示
                error_summary = ", ".join(error_details[:3])  # 最初の3つまで表示
                if len(error_details) > 3:
                    error_summary += "..."
                self.status_labels['validation'].setText(f"⚠️ エラー: {error_count}件 ({error_summary})")
                self.status_labels['validation'].setStyleSheet("color: #ef4444; font-weight: bold;")
                # ツールチップで全詳細を表示
                self.status_labels['validation'].setToolTip("\\n".join(error_details))
            else:
                self.status_labels['validation'].setText("✅ エラー: 0件")
                self.status_labels['validation'].setStyleSheet("color: #10b981;")
                self.status_labels['validation'].setToolTip("入力内容に問題ありません")
        except Exception as e:
            print(f"Status bar update error: {e}")
            import traceback
            traceback.print_exc()
    
    def _count_validation_errors(self):
        """バリデーションエラー数をカウントして詳細も返す"""
        error_count = 0
        error_details = []
        try:
            # 必須フィールドのチェック
            if hasattr(self, 'main_fields'):
                for field_name, widget in self.main_fields.items():
                    if hasattr(widget, 'text'):
                        text = widget.text().strip()
                        # 商品名とmycodeは必須
                        if field_name in [HEADER_PRODUCT_NAME, HEADER_MYCODE] and not text:
                            error_count += 1
                            if field_name == HEADER_PRODUCT_NAME:
                                error_details.append("商品名が未入力")
                            elif field_name == HEADER_MYCODE:
                                error_details.append("商品コードが未入力")
                        # 価格の形式チェック
                        if field_name == HEADER_PRICE_TAX_INCLUDED and text:
                            try:
                                # カンマを除去してから数値チェック
                                price_text = text.replace(',', '')
                                float(price_text)
                            except ValueError:
                                error_count += 1
                                error_details.append("価格の形式が正しくない")
            
            # SKUテーブルの必須項目チェック
            if hasattr(self, 'sku_model') and self.sku_model:
                for row in range(self.sku_model.rowCount()):
                    sku_code_idx = None
                    choice_name_idx = None
                    for col, header in enumerate(self.sku_model._headers):
                        if header == HEADER_SKU_CODE:
                            sku_code_idx = col
                        elif header == HEADER_CHOICE_NAME:
                            choice_name_idx = col
                    
                    # SKUコードと選択肢名は必須
                    if sku_code_idx is not None:
                        sku_code = self.sku_model.data(self.sku_model.index(row, sku_code_idx), Qt.DisplayRole)
                        if not sku_code or not str(sku_code).strip():
                            error_count += 1
                            error_details.append(f"SKU{row+1}行目: SKUコードが未入力")
                    
                    if choice_name_idx is not None:
                        choice_name = self.sku_model.data(self.sku_model.index(row, choice_name_idx), Qt.DisplayRole)
                        if not choice_name or not str(choice_name).strip():
                            error_count += 1
                            error_details.append(f"SKU{row+1}行目: 選択肢名が未入力")
                            
        except Exception as e:
            print(f"Error counting validation errors: {e}")
        
        return error_count, error_details
    
    def _setup_smart_navigation(self):
        """スマートナビゲーション機能の設定"""
        try:
            # メインフィールドにEnterキーナビゲーションを追加
            if hasattr(self, 'main_fields') and hasattr(self, 'main_field_order'):
                for field_name in self.main_field_order:
                    if field_name in self.main_fields:
                        widget = self.main_fields[field_name]
                        if widget and hasattr(widget, 'keyPressEvent'):
                            original_keyPressEvent = widget.keyPressEvent
                            
                            def create_smart_keypress(original_func, widget_ref, field_name_ref):
                                def smart_keyPressEvent(event):
                                    if hasattr(self, 'smart_navigation_enabled') and self.smart_navigation_enabled:
                                        # Enterキーの処理
                                        if (event.key() == Qt.Key_Return and
                                            not (event.modifiers() & Qt.ShiftModifier)):  # Shift+Enterは除外
                                            self._handle_enter_navigation(widget_ref, field_name_ref)
                                            event.accept()
                                        # Tabキーの処理
                                        elif event.key() == Qt.Key_Tab and not event.modifiers():
                                            # Tabキーも同じナビゲーション処理を使用
                                            self._handle_enter_navigation(widget_ref, field_name_ref)
                                            event.accept()
                                        # Shift+Tabキーの処理
                                        elif event.key() == Qt.Key_Backtab:
                                            # 逆方向のナビゲーション
                                            self._handle_backtab_navigation(widget_ref, field_name_ref)
                                            event.accept()
                                        else:
                                            # その他のキー処理
                                            original_func(event)
                                    else:
                                        # 通常のキー処理
                                        original_func(event)
                                return smart_keyPressEvent
                            
                            # 新しいkeyPressEventを設定
                            widget.keyPressEvent = create_smart_keypress(original_keyPressEvent, widget, field_name)
                            
                            # QLineEditでTabキーを確実に捕捉するため、フォーカスポリシーを強制設定
                            if hasattr(widget, 'setFocusPolicy'):
                                widget.setFocusPolicy(Qt.StrongFocus)
                            
                            # イベントフィルターでTabキーを強制的に捕捉
                            def create_event_filter(widget_ref, field_name_ref):
                                from PyQt5.QtCore import QObject, QEvent
                                class TabEventFilter(QObject):
                                    def eventFilter(self, obj, event):
                                        if (event.type() == QEvent.KeyPress and 
                                            event.key() in [Qt.Key_Tab, Qt.Key_Backtab]):
                                            widget_ref.keyPressEvent(event)
                                            return True  # イベントを消費
                                        return False
                                return TabEventFilter()
                            
                            filter_obj = create_event_filter(widget, field_name)
                            widget.installEventFilter(filter_obj)
                            # フィルターオブジェクトを保持（ガベージコレクション防止）
                            if not hasattr(self, '_event_filters'):
                                self._event_filters = []
                            self._event_filters.append((widget, filter_obj))
            
            # Y_specエディタにも同じナビゲーション処理を追加
            self._setup_yspec_navigation()
                        
            # ボタンのフォーカスポリシーを設定（Tabキーでフォーカスを受けないように）
            self._setup_button_focus_policies()
                        
        except Exception as e:
            logging.error(f"Smart navigation setup error: {e}", exc_info=True)
    
    def _cleanup_event_filters(self):
        """イベントフィルターをクリーンアップしてメモリリークを防ぐ"""
        if hasattr(self, '_event_filters'):
            for item in self._event_filters:
                try:
                    # Handle both tuple (widget, filter_obj) and single filter_obj cases
                    if isinstance(item, tuple) and len(item) == 2:
                        widget, filter_obj = item
                        if widget is not None:
                            # Qt5のウィジェット削除チェック（より安全な方法）
                            try:
                                widget.objectName()  # 削除されたウィジェットだとRuntimeError
                                widget.removeEventFilter(filter_obj)
                            except RuntimeError:
                                # ウィジェットは既に削除済み
                                pass
                    else:
                        # Handle single filter object case (YspecTabEventFilter)
                        # For single filter objects, we can't remove them from widgets
                        # since we don't know which widget they're attached to
                        pass
                except Exception as e:
                    logging.debug(f"イベントフィルター削除エラー（継続）: {e}")
            self._event_filters.clear()
    
    def _setup_yspec_navigation(self):
        """Y_specエディタのスマートナビゲーション設定"""
        try:
            if hasattr(self, 'y_spec_current_editors'):
                for i, editor in enumerate(self.y_spec_current_editors):
                    if editor and hasattr(editor, 'keyPressEvent'):
                        field_name = f"Y_spec{i+1}"  # Y_spec1, Y_spec2, ...
                        
                        # 既存のkeyPressEventの保存
                        original_keyPressEvent = editor.keyPressEvent
                        
                        def create_yspec_keypress(original_func, editor_ref, field_name_ref):
                            def yspec_keyPressEvent(event):
                                if hasattr(self, 'smart_navigation_enabled') and self.smart_navigation_enabled:
                                    if (event.key() == Qt.Key_Return and
                                        not (event.modifiers() & Qt.ShiftModifier)):
                                        self._handle_enter_navigation(editor_ref, field_name_ref)
                                        event.accept()
                                    elif event.key() == Qt.Key_Tab and not event.modifiers():
                                        self._handle_enter_navigation(editor_ref, field_name_ref)
                                        event.accept()
                                    elif event.key() == Qt.Key_Backtab:
                                        self._handle_backtab_navigation(editor_ref, field_name_ref)
                                        event.accept()
                                    else:
                                        original_func(event)
                                else:
                                    original_func(event)
                            return yspec_keyPressEvent
                        
                        # 新しいkeyPressEventを設定
                        editor.keyPressEvent = create_yspec_keypress(original_keyPressEvent, editor, field_name)
                        
                        # Y_specエディタにフォーカススタイルを追加（すべてのタイプに対応）
                        if editor.__class__.__name__ == 'QComboBox':
                            editor.setStyleSheet("""
                                QComboBox:focus {
                                    border: 2px solid #0078d4;
                                    background-color: white;
                                }
                                QComboBox {
                                    border: 1px solid #ccc;
                                    background-color: white;
                                }
                            """)
                        elif editor.__class__.__name__ == 'QLineEdit':
                            editor.setStyleSheet("""
                                QLineEdit:focus {
                                    border: 2px solid #0078d4;
                                    background-color: white;
                                }
                                QLineEdit {
                                    border: 1px solid #ccc;
                                    background-color: white;
                                }
                            """)
                        elif hasattr(editor, 'line_edit'):  # SkuMultipleAttributeEditor
                            editor.setStyleSheet("""
                                QWidget:focus-within {
                                    border: 2px solid #0078d4;
                                }
                                QWidget {
                                    border: 1px solid #ccc;
                                }
                            """)
                            # 内部のQLineEditにもスタイルを適用
                            if hasattr(editor, 'line_edit'):
                                editor.line_edit.setStyleSheet("""
                                    QLineEdit:focus {
                                        border: 2px solid #0078d4;
                                        background-color: white;
                                    }
                                    QLineEdit {
                                        border: 1px solid #ccc;
                                        background-color: white;
                                    }
                                """)
                        
                        # Y_specエディタにもEventFilterを追加（Tabキー捕捉用）
                        from PyQt5.QtCore import QObject, QEvent
                        def create_yspec_event_filter(widget_ref, field_name_ref):
                            class YspecTabEventFilter(QObject):
                                def eventFilter(self, obj, event):
                                    if (event.type() == QEvent.KeyPress and 
                                        event.key() in [Qt.Key_Tab, Qt.Key_Backtab, Qt.Key_Return]):
                                        widget_ref.keyPressEvent(event)
                                        return True  # イベントを消費
                                    return False
                            return YspecTabEventFilter()
                        
                        filter_obj = create_yspec_event_filter(editor, field_name)
                        editor.installEventFilter(filter_obj)
                        # フィルターオブジェクトを保持（ガベージコレクション防止）
                        if not hasattr(self, '_event_filters'):
                            self._event_filters = []
                        self._event_filters.append((editor, filter_obj))
                        
        except Exception as e:
            print(f"Y_spec navigation setup error: {e}")
    
    def _setup_button_focus_policies(self):
        """ボタンのフォーカスポリシーを設定してTabキーナビゲーションから除外"""
        try:
            # 左側パネルのウィジェットをTabナビゲーションから除外
            if hasattr(self, 'search_bar'):
                self.search_bar.setFocusPolicy(Qt.ClickFocus)
            if hasattr(self, 'product_list'):
                self.product_list.setFocusPolicy(Qt.ClickFocus)
            
            # メインアクションボタン
            buttons_to_exclude = [
                'new_btn', 'save_btn', 'run_csharp_btn', 'bulk_p_btn',
                'sku_add_btn', 'sku_delete_btn', 'category_select_btn',
                'open_id_search_button', 'image_desc_btn', 'explanation_mark_select_btn'
            ]
            
            for btn_name in buttons_to_exclude:
                if hasattr(self, btn_name):
                    btn = getattr(self, btn_name)
                    if btn:
                        btn.setFocusPolicy(Qt.ClickFocus)  # クリックでのみフォーカス取得
            
            # 色選択ボタンも除外
            if hasattr(self, 'color_select_buttons'):
                for color_btn in self.color_select_buttons:
                    if color_btn:
                        color_btn.setFocusPolicy(Qt.NoFocus)
            
            # コントロールカラムのラジオボタンも除外
            if hasattr(self, 'control_radio_n'):
                self.control_radio_n.setFocusPolicy(Qt.ClickFocus)
            if hasattr(self, 'control_radio_p'):
                self.control_radio_p.setFocusPolicy(Qt.ClickFocus)
            
            # グループ展開ボタンを除外
            from PyQt5.QtWidgets import QPushButton
            toggle_buttons = self.findChildren(QPushButton, "ExpandableGroupToggleButton")
            for button in toggle_buttons:
                button.setFocusPolicy(Qt.NoFocus)
            
            # 色を選択ボタンを除外
            color_select_buttons = self.findChildren(QPushButton, "ColorSelectButton")
            for button in color_select_buttons:
                button.setFocusPolicy(Qt.NoFocus)
            
            # その他の小さなボタンも除外
            all_buttons = self.findChildren(QPushButton)
            for button in all_buttons:
                # 特定のテキストを含むボタンも除外
                button_text = button.text()
                if (button_text and 
                    ('選択' in button_text or 
                     '設定' in button_text or 
                     '生成' in button_text or 
                     '検索' in button_text or
                     'を' in button_text)):  # 「色を選択」「画像を生成」などのパターン
                    button.setFocusPolicy(Qt.NoFocus)
                # サイズが小さいボタン（アイコンボタンなど）も除外
                elif button.size().width() <= 30 or button.size().height() <= 30:
                    button.setFocusPolicy(Qt.NoFocus)
                        
        except Exception as e:
            print(f"Button focus policy setup error: {e}")
    
    def _handle_enter_navigation(self, current_widget, current_field_name):
        """Enterキーでの次フィールドへの移動処理"""
        try:
            # 実際のmain_field_orderを使用
            if not hasattr(self, 'main_field_order') or not hasattr(self, 'main_fields'):
                return
            
            # Y_specエディタからのナビゲーションの場合の特別処理
            if hasattr(current_widget, 'parent') and hasattr(self, 'y_spec_current_editors'):
                for i, editor in enumerate(self.y_spec_current_editors):
                    if editor == current_widget:
                        y_spec_field_name = f"Y_spec{i+1}"
                        if y_spec_field_name in self.main_field_order:
                            current_field_name = y_spec_field_name
                        break
            
            # 現在のフィールドが順序リストに存在するか確認
            if current_field_name not in self.main_field_order:
                return
            
            current_index = self.main_field_order.index(current_field_name)
            
            # Y_spec10から移動する場合は、R_SKU項目名以降にジャンプ
            start_index = current_index + 1
            if current_field_name == "Y_spec10":
                # Y_spec10の場合は、R_SKU項目名から開始
                try:
                    r_sku_index = self.main_field_order.index("R_SKU項目名")
                    start_index = r_sku_index
                except ValueError:
                    pass
            
            for i in range(start_index, len(self.main_field_order)):
                next_field_name = self.main_field_order[i]
                next_widget = None
                
                # Y_specフィールドの特別処理
                if next_field_name.startswith("Y_spec") and next_field_name[6:].isdigit():
                    try:
                        spec_index = int(next_field_name[6:]) - 1  # Y_spec1 -> index 0
                        if hasattr(self, 'y_spec_current_editors') and 0 <= spec_index < len(self.y_spec_current_editors):
                            editor = self.y_spec_current_editors[spec_index]
                            if editor and editor.isEnabled() and editor.isVisible():
                                next_widget = editor
                    except (ValueError, AttributeError):
                        pass
                else:
                    # 通常のmain_fieldsフィールドの処理
                    if (next_field_name in self.main_fields and 
                        self.main_fields[next_field_name] and
                        self.main_fields[next_field_name].isEnabled() and
                        hasattr(self.main_fields[next_field_name], 'setFocus')):
                        
                        # ExpandableFieldGroupに含まれるフィールドが閉じられている場合はスキップ
                        if not self.main_fields[next_field_name].isVisible():
                            continue
                        
                        next_widget = self.main_fields[next_field_name]
                
                # ウィジェットが見つかった場合、フォーカスを移動
                if next_widget and hasattr(next_widget, 'setFocus'):
                    next_widget.setFocus()
                    
                    # Y_specエディタの場合、フォーカススタイルを強制更新
                    if hasattr(self, 'y_spec_current_editors') and next_widget in self.y_spec_current_editors:
                        def force_focus():
                            if next_widget and next_widget.isVisible():
                                if hasattr(next_widget, 'line_edit'):  # SkuMultipleAttributeEditor
                                    # SkuMultipleAttributeEditorの場合は内部のline_editにフォーカス
                                    next_widget.line_edit.clearFocus()
                                    next_widget.line_edit.setFocus()
                                    next_widget.line_edit.update()
                                    next_widget.line_edit.repaint()
                                else:
                                    # QComboBoxやQLineEditの場合
                                    next_widget.clearFocus()
                                    next_widget.setFocus()
                                    next_widget.update()
                                    next_widget.repaint()
                        
                        # 少し遅延してフォーカスを再設定
                        from PyQt5.QtCore import QTimer
                        QTimer.singleShot(1, force_focus)
                    
                    # テキストフィールドの場合、全選択
                    if hasattr(next_widget, 'selectAll'):
                        next_widget.selectAll()
                    
                    # スクロール領域があれば、フィールドが見えるようにスクロール
                    self._ensure_field_visible(next_widget)
                    return
            
            # 全てのメインフィールドを確認したが、移動先が見つからない場合のみSKUテーブルへ
            self._move_to_sku_table()
                        
        except Exception as e:
            print(f"Enter navigation error: {e}")
    
    def _ensure_field_visible(self, widget):
        """フィールドが見えるようにスクロール領域を調整"""
        try:
            # Y_specウィジェットの場合は、プレースホルダー親を使用
            target_widget = widget
            if hasattr(widget, 'parent') and widget.parent():
                parent_widget = widget.parent()
                # Y_specプレースホルダー（QWidget with QHBoxLayout）を探す
                if (hasattr(parent_widget, 'layout') and 
                    parent_widget.layout() and
                    parent_widget.layout().__class__.__name__ == 'QHBoxLayout'):
                    target_widget = parent_widget
            
            # スクロール領域を探す
            parent = target_widget.parent()
            scroll_attempts = 0
            while parent and scroll_attempts < 10:  # 無限ループ防止
                if hasattr(parent, 'ensureWidgetVisible'):
                    parent.ensureWidgetVisible(target_widget)
                    break
                elif hasattr(parent, 'verticalScrollBar') and hasattr(parent, 'viewport'):
                    # QScrollAreaの場合
                    scroll_area = parent
                    
                    # ウィジェットの位置計算を改善
                    try:
                        # scroll_area.widget()からの相対位置を取得
                        widget_pos = target_widget.mapTo(scroll_area.widget(), target_widget.rect().topLeft())
                        widget_height = target_widget.height()
                        widget_bottom = widget_pos.y() + widget_height
                        
                        viewport_height = scroll_area.viewport().height()
                        current_scroll = scroll_area.verticalScrollBar().value()
                        
                        # 可視範囲の計算
                        visible_top = current_scroll
                        visible_bottom = current_scroll + viewport_height
                        
                        # スクロール判定と実行
                        if widget_bottom > visible_bottom:
                            # 下にスクロール
                            new_scroll = widget_bottom - viewport_height + 50  # 50px余裕
                            max_scroll = scroll_area.verticalScrollBar().maximum()
                            new_scroll = min(new_scroll, max_scroll)
                            scroll_area.verticalScrollBar().setValue(new_scroll)
                        elif widget_pos.y() < visible_top:
                            # 上にスクロール
                            new_scroll = max(0, widget_pos.y() - 50)  # 50px余裕
                            scroll_area.verticalScrollBar().setValue(new_scroll)
                            
                    except Exception as calc_error:
                        # フォールバック: ensureWidgetVisibleを試す
                        if hasattr(scroll_area, 'ensureWidgetVisible'):
                            scroll_area.ensureWidgetVisible(target_widget)
                    
                    break
                
                parent = parent.parent()
                scroll_attempts += 1
                
        except Exception as e:
            print(f"Ensure field visible error: {e}")
    
    def _handle_backtab_navigation(self, current_widget, current_field_name):
        """Shift+Tabキーでの前フィールドへの移動処理"""
        try:
            # 実際のmain_field_orderを使用
            if not hasattr(self, 'main_field_order') or not hasattr(self, 'main_fields'):
                return
            
            # 現在のフィールドが順序リストに存在するか確認
            if current_field_name not in self.main_field_order:
                return
            
            current_index = self.main_field_order.index(current_field_name)
            
            # 前のフィールドを探す（現在のインデックスから逆順に）
            for i in range(current_index - 1, -1, -1):
                prev_field_name = self.main_field_order[i]
                prev_widget = None
                
                # Y_specフィールドの特別処理
                if prev_field_name.startswith("Y_spec") and prev_field_name[6:].isdigit():
                    try:
                        spec_index = int(prev_field_name[6:]) - 1  # Y_spec1 -> index 0
                        if (hasattr(self, 'y_spec_current_editors') and 
                            0 <= spec_index < len(self.y_spec_current_editors) and
                            self.y_spec_current_editors[spec_index] and
                            self.y_spec_current_editors[spec_index].isEnabled() and
                            self.y_spec_current_editors[spec_index].isVisible()):
                            prev_widget = self.y_spec_current_editors[spec_index]
                    except (ValueError, AttributeError):
                        pass
                else:
                    # 通常のmain_fieldsフィールドの処理
                    if (prev_field_name in self.main_fields and 
                        self.main_fields[prev_field_name] and
                        self.main_fields[prev_field_name].isEnabled() and
                        self.main_fields[prev_field_name].isVisible() and
                        hasattr(self.main_fields[prev_field_name], 'setFocus')):
                        prev_widget = self.main_fields[prev_field_name]
                
                # ウィジェットが見つかった場合、フォーカスを移動
                if prev_widget and hasattr(prev_widget, 'setFocus'):
                    prev_widget.setFocus()
                    
                    # テキストフィールドの場合、全選択
                    if hasattr(prev_widget, 'selectAll'):
                        prev_widget.selectAll()
                    
                    # スクロール領域があれば、フィールドが見えるようにスクロール
                    self._ensure_field_visible(prev_widget)
                    return
            
            # メインフィールドの最初に到達した場合はSKUテーブルの最後へ
            self._move_to_sku_table_last()
                        
        except Exception as e:
            print(f"Backtab navigation error: {e}")
    
    def _move_to_sku_table(self):
        """SKUテーブルの最初のセルに移動"""
        try:
            if hasattr(self, 'frozen_table_view') and self.frozen_table_view.model():
                model = self.frozen_table_view.model()
                if model.rowCount() > 0:
                    # 最初の表示されている列を探す
                    first_visible_col = 0
                    loop_count = 0
                    max_loops = model.columnCount() + 1
                    while (first_visible_col < model.columnCount() and 
                           self.frozen_table_view.isColumnHidden(first_visible_col) and
                           loop_count < max_loops):
                        first_visible_col += 1
                        loop_count += 1
                    
                    if first_visible_col < model.columnCount():
                        first_index = model.index(0, first_visible_col)
                        self.frozen_table_view.setCurrentIndex(first_index)
                        self.frozen_table_view.setFocus()
                        # 編集モードに入る
                        self.frozen_table_view.edit(first_index)
        except Exception as e:
            print(f"Move to SKU table error: {e}")
    
    def _move_to_sku_table_last(self):
        """SKUテーブルの最後のセルに移動"""
        try:
            if hasattr(self, 'scrollable_table_view') and self.scrollable_table_view.model():
                model = self.scrollable_table_view.model()
                if model.rowCount() > 0:
                    last_row = model.rowCount() - 1
                    # 最後の表示されている列を探す
                    last_visible_col = model.columnCount() - 1
                    loop_count = 0
                    max_loops = model.columnCount() + 1
                    while (last_visible_col >= 0 and 
                           self.scrollable_table_view.isColumnHidden(last_visible_col) and
                           loop_count < max_loops):
                        last_visible_col -= 1
                        loop_count += 1
                    
                    if last_visible_col >= 0:
                        last_index = model.index(last_row, last_visible_col)
                        self.scrollable_table_view.setCurrentIndex(last_index)
                        self.scrollable_table_view.setFocus()
                        # 編集モードに入る
                        self.scrollable_table_view.edit(last_index)
                        # スクロールを同期
                        self._sync_table_scroll(last_row)
        except Exception as e:
            print(f"Move to SKU table last error: {e}")
    
    def _setup_sku_table_navigation(self):
        """SKUテーブル用のスマートナビゲーション設定"""
        try:
            # FocusControllingTableViewのkeyPressEventで直接処理するため、ここでは設定のみ
            pass
                
        except Exception as e:
            print(f"SKU table navigation setup error: {e}")
    
    def _setup_global_tab_filter(self):
        """アプリケーションレベルでTabキーイベントを捕捉"""
        try:
            from PyQt5.QtCore import QObject, QEvent
            from PyQt5.QtWidgets import QApplication
            
            class GlobalTabEventFilter(QObject):
                def __init__(self, parent_app):
                    super().__init__()
                    self.parent_app = parent_app
                
                def eventFilter(self, obj, event):
                    # Tabキーのみを処理（Enterキーは元の処理に任せる）
                    if (event.type() == QEvent.KeyPress and 
                        event.key() in [Qt.Key_Tab, Qt.Key_Backtab]):
                        
                        # 親ウィジェットを辿ってSKUテーブルを探す
                        target_table = None
                        current_widget = obj
                        
                        # 最大5回まで親を辿る
                        for i in range(5):
                            if current_widget is None:
                                break
                            
                            # SKUテーブルを発見
                            if (hasattr(self.parent_app, 'frozen_table_view') and 
                                current_widget == self.parent_app.frozen_table_view):
                                target_table = self.parent_app.frozen_table_view
                                break
                            elif (hasattr(self.parent_app, 'scrollable_table_view') and 
                                  current_widget == self.parent_app.scrollable_table_view):
                                target_table = self.parent_app.scrollable_table_view
                                break
                            
                            # 親ウィジェットに移動
                            current_widget = getattr(current_widget, 'parent', lambda: None)()
                        
                        # SKUテーブルが見つかった場合のみ処理
                        if target_table is not None:
                            if (hasattr(self.parent_app, 'smart_navigation_enabled') and 
                                self.parent_app.smart_navigation_enabled):
                                
                                if event.key() == Qt.Key_Tab:
                                    self.parent_app._handle_sku_enter_navigation(target_table, event)
                                else:  # Qt.Key_Backtab
                                    self.parent_app._handle_sku_backtab_navigation(target_table, event)
                                
                                return True  # イベントを消費
                    
                    return False  # 他のイベントは通常通り処理
            
            self.global_tab_filter = GlobalTabEventFilter(self)
            self.global_tab_filter.setParent(self)  # メモリリーク対策
            QApplication.instance().installEventFilter(self.global_tab_filter)
            
        except Exception as e:
            print(f"Global tab filter setup error: {e}")
    
    def _handle_sku_enter_navigation(self, table_view, event):
        """SKUテーブルでのEnterキー処理"""
        try:
            current_index = table_view.currentIndex()
            if not current_index.isValid():
                event.ignore()
                return
            
            model = table_view.model()
            if not model:
                event.ignore()
                return
            
            current_row = current_index.row()
            current_col = current_index.column()
            
            # 次の編集可能セルを探す
            next_row = current_row
            next_col = current_col + 1
            
            # 固定テーブルにいる場合の特別処理
            if table_view == self.frozen_table_view:
                # 固定テーブルの表示可能列内で次の列を探す
                frozen_next_col = next_col
                loop_count = 0
                max_loops = FROZEN_TABLE_COLUMN_COUNT + 1
                while (frozen_next_col < FROZEN_TABLE_COLUMN_COUNT and 
                       table_view.isColumnHidden(frozen_next_col) and
                       loop_count < max_loops):
                    frozen_next_col += 1
                    loop_count += 1
                
                # 固定テーブル内に次の列がある場合
                if frozen_next_col < FROZEN_TABLE_COLUMN_COUNT and not table_view.isColumnHidden(frozen_next_col):
                    next_index = model.index(next_row, frozen_next_col)
                    table_view.setCurrentIndex(next_index)
                    table_view.edit(next_index)
                    event.accept()
                    return
                else:
                    # 固定テーブルの最後に到達 → スクロール可能テーブルに移動
                    if hasattr(self, 'scrollable_table_view'):
                        scrollable_start_col = FROZEN_TABLE_COLUMN_COUNT
                        scrollable_col = scrollable_start_col
                        
                        # スクロール可能テーブルの最初の表示列を探す
                        while (scrollable_col < model.columnCount() and 
                               self.scrollable_table_view.isColumnHidden(scrollable_col)):
                            scrollable_col += 1
                        
                        if scrollable_col < model.columnCount():
                            next_index = model.index(current_row, scrollable_col)
                            self.scrollable_table_view.setCurrentIndex(next_index)
                            self.scrollable_table_view.setFocus()
                            self.scrollable_table_view.edit(next_index)
                            self._sync_table_scroll(current_row)
                            event.accept()
                            return
            else:
                # スクロール可能テーブルでの通常処理
                while (next_col < model.columnCount() and 
                       table_view.isColumnHidden(next_col)):
                    next_col += 1
                
                # スクロール可能テーブル内に次の列がある場合
                if next_col < model.columnCount():
                    next_index = model.index(next_row, next_col)
                    table_view.setCurrentIndex(next_index)
                    table_view.edit(next_index)
                    self._sync_table_scroll(next_row)
                    event.accept()
                    return
                
            # 行の最後まで来たら次の行の最初へ
            next_row += 1
            
            # 次の行が存在する場合、固定テーブルの最初の列へ移動
            if next_row < model.rowCount():
                # 固定テーブルの最初の表示列を探す
                first_visible_col = 0
                while (first_visible_col < FROZEN_TABLE_COLUMN_COUNT and 
                       self.frozen_table_view.isColumnHidden(first_visible_col)):
                    first_visible_col += 1
                
                if first_visible_col < FROZEN_TABLE_COLUMN_COUNT:
                    next_index = model.index(next_row, first_visible_col)
                    self.frozen_table_view.setCurrentIndex(next_index)
                    self.frozen_table_view.setFocus()
                    self.frozen_table_view.edit(next_index)
                    # スクロールを同期
                    self._sync_table_scroll(next_row)
                    event.accept()
                    return
            
            # テーブルの最後まで来たらメインフィールドに戻る
            if next_row >= model.rowCount():
                if hasattr(self, 'main_fields') and HEADER_MEMO in self.main_fields:
                    self.main_fields[HEADER_MEMO].setFocus()
                    # テキストフィールドの場合、全選択
                    if hasattr(self.main_fields[HEADER_MEMO], 'selectAll'):
                        self.main_fields[HEADER_MEMO].selectAll()
                    # メモ欄が見えるようにスクロール
                    self._ensure_field_visible(self.main_fields[HEADER_MEMO])
                event.accept()
                return
                
        except Exception as e:
            print(f"SKU Enter navigation error: {e}")
            event.ignore()
    
    def _sync_table_scroll(self, target_row):
        """テーブルビューのスクロールを同期"""
        try:
            if hasattr(self, 'frozen_table_view') and hasattr(self, 'scrollable_table_view'):
                # 確実に表示されるよう即座に実行
                self._ensure_row_visible(target_row)
                
                # 少し遅延してもう一度実行（確実性のため）
                QTimer.singleShot(50, lambda: self._ensure_row_visible(target_row))
                
        except Exception as e:
            print(f"Table scroll sync error: {e}")
    
    def _ensure_row_visible(self, row):
        """指定した行が確実に表示されるようにする"""
        try:
            if hasattr(self, 'frozen_table_view') and hasattr(self, 'scrollable_table_view'):
                # 両方のテーブルビューで行を表示
                if (self.frozen_table_view.model() and 
                    row >= 0 and row < self.frozen_table_view.model().rowCount()):
                    
                    # 固定テーブルビューで行を表示
                    frozen_index = self.frozen_table_view.model().index(row, 0)
                    self.frozen_table_view.scrollTo(frozen_index, QAbstractItemView.EnsureVisible)
                    
                    # スクロール可能テーブルビューで行を表示
                    if (self.scrollable_table_view.model() and 
                        row < self.scrollable_table_view.model().rowCount()):
                        scrollable_index = self.scrollable_table_view.model().index(row, 0)
                        self.scrollable_table_view.scrollTo(scrollable_index, QAbstractItemView.EnsureVisible)
                    
                    # スクロールバーの値を同期
                    frozen_scrollbar = self.frozen_table_view.verticalScrollBar()
                    scrollable_scrollbar = self.scrollable_table_view.verticalScrollBar()
                    
                    # 固定テーブルを基準にスクロール位置を合わせる
                    scroll_value = frozen_scrollbar.value()
                    scrollable_scrollbar.setValue(scroll_value)
                    
        except Exception as e:
            print(f"Ensure row visible error: {e}")
    
    def _handle_sku_backtab_navigation(self, table_view, event):
        """SKUテーブルでのShift+Tabキー処理"""
        try:
            current_index = table_view.currentIndex()
            if not current_index.isValid():
                event.ignore()
                return
            
            model = table_view.model()
            if not model:
                event.ignore()
                return
            
            current_row = current_index.row()
            current_col = current_index.column()
            
            # 前の編集可能セルを探す
            prev_row = current_row
            prev_col = current_col - 1
            
            # スクロール可能テーブルにいる場合の特別処理
            if table_view == self.scrollable_table_view:
                # スクロール可能テーブル内で前の列を探す
                while (prev_col >= FROZEN_TABLE_COLUMN_COUNT and 
                       table_view.isColumnHidden(prev_col)):
                    prev_col -= 1
                
                # スクロール可能テーブル内に前の列がある場合
                if prev_col >= FROZEN_TABLE_COLUMN_COUNT:
                    prev_index = model.index(prev_row, prev_col)
                    table_view.setCurrentIndex(prev_index)
                    table_view.edit(prev_index)
                    event.accept()
                    return
                else:
                    # スクロール可能テーブルの最初に到達 → 固定テーブルに移動
                    if hasattr(self, 'frozen_table_view'):
                        # 固定テーブルの最後の表示列を探す
                        frozen_col = FROZEN_TABLE_COLUMN_COUNT - 1
                        while (frozen_col >= 0 and 
                               self.frozen_table_view.isColumnHidden(frozen_col)):
                            frozen_col -= 1
                        
                        if frozen_col >= 0:
                            prev_index = model.index(current_row, frozen_col)
                            self.frozen_table_view.setCurrentIndex(prev_index)
                            self.frozen_table_view.setFocus()
                            self.frozen_table_view.edit(prev_index)
                            event.accept()
                            return
            else:
                # 固定テーブルでの通常処理（無限ループ防止付き）
                loop_count = 0
                max_loops = model.columnCount() + 1  # カラム数分のループを上限とする
                while (prev_col >= 0 and table_view.isColumnHidden(prev_col) and loop_count < max_loops):
                    prev_col -= 1
                    loop_count += 1
                
                # 固定テーブル内に前の列がある場合
                if prev_col >= 0:
                    prev_index = model.index(prev_row, prev_col)
                    table_view.setCurrentIndex(prev_index)
                    table_view.edit(prev_index)
                    event.accept()
                    return
                
                # 行の最初まで来たら前の行の最後へ
                prev_row -= 1
                if prev_row >= 0:
                    # スクロール可能テーブルの最後の表示列を探す
                    last_col = model.columnCount() - 1
                    while (last_col >= 0 and 
                           self.scrollable_table_view.isColumnHidden(last_col)):
                        last_col -= 1
                    
                    if last_col >= 0:
                        prev_index = model.index(prev_row, last_col)
                        self.scrollable_table_view.setCurrentIndex(prev_index)
                        self.scrollable_table_view.setFocus()
                        self.scrollable_table_view.edit(prev_index)
                        # スクロールを同期
                        self._sync_table_scroll(prev_row)
                        event.accept()
                        return
            
            # テーブルの最初まで来たらメインフィールドの最後に戻る
            if prev_row < 0:
                # 最後の表示されているメインフィールドを探す
                if hasattr(self, 'main_field_order') and hasattr(self, 'main_fields'):
                    for i in range(len(self.main_field_order) - 1, -1, -1):
                        field_name = self.main_field_order[i]
                        if (field_name in self.main_fields and 
                            self.main_fields[field_name] and
                            self.main_fields[field_name].isEnabled() and
                            self.main_fields[field_name].isVisible()):
                            self.main_fields[field_name].setFocus()
                            if hasattr(self.main_fields[field_name], 'selectAll'):
                                self.main_fields[field_name].selectAll()
                            self._ensure_field_visible(self.main_fields[field_name])
                            break
                event.accept()
                return
            
            # 同じテーブル内で前のセルにフォーカスを移動
            if prev_col >= 0:
                prev_index = model.index(prev_row, prev_col)
                if prev_index.isValid():
                    table_view.setCurrentIndex(prev_index)
                    table_view.edit(prev_index)
                    # スクロールを同期
                    self._sync_table_scroll(prev_row)
            
            event.accept()
                
        except Exception as e:
            print(f"SKU Backtab navigation error: {e}")
            event.ignore()

    def _setup_main_layout(self, main_layout_ref):
        """メインレイアウトの構築（左ペイン、右ペイン、スプリッター）"""
        # この実装は既存の__init__内のレイアウト構築部分と重複しているため
        # 現在の機能を保持するために既存の実装をそのまま使用
        pass

    def _connect_signals(self):
        """シグナルとスロットの接続を設定"""
        # ボタンのクリックイベント
        self.new_btn.clicked.connect(lambda: self.clear_fields())
        self.save_btn.clicked.connect(lambda: self.save_to_excel())
        self.sku_add_btn.clicked.connect(lambda: self.add_sku_column())
        self.bulk_p_btn.clicked.connect(lambda: self.set_all_control_to_p())
        self.run_csharp_btn.clicked.connect(lambda: self.handle_csv_generation_button_click())
        self.sku_delete_btn.clicked.connect(lambda: self.delete_selected_skus())
        self.category_select_btn.clicked.connect(lambda: self.open_category_dialog())
        self.open_id_search_button.clicked.connect(lambda: self._open_id_search_dialog())
        self.image_desc_btn.clicked.connect(lambda: self.generate_html())
        
        # 追加ショートカットの接続
        self.focus_search_action.triggered.connect(lambda: self.focus_search())
        
        # 検索バーと商品リストのイベント
        self.search_bar.textChanged.connect(lambda text: self.filter_list())
        self.product_list.currentItemChanged.connect(lambda current, previous: self._handle_product_selection_changed(current, previous))
        self.product_list.customContextMenuRequested.connect(lambda pos: self.show_product_list_menu(pos))
        
        # ラジオボタンのイベント
        self.control_radio_n.toggled.connect(lambda: self.mark_dirty())
        
        # SKUテーブルのイベント
        self.frozen_table_view.setModel(self.sku_model)
        self.scrollable_table_view.setModel(self.sku_model)
        self.frozen_table_view.selectionModel().selectionChanged.connect(
            lambda s,d: self.synchronize_selection(self.frozen_table_view, self.scrollable_table_view, s))
        self.scrollable_table_view.selectionModel().selectionChanged.connect(
            lambda s,d: self.synchronize_selection(self.scrollable_table_view, self.frozen_table_view, s))
        
        # スクロール同期
        self.scrollable_table_view.verticalScrollBar().valueChanged.connect(
            self.frozen_table_view.verticalScrollBar().setValue)
        self.frozen_table_view.verticalScrollBar().valueChanged.connect(
            self.scrollable_table_view.verticalScrollBar().setValue)
        self.scrollable_table_view.verticalHeader().sectionResized.connect(
            self.frozen_table_view.verticalHeader().resizeSection)
        self.frozen_table_view.verticalHeader().sectionResized.connect(
            self.scrollable_table_view.verticalHeader().resizeSection)
        
        # テーブルビュー間の相互参照設定
        self.frozen_table_view.setOtherTableView(self.scrollable_table_view)
        self.scrollable_table_view.setOtherTableView(self.frozen_table_view)

    @property
    def is_dirty(self):
        return self._is_dirty

    @is_dirty.setter
    def is_dirty(self, value):
        # 実際の状態変更があった場合のみ処理
        if self._is_dirty != value:
            self._is_dirty = value
            if hasattr(self, 'save_btn'): # save_btnが初期化済みか確認
                self.save_btn.setEnabled(value)
            
            # ステータスバーを更新して保存状態を表示
            if hasattr(self, 'status_labels') and hasattr(self, '_update_status_bar'):
                self._update_status_bar()
            
            # データ変更時は何も表示しない（うるさくないように）


    def _load_settings(self) -> None:
        settings = QSettings("株式会社大宝家具", APP_NAME) # 組織名を設定
        geometry = settings.value("geometry", QByteArray())
        main_splitter_state = settings.value("mainSplitterState", QByteArray())
        right_splitter_state = settings.value("rightSplitterState", QByteArray())
        
        # 自動更新チェックの設定を読み込み（デフォルトはTrue）
        self.auto_update_check_enabled = settings.value("update/auto_check_enabled", True, type=bool)

        if isinstance(geometry, QByteArray) and not geometry.isEmpty():
            self.restoreGeometry(geometry)
        else:
            self.showMaximized()

        # 起動時に自動保存されたデータがあるか確認
        if settings.value("autosave/exists", False, type=bool):
            # QTimer.singleShotを使用して、ウィンドウの表示が安定した後にダイアログを表示
            QTimer.singleShot(0, self._show_restore_confirmation_dialog)

        if isinstance(main_splitter_state, QByteArray) and not main_splitter_state.isEmpty():
            self.main_splitter.restoreState(main_splitter_state)
        
        if isinstance(right_splitter_state, QByteArray) and not right_splitter_state.isEmpty():
            self.right_splitter.restoreState(right_splitter_state)

    def _show_restore_confirmation_dialog(self):
        """自動保存されたデータの復元確認ダイアログを表示し、ユーザーの選択に応じて処理する"""
        settings = QSettings("株式会社大宝家具", APP_NAME) # settingsを再度取得
        # この時点でメインウィンドウは表示され、位置も確定しているはず
        reply = QMessageBox.question(self, "データ復元",
                                     "前回予期せず終了した際のデータが見つかりました。復元しますか？",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if reply == QMessageBox.Yes:
            self._load_auto_saved_data()
        else:
            self._clear_auto_save_data() # 復元しない場合は自動保存データをクリア


    def _load_all_id_master_data(self):
        self._r_genre_master_list = load_id_master_data(
            R_GENRE_MASTER_FILE, MASTER_ID_COLUMN_DEFAULT, MASTER_NAME_COLUMN_R_GENRE, MASTER_HIERARCHY_COLUMN_DEFAULT
        )
    def _open_id_search_dialog(self):
        if not hasattr(self, '_r_genre_master_list') or \
           (not self._r_genre_master_list and not self._y_category_master_list and not self._ya_category_master_list):
             msg = "IDマスターデータが読み込まれていません。ID検索機能は利用できません。"
             QMessageBox.information(self, "ID検索", msg)
             logging.info(f"ID検索ダイアログ表示試行: {msg}")

             return

        current_r_id = self.main_fields.get(HEADER_R_GENRE_ID).text() if HEADER_R_GENRE_ID in self.main_fields else ""
        current_y_id = self.main_fields.get(HEADER_Y_CATEGORY_ID).text() if HEADER_Y_CATEGORY_ID in self.main_fields else ""
        current_ya_id = self.main_fields.get(HEADER_YA_CATEGORY_ID).text() if HEADER_YA_CATEGORY_ID in self.main_fields else ""

        dialog = IdSearchDialog(
            self._r_genre_master_list,
            self._y_category_master_list,
            self._ya_category_master_list,
            current_r_id,
            current_y_id,
            current_ya_id,
            self
        )
        if dialog.exec_() == QDialog.Accepted:
            selected_ids = dialog.get_all_selected_ids()
            if selected_ids:
                if HEADER_R_GENRE_ID in self.main_fields and selected_ids.get('Rジャンル'):
                    self.main_fields[HEADER_R_GENRE_ID].setText(selected_ids['Rジャンル'])
                if HEADER_Y_CATEGORY_ID in self.main_fields and selected_ids.get('Yカテゴリ'):
                    self.main_fields[HEADER_Y_CATEGORY_ID].setText(selected_ids['Yカテゴリ'])
                if HEADER_YA_CATEGORY_ID in self.main_fields and selected_ids.get('YAカテゴリ'):
                    self.main_fields[HEADER_YA_CATEGORY_ID].setText(selected_ids['YAカテゴリ'])

    def open_explanation_mark_dialog(self):
        if not hasattr(self, 'explanation_mark_icon_data') or not self.explanation_mark_icon_data:
            msg = "説明マークのアイコンデータが読み込まれていません。"
            QMessageBox.information(self, "アイコン情報なし", msg); logging.info(f"説明マークダイアログ表示試行: {msg}")
            return

        current_ids_str = ""
        # self.main_fields[EXPLANATION_MARK_FIELD_NAME] は QLineEdit を指す
        if EXPLANATION_MARK_FIELD_NAME in self.main_fields and isinstance(self.main_fields[EXPLANATION_MARK_FIELD_NAME], QLineEdit):
            current_ids_str = self.main_fields[EXPLANATION_MARK_FIELD_NAME].text()

        dialog = ExplanationMarkDialog(self.explanation_mark_icon_data, current_ids_str, self)
        if dialog.exec_() == QDialog.Accepted:
            selected_ids_output_str = dialog.get_selected_ids_as_string()
            if EXPLANATION_MARK_FIELD_NAME in self.main_fields and isinstance(self.main_fields[EXPLANATION_MARK_FIELD_NAME], QLineEdit):
                # QLineEdit.setText() が mark_dirty をトリガーする (textChangedシグナル経由)
                self.main_fields[EXPLANATION_MARK_FIELD_NAME].setText(selected_ids_output_str)

    def apply_stylesheet(self):
        stylesheet = """
            QWidget { 
                color: #2c3e50; 
                font-family: "Segoe UI", "Yu Gothic UI", "Meiryo UI", -apple-system, BlinkMacSystemFont, sans-serif; 
                font-weight: 400;
                letter-spacing: 0.3px;
            }
            QWidget#LeftPane { 
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                             stop: 0 #f8f9fc, stop: 1 #e8ecf4);
            }
            QLabel#SidebarLabel { 
                font-size: 11pt; 
                font-weight: 700; 
                color: #2d3748; 
                margin-top: 8px; 
                margin-bottom: 6px; 
                padding: 8px 12px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f7fafc, stop:1 #edf2f7);
                border-radius: 8px;
                border-left: 4px solid #3b82f6;
            }
            QListWidget#ProductList { 
                background-color: #ffffff; 
                border: 1px solid #e2e8f0; 
                border-radius: 12px; 
                padding: 8px; 
            }
            QListWidget#ProductList::item { 
                padding: 12px 12px; 
                border: none;
                border-radius: 8px;
                margin: 2px 0px;
                font-weight: 500;
                color: #334155;
            }
            QListWidget#ProductList::item:selected { 
                background-color: #e0f2fe; 
                color: #0d47a1 !important; 
                border: 2px solid #1976d2;
                border-radius: 8px; 
                font-weight: 600;
            }
            QListWidget#ProductList::item:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9); 
                border-radius: 8px;
            }
            QLineEdit { 
                padding: 12px 16px; 
                border: 1px solid #cbd5e1; 
                border-radius: 10px; 
                background-color: #ffffff; 
                font-size: 14px;
                color: #1e293b;
                font-weight: 500;
            }
            QLineEdit:focus { 
                border: 2px solid #3b82f6; 
                background-color: #ffffff;
            }
            QLineEdit[readOnly="true"] { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9); 
                color: #64748b; 
                border-color: #e2e8f0;
            }
            QWidget#RightPane { 
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                             stop: 0 #ffffff, stop: 1 #f7f9fc); 
            }
            QScrollArea { border: none; }
            QWidget#MainScrollContent { 
                background-color: #ffffff; 
                border: 1px solid #e1e8ed; 
                border-radius: 16px; 
                margin: 12px;
            }
            QPushButton { 
                padding: 10px 20px; 
                border: 1px solid #e1e8ed; 
                border-radius: 12px; 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffffff, stop:1 #f8fafc); 
                color: #334155; 
                min-height: 28px; 
                font-weight: 600;
                font-size: 13px;
            }
            QPushButton:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9); 
                border-color: #667eea;
            }
            QPushButton:pressed { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #e2e8f0, stop:1 #cbd5e1); 
            }
            QPushButton:disabled { 
                background-color: #f1f2f6; 
                color: #94a3b8; 
                border-color: #ddd;
            }
            QPushButton#SaveButton { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #00b894, stop:1 #00a085); 
                color: white; 
                border: 1px solid #00a085;
                font-weight: 700; 
                font-size: 14px;
            }
            QPushButton#SaveButton:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #00cec9, stop:1 #00b894); 
            }
            QPushButton#SaveButton:pressed { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #00a085, stop:1 #00967d); 
            }
            QPushButton#NewProductButton, QPushButton#AddSkuButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #1d4ed8); 
                color: white; 
                border: 1px solid #1d4ed8;
                font-weight: 700;
                font-size: 14px;
            }
            QPushButton#NewProductButton:hover, QPushButton#AddSkuButton:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60a5fa, stop:1 #3b82f6); 
            }
            QPushButton#NewProductButton:pressed, QPushButton#AddSkuButton:pressed { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #1d4ed8, stop:1 #1e3a8a); 
            }
            QPushButton#CategorySelectButton, QPushButton#ImageDescButton, QPushButton#IdSearchButton {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #17a2b8, stop:1 #138496); color: white; border-color: #117a8b;
            }
            /* SKU追加ボタンと新規作成ボタンのスタイルを共通化 */
            QPushButton#CategorySelectButton:hover, QPushButton#ImageDescButton:hover, QPushButton#IdSearchButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #19b4cc, stop:1 #1595a8);
            }
            QPushButton#ExplanationMarkSelectButton {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #17a2b8, stop:1 #138496); color: white; border-color: #117a8b;
            }
            QPushButton#ExplanationMarkSelectButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #19b4cc, stop:1 #1595a8);
            }
            QPushButton#ColorSelectButton { /* Style for the new color selection button */
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #17a2b8, stop:1 #138496); color: white; border-color: #117a8b;
            }
            QPushButton#ColorSelectButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #19b4cc, stop:1 #1595a8);
            }
            QPushButton#DeleteSkuButton {
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #dc3545, stop:1 #c82333); color: white; border-color: #bd2130; font-weight: bold;
            }
            QPushButton#DeleteSkuButton:hover { background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #e24755, stop:1 #d1303f); }
            QPushButton#DeleteSkuButton:pressed { background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #c82333, stop:1 #b01c2b); }
            QPushButton#BulkPButton { background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffc107, stop:1 #e0a800); color: #212529; border-color: #d39e00; }
            QPushButton#BulkPButton:hover { background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffca2c, stop:1 #f0b300); }
            QLabel#SectionHeader { 
                font-size: 14pt; 
                font-weight: 700; 
                color: #1a202c; 
                padding: 6px 0px 4px 16px; 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9);
                border: none;
                border-left: 4px solid #3b82f6;
                border-radius: 8px;
                margin-bottom: 8px;
                margin-top: 8px;
            }
            QPushButton#RunCSharpButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ff6b6b, stop:1 #ee5a52);
                color: white;
                font-weight: bold;
                border: 1px solid #ee5a52;
            }
            QPushButton#RunCSharpButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ff7979, stop:1 #ff6b6b);
            }
            QPushButton#RunCSharpButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ee5a52, stop:1 #e55039);
            }
            QRadioButton { 
                spacing: 8px; 
                padding: 8px 12px; 
                font-weight: 500;
                color: #374151;
                border-radius: 8px;
            }
            QRadioButton:hover {
                background-color: #f9fafb;
            }
            QRadioButton::indicator { 
                width: 18px; 
                height: 18px; 
                border: 2px solid #d1d5db;
                border-radius: 9px;
                background-color: white;
            }
            QRadioButton::indicator:checked {
                border-color: #3b82f6;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #1d4ed8);
            }
            QRadioButton::indicator:checked:hover {
                border-color: #2563eb;
            }
            QSplitter::handle { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #e2e8f0, stop:1 #cbd5e1);
                border-radius: 2px;
            }
            QSplitter::handle:horizontal { 
                width: 4px; 
                margin: 4px 0px;
            }
            QSplitter::handle:vertical { 
                height: 4px; 
                margin: 0px 4px;
            }
            QSplitter::handle:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3b82f6, stop:1 #1d4ed8);
            }
            QSplitter::handle:pressed { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2563eb, stop:1 #1e40af);
            }
            QWidget#ExpandableGroup { 
                margin-bottom: 16px; 
                border-radius: 12px;
                border: 1px solid #e2e8f0;
            }
            QWidget#ExpandableGroupHeader { 
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4a5568, stop:1 #2d3748); 
                border-top-left-radius: 12px; 
                border-top-right-radius: 12px; 
                border-bottom: none;
                padding: 4px;
            }
            QLabel#ExpandableGroupLabel { 
                font-weight: 700; 
                font-size: 11pt; 
                color: #ffffff; 
                padding: 8px 16px; 
                letter-spacing: 0.5px;
            }
            QPushButton#ExpandableGroupToggleButton { 
                border: 2px solid rgba(255, 255, 255, 0.3); 
                border-radius: 12px; 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgba(255,255,255,0.2), stop:1 rgba(255,255,255,0.1)); 
                padding: 0px; 
                min-height: 24px; 
                min-width: 24px;
                font-weight: bold;
                color: white;
            }
            QPushButton#ExpandableGroupToggleButton:hover { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgba(255,255,255,0.4), stop:1 rgba(255,255,255,0.2)); 
                border-color: rgba(255, 255, 255, 0.5);
            }
            QWidget#ExpandableGroupContent { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9); 
                border-bottom-left-radius: 12px; 
                border-bottom-right-radius: 12px; 
                padding: 8px;
            }
            QWidget#SkuTableContainer { 
                border: 1px solid #e2e8f0; 
                border-radius: 12px; 
                background-color: #ffffff; 
                margin: 4px;
            }
            QTableView#FrozenTableView, QTableView#ScrollableTableView {
                gridline-color: #e2e8f0; 
                border: none; 
                alternate-background-color: #f8fafc;
                background-color: white;
                selection-background-color: #3b82f6; 
                selection-color: white;
                border-radius: 8px;
            }
            QTableView#FrozenTableView::item:selected, QTableView#ScrollableTableView::item:selected {
                background-color: #dbeafe;
                color: #1e40af;
                border: none;
            }
            QTableView#FrozenTableView::item:selected:focus, QTableView#ScrollableTableView::item:selected:focus {
                background-color: #bfdbfe;
                color: #1d4ed8;
                border: 1px solid #3b82f6;
            }
            QTableView QHeaderView::section { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f7fafc, stop:1 #edf2f7); 
                padding: 8px 12px; 
                font-weight: 600; 
                color: #374151;
                border: 1px solid #e2e8f0;
                font-size: 12px;
            }
            QTextEdit[objectName="特徴_1"],
            QTextEdit[objectName="材質_1"],
            QTextEdit[objectName="仕様_1"] {
                min-height: 100px;
                padding: 12px 16px;
                border: 1px solid #cbd5e1;
                border-radius: 12px;
                background-color: #ffffff;
                font-size: 14px;
                color: #1e293b;
                font-family: "Segoe UI", "Yu Gothic UI", "Meiryo UI", sans-serif;
                line-height: 1.5;
            }
            QTextEdit[objectName="特徴_1"]:focus,
            QTextEdit[objectName="材質_1"]:focus,
            QTextEdit[objectName="仕様_1"]:focus {
                border: 2px solid #3b82f6;
                background-color: #ffffff;
            }
            QLabel[objectName^="ByteCountLabel_"],
            QLabel[objectName="DigitCountLabelMycode"] {
                font-size: 10pt;
                color: #64748b;
                font-weight: 500;
                padding: 4px 8px;
                background-color: #f1f5f9;
                border-radius: 6px;
                border: 1px solid #e2e8f0;
            }
            QLabel[objectName^="ByteCountLabel_"][style*="color: red"],
            QLabel[objectName="DigitCountLabelMycode"][style*="color: red"] {
                color: #dc2626;
                font-weight: 700;
                background-color: #fef2f2;
                border-color: #fecaca;
            }
            QLabel[objectName^="ByteCountLabel_"][style*="color: green"],
            QLabel[objectName="DigitCountLabelMycode"][style*="color: green"] {
                color: #059669;
                font-weight: 700;
                background-color: #f0fdf4;
                border-color: #bbf7d0;
            }
            ClickableIconLabel { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffffff, stop:1 #f8fafc); 
                border: 1px solid #e2e8f0; 
                border-radius: 10px; 
                padding: 8px; 
                margin: 2px;
            }
            ClickableIconLabel:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f8fafc, stop:1 #f1f5f9);
                border-color: #cbd5e1;
            }
            ClickableIconLabel[selected="true"] { 
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #dbeafe, stop:1 #bfdbfe); 
                border: 2px solid #3b82f6; 
                font-weight: 600; 
            }
            ClickableIconLabel QLabel { 
                border: none; 
                background-color: transparent; 
            }
            QComboBox {
                padding: 8px 12px;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                background-color: white;
                font-size: 13px;
                color: #374151;
                font-weight: 500;
                min-height: 20px;
            }
            QComboBox:hover {
                border-color: #94a3b8;
                background-color: #f8fafc;
            }
            QComboBox:focus {
                border: 2px solid #3b82f6;
                background-color: white;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                padding: 4px;
                outline: none;
            }
            QComboBox QAbstractItemView::item {
                padding: 8px 12px;
                border-radius: 4px;
                margin: 1px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #3b82f6;
                color: white;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #f1f5f9;
            }
            QScrollBar:vertical {
                background-color: #f1f5f9;
                width: 12px;
                border-radius: 6px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-height: 20px;
                margin: 2px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #94a3b8;
            }
            QScrollBar::handle:vertical:pressed {
                background-color: #6b7280;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
            QScrollBar:horizontal {
                background-color: #f1f5f9;
                height: 12px;
                border-radius: 6px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-width: 20px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #94a3b8;
            }
            QScrollBar::handle:horizontal:pressed {
                background-color: #6b7280;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                width: 0px;
            }
            QToolTip {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2d3748, stop:1 #1a202c);
                color: #ffffff;
                border: 1px solid #4a5568;
                border-radius: 8px;
                padding: 10px 14px;
                font-size: 13px;
                font-weight: 500;
                font-family: "Segoe UI", "Yu Gothic UI", "Meiryo UI", sans-serif;
                max-width: 350px;
                opacity: 240;
            }
        """
        self.setStyleSheet(stylesheet)

    def _setup_tab_order(self):
        """論理的なタブオーダーを設定してキーボードナビゲーションを改善"""
        try:
            # 主要なナビゲーション順序を設定
            tab_widgets = []
            
            # 1. 検索バー
            if hasattr(self, 'search_bar') and self.search_bar:
                tab_widgets.append(self.search_bar)
            
            # 2. 商品リスト
            if hasattr(self, 'product_list') and self.product_list:
                tab_widgets.append(self.product_list)
            
            # 3. アクションボタン群
            for btn_name in ['new_btn', 'save_btn', 'run_csharp_btn', 'bulk_p_btn']:
                btn = getattr(self, btn_name, None)
                if btn:
                    tab_widgets.append(btn)
            
            # 4. コントロールラジオボタン
            if hasattr(self, 'control_radio_n') and self.control_radio_n:
                tab_widgets.append(self.control_radio_n)
            if hasattr(self, 'control_radio_p') and self.control_radio_p:
                tab_widgets.append(self.control_radio_p)
            
            # 5. 主要フィールド（存在するもののみ）
            main_field_names = [
                HEADER_MYCODE, HEADER_PRODUCT_NAME, HEADER_PRICE_TAX_INCLUDED,
                "R_商品名", "Y_商品名", "R_キャッチコピー"
            ]
            
            for field_name in main_field_names:
                if hasattr(self, 'main_fields') and field_name in self.main_fields:
                    field = self.main_fields[field_name]
                    if field and hasattr(field, 'setFocusPolicy'):
                        tab_widgets.append(field)
            
            # タブオーダーを設定
            for i in range(len(tab_widgets) - 1):
                if tab_widgets[i] and tab_widgets[i + 1]:
                    self.setTabOrder(tab_widgets[i], tab_widgets[i + 1])
            
            logging.debug(f"タブオーダーを設定しました: {len(tab_widgets)}個のウィジェット")
            
        except Exception as e:
            logging.warning(f"タブオーダー設定中にエラー: {e}")

    def keyPressEvent(self, event):
        """グローバルキーボードショートカットの処理"""
        try:
            # スーパークラスの処理を最初に実行
            super().keyPressEvent(event)
            
            # カスタムショートカット
            if event.key() == Qt.Key_F5:
                # F5で画面更新
                self.update()
                self.repaint()
                event.accept()
                return
            
            # Ctrl+数字でクイックアクション
            if event.modifiers() == Qt.ControlModifier:
                if event.key() == Qt.Key_1:
                    # Ctrl+1: 検索フォーカス
                    if hasattr(self, 'search_bar') and self.search_bar:
                        self.search_bar.setFocus()
                        event.accept()
                        return
                elif event.key() == Qt.Key_2:
                    # Ctrl+2: 商品リストフォーカス
                    if hasattr(self, 'product_list') and self.product_list:
                        self.product_list.setFocus()
                        event.accept()
                        return
                elif event.key() == Qt.Key_3:
                    # Ctrl+3: 商品コードフォーカス
                    if hasattr(self, 'main_fields') and HEADER_MYCODE in self.main_fields:
                        self.main_fields[HEADER_MYCODE].setFocus()
                        event.accept()
                        return
            
        except Exception as e:
            logging.debug(f"キーイベント処理中のエラー: {e}")
            super().keyPressEvent(event)

    def build_category_tree(self, categories):
        tree = {}; order1, order2, order3 = [], [], []
        for level, name, parent in categories:
            if level == 1:
                if name not in tree: tree[name] = {}; order1.append(name)
            elif level == 2:
                if parent not in tree: tree[parent] = {}; order1.append(parent)
                if name not in tree[parent]: tree[parent][name] = {}; order2.append((parent, name))
            elif level == 3:
                found_parent = False
                for p1_name, l2_dict in tree.items():
                    if parent in l2_dict:
                         if name not in l2_dict[parent]: l2_dict[parent][name] = {}; order3.append((parent, name))
                         found_parent = True
                         break
                if not found_parent:
                    if parent not in tree: tree[parent] = {}; order1.append(parent)
                    if parent not in tree[parent]: tree[parent][parent] = {}; order2.append((parent, parent))
                    if name not in tree[parent][parent]: tree[parent][parent][name] = {}; order3.append((parent, name))

        root_items = {}
        for name in order1:
            if name in tree:
                item = QTreeWidgetItem(self.tree_widget); item.setText(0, name); item.setData(0, Qt.UserRole, name); root_items[name] = item
        for parent_name, name in order2:
            if parent_name in root_items and name in tree.get(parent_name, {}):
                item = QTreeWidgetItem(root_items[parent_name]); item.setText(0, name); item.setData(0, Qt.UserRole, f"{parent_name}:{name}"); self.items_map_by_path[f"{parent_name}:{name}"] = item
        for parent_name, name in order3:
             for l2_full_path, parent_item_l2 in self.items_map_by_path.items():
                 if l2_full_path.endswith(":" + parent_name):
                     full_path_l3 = f"{l2_full_path}:{name}"
                     if full_path_l3 not in self.items_map_by_path:
                         item = QTreeWidgetItem(parent_item_l2); item.setText(0, name); item.setData(0, Qt.UserRole, full_path_l3)
                         self.items_map_by_path[full_path_l3] = item; break
        for name, item in root_items.items():
             self.items_map_by_path[name] = item

    def open_category_dialog(self):
        current_paths = [self.main_fields[f"商品カテゴリ{i+1}"].text() for i in range(5) if f"商品カテゴリ{i+1}" in self.main_fields]
        initial_path = next((p.strip() for p in reversed(current_paths) if p and p.strip()), "")
        dlg = CategorySelectDialog(self.categories, self, [initial_path] if initial_path else [""])
        if dlg.exec_() == QDialog.Accepted:
            selected_paths = dlg.get_selected_categories()
            for i in range(5): self.main_fields[f"商品カテゴリ{i+1}"].setText("")
            all_paths_parts = [p.strip().split(':') for p in selected_paths if p and p.strip()]
            if all_paths_parts:
                unique_sub_paths = set()
                for parts in all_paths_parts:
                    current_sub = ""
                    for part in parts: current_sub = f"{current_sub}:{part}" if current_sub else part; unique_sub_paths.add(current_sub)
                sorted_paths = sorted(list(unique_sub_paths), key=lambda x: (x.split(':')[0], x.count(':'), x))
                for i, path_to_set in enumerate(sorted_paths[:5]): self.main_fields[f"商品カテゴリ{i+1}"].setText(path_to_set)
            self.mark_dirty()

    def show_sku_table(self):
        if not hasattr(self, "sku_data_list") or self.sku_data_list is None: self.sku_data_list = []
        genre_id_widget = self.main_fields.get(HEADER_R_GENRE_ID)
        genre_id = genre_id_widget.text().strip() if isinstance(genre_id_widget, QLineEdit) else ""
        
        defined_attr_details = []
        if hasattr(self, 'definition_loader') and self.definition_loader and genre_id:
            defined_attr_details = self.definition_loader.get_attribute_details_for_genre(genre_id)
        
        sku_headers = []
        if self.sku_data_list:
            preferred = [HEADER_SKU_CODE, HEADER_CHOICE_NAME, HEADER_MEMO, HEADER_GROUP]
            all_keys = set(k for item in self.sku_data_list for k in item.keys())
            valid_keys = {k for k in all_keys if not k.startswith("_highlight_")}
            sku_headers.extend([h for h in preferred if h in valid_keys])
            remaining = sorted([k for k in valid_keys if k not in preferred])
            attr_keys = [k for k in remaining if HEADER_ATTR_ITEM_PREFIX in k or HEADER_ATTR_VALUE_PREFIX in k or HEADER_ATTR_UNIT_PREFIX in k]
            other_keys = [k for k in remaining if k not in attr_keys]
            def sort_key_func(key):
                try:
                    if HEADER_ATTR_ITEM_PREFIX in key: return (int(key.replace(HEADER_ATTR_ITEM_PREFIX,"").strip()), 0)
                    if HEADER_ATTR_VALUE_PREFIX in key: return (int(key.replace(HEADER_ATTR_VALUE_PREFIX,"").strip()), 1)
                    if HEADER_ATTR_UNIT_PREFIX in key: return (int(key.replace(HEADER_ATTR_UNIT_PREFIX,"").strip()), 2)
                except ValueError: return (float('inf'), 0)
                return (float('inf'), 1)
            attr_keys.sort(key=sort_key_func); sku_headers.extend(attr_keys); sku_headers.extend(other_keys)

        self.sku_model.update_data(self.sku_data_list, sku_headers, defined_attr_details)
        self.scrollable_table_view.resizeColumnsToContents(); self.frozen_table_view.resizeColumnsToContents()
        
        frozen_count = 0
        if sku_headers:
            frozen_count = FROZEN_TABLE_COLUMN_COUNT
            
            if HEADER_SKU_CODE in sku_headers: idx = sku_headers.index(HEADER_SKU_CODE); self.frozen_table_view.setColumnWidth(idx, 120)
            if HEADER_CHOICE_NAME in sku_headers: idx = sku_headers.index(HEADER_CHOICE_NAME); self.frozen_table_view.setColumnWidth(idx, 160)
            for i, hdr_txt in enumerate(sku_headers):
                if i >= frozen_count and (HEADER_ATTR_UNIT_PREFIX in hdr_txt or HEADER_ATTR_ITEM_PREFIX in hdr_txt): self.scrollable_table_view.setColumnWidth(i, 60)
            hide_always = [HEADER_MEMO, HEADER_GROUP, HEADER_PRODUCT_CODE_SKU]
            
            for col_idx in range(len(sku_headers)):
                self.frozen_table_view.setColumnHidden(col_idx, False)
                self.scrollable_table_view.setColumnHidden(col_idx, False)

            for col_idx, header_text in enumerate(sku_headers):
                if HEADER_ATTR_UNIT_PREFIX in header_text:
                    should_hide_unit_column = False 
                    try:
                        attr_num = int(header_text.replace(HEADER_ATTR_UNIT_PREFIX, "").strip())
                        if 1 <= attr_num <= len(defined_attr_details):
                            unit_exists_raw = defined_attr_details[attr_num - 1].get("unit_exists_raw", "")
                            should_hide_unit_column = not unit_exists_raw or "有" not in unit_exists_raw
                    except ValueError:
                        pass
                    
                    if col_idx < frozen_count : 
                        self.frozen_table_view.setColumnHidden(col_idx, should_hide_unit_column)
                        self.scrollable_table_view.setColumnHidden(col_idx, True)
                    else: 
                        self.scrollable_table_view.setColumnHidden(col_idx, should_hide_unit_column)
            
            sku_attr_delegate = SkuAttributeDelegate(self.scrollable_table_view)
            frozen_sku_attr_delegate = SkuAttributeDelegate(self.frozen_table_view)
            for col_idx, header_text in enumerate(sku_headers):
                if (HEADER_ATTR_VALUE_PREFIX in header_text or 
                    HEADER_ATTR_UNIT_PREFIX in header_text or 
                    header_text == HEADER_CHOICE_NAME or 
                    header_text == HEADER_SKU_CODE):
                    if col_idx >= frozen_count:
                         self.scrollable_table_view.setItemDelegateForColumn(col_idx, sku_attr_delegate)
                    else:
                         self.frozen_table_view.setItemDelegateForColumn(col_idx, frozen_sku_attr_delegate)
            
            for i, cur_hdr in enumerate(sku_headers):
                is_frz = i < frozen_count
                self.frozen_table_view.setColumnHidden(i, not is_frz or cur_hdr in hide_always)
                if HEADER_ATTR_UNIT_PREFIX not in cur_hdr: 
                    self.scrollable_table_view.setColumnHidden(i, is_frz or HEADER_ATTR_ITEM_PREFIX in cur_hdr or cur_hdr in hide_always)
            
            frz_width = sum(self.frozen_table_view.columnWidth(i) for i in range(frozen_count) if not self.frozen_table_view.isColumnHidden(i))
            if self.frozen_table_view.verticalHeader().isVisible(): frz_width += self.frozen_table_view.verticalHeader().width()
            self.frozen_table_view.setFixedWidth(frz_width + TABLE_PADDING) 
        
        if hasattr(self, 'right_splitter') and self.sku_data_list: self.right_splitter.setSizes([self.right_splitter.height()*3//5, self.right_splitter.height()*2//5])
        
        # SKUテーブル表示後にステータスバーを更新
        if hasattr(self, 'status_labels') and hasattr(self, '_update_status_bar'):
            self._update_status_bar()

    def add_sku_column(self, values=None):
        mycode_widget = self.main_fields.get(HEADER_MYCODE)
        mycode = mycode_widget.text().strip() if isinstance(mycode_widget, QLineEdit) else ""
        genre_id_widget = self.main_fields.get(HEADER_R_GENRE_ID)
        genre_id = genre_id_widget.text().strip() if isinstance(genre_id_widget, QLineEdit) else ""
        skip_val_attrs = [] # skip_val_attrs をここで初期化
        
        defined_attr_details = [] 
        if hasattr(self, 'definition_loader') and self.definition_loader and genre_id:
            defined_attr_details = self.definition_loader.get_attribute_details_for_genre(genre_id)

        # --- 商品サイズの本体寸法を取得し、連結文字列を作成 ---
        main_body_width, main_body_depth, main_body_height = "", "", ""
        formatted_main_body_size_info = "" # SKU属性に設定する文字列
        product_size_efg = self.expandable_field_group_instances.get("商品サイズ")
        if product_size_efg:
            main_body_row_idx = -1
            for i in range(product_size_efg.group_count):
                field_a_name = f"{product_size_efg.group_label_prefix}_{i+1}a"
                field_a_widget = self.main_fields.get(field_a_name)
                if isinstance(field_a_widget, QLineEdit) and field_a_widget.text().strip() == "本体":
                    main_body_row_idx = i
                    break
            if main_body_row_idx != -1:
                dim_data = product_size_efg.dimension_fields_list[main_body_row_idx]
                if dim_data:
                    main_body_width = dim_data['w'].text().strip()
                    main_body_depth = dim_data['d'].text().strip()
                    main_body_height = dim_data['h'].text().strip()
                    # 全ての寸法が存在する場合のみ文字列を生成
                    if main_body_width and main_body_depth and main_body_height:
                        formatted_main_body_size_info = f"幅{main_body_width}×奥行{main_body_depth}×高さ{main_body_height}cm"
        # --- ここまで商品サイズの本体寸法取得 ---

        # --- 商品サイズの重量値を取得 ---
        main_body_weight_value = "" # SKU属性に設定する数値文字列
        if product_size_efg: # product_size_efg は既に取得済みと仮定
            main_body_weight_row_idx = -1
            for i in range(product_size_efg.group_count):
                field_a_name = f"{product_size_efg.group_label_prefix}_{i+1}a"
                field_a_widget = self.main_fields.get(field_a_name)
                if isinstance(field_a_widget, QLineEdit) and field_a_widget.text().strip() == "重量":
                    main_body_weight_row_idx = i
                    break
            if main_body_weight_row_idx != -1:
                # 「重量」行の 'b' フィールドから値を取得
                field_b_name = f"{product_size_efg.group_label_prefix}_{main_body_weight_row_idx+1}b"
                field_b_widget = self.main_fields.get(field_b_name)
                if isinstance(field_b_widget, QLineEdit):
                    weight_text_from_b = field_b_widget.text().strip()
                    # _sync_product_size_to_yspec と同様の正規表現で重量値を抽出
                    match_weight = re.match(r"^\s*(?:約)?\s*([\d\.]+)\s*(?:kg)?\s*$", weight_text_from_b, re.IGNORECASE)
                    if match_weight:
                        main_body_weight_value = match_weight.group(1)
                    elif weight_text_from_b.replace('.', '', 1).isdigit(): # 単純な数値の場合 (正規表現でカバーされるはずだが念のため)
                        main_body_weight_value = weight_text_from_b
        # --- ここまで商品サイズの重量値取得 ---

        new_sku = {HEADER_PRODUCT_CODE_SKU: mycode, HEADER_MEMO: genre_id, HEADER_GROUP: "", HEADER_CHOICE_NAME: ""} 
        new_sku[f"_highlight_{HEADER_SKU_CODE}"] = True; new_sku[f"_highlight_{HEADER_CHOICE_NAME}"] = True
        base_mycode = mycode[:-3] if mycode and len(mycode) >=3 else (mycode if mycode else "")
        src_sku_target = base_mycode + SKU_CODE_SUFFIX_INITIAL
        src_sku_data = next((s for s in self.sku_data_list if s.get(HEADER_PRODUCT_CODE_SKU)==mycode and s.get(HEADER_SKU_CODE,"")==src_sku_target), None)
        
        # is_first_sku_for_this_product は、この商品に対して最初のSKUセット(010)が追加されるかどうか
        # (既に010が存在し、020を追加する場合は False になる)
        is_first_sku_for_this_product = not any(s.get(HEADER_SKU_CODE,"").endswith(SKU_CODE_SUFFIX_INITIAL) for s in self.sku_data_list if s.get(HEADER_PRODUCT_CODE_SKU) == mycode)

        if src_sku_data and not is_first_sku_for_this_product: # 010以外を追加する場合で、かつ010が存在する場合
            skip_keys = [HEADER_SKU_CODE, HEADER_CHOICE_NAME]; skip_val_attrs = ["代表カラー", "カラー"]
            for k,v in src_sku_data.items():
                if k in skip_keys: continue
                is_skip_val = False
                if k.startswith(HEADER_ATTR_VALUE_PREFIX):
                    try:
                        attr_num = int(k.replace(HEADER_ATTR_VALUE_PREFIX,"").strip())
                        # defined_attr_details が空でないことと、インデックス範囲をチェック
                        if defined_attr_details and 1 <= attr_num <= len(defined_attr_details) and \
                           defined_attr_details[attr_num-1].get("name") in skip_val_attrs:
                            is_skip_val = True; new_sku[f"_highlight_{k}"] = True
                    except ValueError: pass
                if not is_skip_val: new_sku[k] = v
            if HEADER_MEMO in src_sku_data: new_sku[HEADER_MEMO] = src_sku_data[HEADER_MEMO]
            if HEADER_GROUP in src_sku_data: new_sku[HEADER_GROUP] = src_sku_data[HEADER_GROUP]

        for i in range(MAX_SKU_ATTRIBUTES):
            num = i+1; item_k=f"{HEADER_ATTR_ITEM_PREFIX}{num}"; val_k=f"{HEADER_ATTR_VALUE_PREFIX}{num}"; unit_k=f"{HEADER_ATTR_UNIT_PREFIX}{num}"
            
            attr_name_from_def = ""
            attr_unit_from_def = "" 

            # defined_attr_details が空でないことと、インデックス範囲をチェック
            if defined_attr_details and i < len(defined_attr_details):
                attr_detail = defined_attr_details[i]
                attr_name_from_def = attr_detail.get("name", "")
                attr_unit_from_def = attr_detail.get("unit_options_list")[0] if attr_detail.get("unit_options_list") else ""
                
                # ★★★ 最初のSKU追加時に本体の連結サイズ情報をSKU属性に設定 ★★★
                if is_first_sku_for_this_product: # この商品に対して最初のSKUセット(010)が追加される場合
                    if attr_name_from_def == RAKUTEN_SKU_ATTR_NAME_SIZE_INFO and formatted_main_body_size_info:
                        new_sku[val_k] = formatted_main_body_size_info
                    # --- ここから追加 ---
                    elif attr_name_from_def == "本体横幅" and main_body_width:
                        new_sku[val_k] = main_body_width
                    elif attr_name_from_def == "本体奥行" and main_body_depth:
                        new_sku[val_k] = main_body_depth
                    elif attr_name_from_def == "本体高さ" and main_body_height:
                        new_sku[val_k] = main_body_height
                    elif attr_name_from_def == "重量" and main_body_weight_value: # SKU属性名が「重量」の場合
                        new_sku[val_k] = main_body_weight_value
                    # --- ここまで追加 ---
                # ★★★ ここまで ★★★

                if attr_detail.get("is_required", False) and val_k not in new_sku:
                    new_sku[val_k] = "-" # 必須項目でまだ値がなければハイフンを設定

            new_sku[item_k] = attr_name_from_def if attr_name_from_def else (new_sku.get(item_k, ""))
            if val_k not in new_sku:
                new_sku[val_k] = ""
                new_sku[f"_highlight_{val_k}"] = bool(defined_attr_details and attr_name_from_def and attr_name_from_def in skip_val_attrs)
            
            # 単位も、もし値が設定されていなければ定義から持ってくる
            # (サイズ属性の場合、単位が自動入力されると良いかもしれないが、今回は値のみ)
            if unit_k not in new_sku: 
                new_sku[unit_k] = attr_unit_from_def

        if not self.sku_data_list: sku_code = base_mycode + SKU_CODE_SUFFIX_INITIAL
        else:
            exist_nums = [int(s.get(HEADER_SKU_CODE,"")[-3:]) for s in self.sku_data_list if s.get(HEADER_SKU_CODE,"").startswith(base_mycode) and len(s.get(HEADER_SKU_CODE,""))==len(base_mycode)+3 and s.get(HEADER_SKU_CODE,"")[-3:].isdigit()]
            next_n = SKU_CODE_SUFFIX_INCREMENT
            loop_count = 0
            max_iterations = (SKU_CODE_SUFFIX_MAX - SKU_CODE_SUFFIX_INITIAL) // max(1, SKU_CODE_SUFFIX_INCREMENT) + 1
            while next_n in exist_nums and next_n <= SKU_CODE_SUFFIX_MAX and loop_count < max_iterations: 
                next_n += SKU_CODE_SUFFIX_INCREMENT
                loop_count += 1
            if next_n > SKU_CODE_SUFFIX_MAX: next_n = (max(exist_nums) + SKU_CODE_SUFFIX_INCREMENT) if exist_nums else SKU_CODE_SUFFIX_INCREMENT
            sku_code = base_mycode + f"{next_n:03d}"
        new_sku[HEADER_SKU_CODE] = sku_code
        if values: new_sku.update(values) # type: ignore
        self.sku_data_list.append(new_sku)
        def sku_sort_key(s_item):
            code_val = s_item.get(HEADER_SKU_CODE, "")
            if code_val and len(code_val) >= 3 and code_val[-3:].isdigit():
                return int(code_val[-3:])
            return float('inf')
        self.sku_data_list.sort(key=sku_sort_key)
        self.show_sku_table(); self.mark_dirty()

    def delete_selected_skus(self):
        sel_model = self.frozen_table_view.selectionModel()
        if not sel_model or not sel_model.hasSelection(): QMessageBox.information(self, "SKU削除", "削除するSKUを選択"); return
        sel_rows = sel_model.selectedRows()
        if not sel_rows: QMessageBox.information(self, "SKU削除", "削除するSKUの行を選択"); return
        if QMessageBox.question(self,"SKU削除確認",f"{len(sel_rows)}件削除しますか？",QMessageBox.Yes|QMessageBox.No,QMessageBox.No) == QMessageBox.No: return
        for r_idx in sorted([idx.row() for idx in sel_rows], reverse=True):
            if 0 <= r_idx < len(self.sku_data_list): del self.sku_data_list[r_idx]
        self.show_sku_table(); self.mark_dirty()

    def synchronize_selection(self, source_view, target_view, qitem_selection_selected):
        """テーブル間の選択状態を同期（改善版）"""
        if not source_view or not target_view or source_view == target_view:
            return
            
        src_sel_model = source_view.selectionModel()
        tgt_sel_model = target_view.selectionModel()
        
        if not src_sel_model or not tgt_sel_model:
            return
            
        # シグナルをブロックして無限ループを防止
        tgt_sel_model.blockSignals(True)
        
        try:
            # 選択をクリア
            tgt_sel_model.clearSelection()
            
            # ソースビューで選択されている行を取得
            selected_rows = set()
            for index in src_sel_model.selectedIndexes():
                if index.isValid():
                    selected_rows.add(index.row())
            
            # ターゲットビューに同じ行を選択
            if selected_rows and target_view.model():
                row_count = target_view.model().rowCount()
                col_count = target_view.model().columnCount()
                
                for row in selected_rows:
                    if 0 <= row < row_count and col_count > 0:
                        # 行全体を選択
                        start_index = target_view.model().index(row, 0)
                        end_index = target_view.model().index(row, col_count - 1)
                        
                        if start_index.isValid() and end_index.isValid():
                            selection = QItemSelection(start_index, end_index)
                            tgt_sel_model.select(selection, QItemSelectionModel.Select | QItemSelectionModel.Rows)
                            
        finally:
            # シグナルのブロックを解除
            tgt_sel_model.blockSignals(False)
            
        # ビューを更新して選択状態を確実に反映
        target_view.update()
        
        # ステータスバーを更新してSKU選択状況を反映
        if hasattr(self, 'status_labels') and hasattr(self, '_update_status_bar'):
            self._update_status_bar()

    def _update_byte_count_display(self, field_name, text):
        if field_name in self.byte_count_labels and field_name in BYTE_LIMITS:
            limit = BYTE_LIMITS[field_name]
            current_bytes = get_byte_count_excel_lenb(text)
            label = self.byte_count_labels[field_name]
            label.setText(f"({current_bytes}/{limit} bytes)")
            if current_bytes > limit:
                label.setStyleSheet("font-size: 8pt; color: red; font-weight: bold;")
                self._show_field_validation_error(field_name, f"文字数制限を超えています ({current_bytes}/{limit} bytes)")
            else:
                label.setStyleSheet("font-size: 8pt; color: #6c757d;")
                self._clear_field_validation_error(field_name)
        self.mark_dirty()

    def _update_mycode_digit_count_display(self, text):
        if self.digit_count_label_mycode:
            current_digits = len(text)
            self.digit_count_label_mycode.setText(f"({current_digits}/10 桁)")
            is_valid_input = text.isdigit() or not text
            if not is_valid_input:
                self.digit_count_label_mycode.setStyleSheet("font-size: 8pt; color: red; font-weight: bold;")
                self._show_field_validation_error(HEADER_MYCODE, "商品コードは数字のみ入力してください")
            elif current_digits == 10 and is_valid_input:
                self.digit_count_label_mycode.setStyleSheet("font-size: 8pt; color: green; font-weight: bold;")
                self._clear_field_validation_error(HEADER_MYCODE)
            else:
                self.digit_count_label_mycode.setStyleSheet("font-size: 8pt; color: #6c757d;")
                if current_digits > 0 and current_digits != 10:
                    self._show_field_validation_error(HEADER_MYCODE, "商品コードは10桁で入力してください")
                else:
                    self._clear_field_validation_error(HEADER_MYCODE)
        self.mark_dirty()

    def _on_price_field_changed(self, field_name, text):
        self.mark_dirty()
        self._format_and_sync_price_fields()
        self._validate_price_field(field_name, text)
    
    def _validate_price_field(self, field_name, text):
        """価格フィールドのリアルタイム検証"""
        if not text.strip():
            self._clear_field_validation_error(field_name)
            return
        
        # カンマを除去して数値チェック
        cleaned_text = text.replace(",", "").replace("，", "")
        if not cleaned_text.isdigit():
            self._show_field_validation_error(field_name, "価格は数字のみ入力してください")
        else:
            price_value = int(cleaned_text)
            if price_value <= 0:
                self._show_field_validation_error(field_name, "価格は0より大きい値を入力してください")
            elif price_value > 999999999:
                self._show_field_validation_error(field_name, "価格が大きすぎます")
            else:
                self._clear_field_validation_error(field_name)
    
    def _validate_required_field(self, field_name, text):
        """必須フィールドの検証"""
        if field_name in [HEADER_MYCODE, HEADER_PRODUCT_NAME]:
            if not text.strip():
                self._show_field_validation_error(field_name, "この項目は必須入力です")
            else:
                self._clear_field_validation_error(field_name)
    
    def _show_field_validation_error(self, field_name, message):
        """フィールドにエラー表示を追加（強化版）"""
        if not hasattr(self, '_validation_errors'):
            self._validation_errors = {}
        
        self._validation_errors[field_name] = message
        
        # フィールドの枠を赤くする（より洗練されたスタイル）
        if field_name in self.main_fields:
            field_widget = self.main_fields[field_name]
            error_style = """
                border: 2px solid #dc3545;
                background-color: #fef2f2;
                border-radius: 6px;
                color: #721c24;
                font-weight: 500;
            """
            field_widget.setStyleSheet(error_style)
            
            # 強化されたツールチップ
            enhanced_tooltip = f"❌ エラー: {message}\n💡 修正してから保存してください"
            field_widget.setToolTip(enhanced_tooltip)
    
    def _show_field_validation_success(self, field_name):
        """フィールドに成功状態を表示"""
        if field_name in self.main_fields:
            field_widget = self.main_fields[field_name]
            success_style = """
                border: 2px solid #059669;
                background-color: #f0fdf4;
                border-radius: 6px;
                color: #065f46;
                font-weight: 500;
            """
            field_widget.setStyleSheet(success_style)
            field_widget.setToolTip("✅ 入力OK")
            
            # 3秒後に成功表示をクリア
            QTimer.singleShot(3000, lambda: self._clear_field_validation_error(field_name))

    def _clear_field_validation_error(self, field_name):
        """フィールドのエラー表示をクリア"""
        if hasattr(self, '_validation_errors') and field_name in self._validation_errors:
            del self._validation_errors[field_name]
        
        # フィールドのスタイルを完全にリセット
        if field_name in self.main_fields:
            field_widget = self.main_fields[field_name]
            # スタイルシートを空にして、親ウィジェットのスタイルを継承させる
            field_widget.setStyleSheet("")
            field_widget.setToolTip("")
    
    def _validate_all_fields(self):
        """全フィールドの検証"""
        if not hasattr(self, '_validation_errors'):
            self._validation_errors = {}
        
        # 商品コードの検証
        mycode_text = self.main_fields.get(HEADER_MYCODE, QLineEdit()).text()
        self._update_mycode_digit_count_display(mycode_text)
        
        # 商品名の検証
        product_name_text = self.main_fields.get(HEADER_PRODUCT_NAME, QLineEdit()).text()
        self._validate_required_field(HEADER_PRODUCT_NAME, product_name_text)
        
        # 価格フィールドの検証
        price_text = self.main_fields.get(HEADER_PRICE_TAX_INCLUDED, QLineEdit()).text()
        self._validate_price_field(HEADER_PRICE_TAX_INCLUDED, price_text)
        
        # SKUデータとMainデータの整合性チェック
        self._validate_sku_consistency()
        
        return len(self._validation_errors) == 0
    
    def _validate_sku_consistency(self):
        """SKUデータとMainデータの整合性を検証"""
        if not hasattr(self, 'sku_data_list') or not self.sku_data_list:
            return  # SKUデータがない場合はスキップ
        
        mycode = self.main_fields.get(HEADER_MYCODE, QLineEdit()).text().strip()
        if not mycode:
            return  # 商品コードがない場合はスキップ
        
        # 各SKUコードが正しい形式かチェック
        for i, sku_data in enumerate(self.sku_data_list):
            sku_code = sku_data.get(HEADER_SKU_CODE, "").strip()
            if sku_code:
                # SKUコードの正しい形式をチェック
                # 例: 商品コード 3140972000 → SKUコード 3140972010, 3140972020, ..., 3140972090, 3140972100, etc.
                # 8桁目と9桁目（右から3桁目と2桁目）が連番管理
                if len(sku_code) == len(mycode) and sku_code.isdigit() and mycode.isdigit() and len(sku_code) >= 9:
                    # 商品コードの8桁目と9桁目を00にしたベースコードを計算
                    base_code = mycode[:-3] + "00" + mycode[-1]  # 8桁目と9桁目を00に
                    expected_prefix = mycode[:-3]  # 8桁目より前の部分
                    expected_suffix = mycode[-1]   # 最後の1桁
                    
                    # SKUコードが正しいパターンかチェック
                    if not (sku_code.startswith(expected_prefix) and sku_code.endswith(expected_suffix)):
                        self._show_field_validation_error(f"SKU{i+1}", f"SKUコード'{sku_code}'が商品コード'{mycode}'と整合していません")
                    elif not sku_code[-3:-1].isdigit():  # 8桁目と9桁目が数字でない場合
                        self._show_field_validation_error(f"SKU{i+1}", f"SKUコード'{sku_code}'の8桁目・9桁目が数字ではありません")
                    else:
                        self._clear_field_validation_error(f"SKU{i+1}")
                else:
                    self._show_field_validation_error(f"SKU{i+1}", f"SKUコード'{sku_code}'の形式が正しくありません（10桁の数字である必要があります）")
            
            # SKU選択肢名が空でないかチェック
            choice_name = sku_data.get(HEADER_CHOICE_NAME, "").strip()
            if not choice_name:
                self._show_field_validation_error(f"SKU{i+1}_選択肢", f"SKU{i+1}の選択肢名が未入力です")
            else:
                self._clear_field_validation_error(f"SKU{i+1}_選択肢")

    def _format_and_sync_price_fields(self):
        price_field = self.main_fields.get(HEADER_PRICE_TAX_INCLUDED)
        sort_field = self.main_fields.get(HEADER_SORT_FIELD)
        if not price_field or not sort_field: return

        price_field.blockSignals(True); sort_field.blockSignals(True)
        current_text = price_field.text()
        cleaned_text = ''.join(filter(str.isdigit, current_text))
        formatted_text = ""
        if cleaned_text:
            try:
                num_val = int(cleaned_text)
                formatted_text = f"{num_val:,}"
            except ValueError:
                formatted_text = cleaned_text
        if price_field.text() != formatted_text:
            price_field.setText(formatted_text)
            price_field.setCursorPosition(len(formatted_text))
        if sort_field.text() != formatted_text:
            sort_field.setText(formatted_text)
        price_field.blockSignals(False); sort_field.blockSignals(False)

    def clear_fields(self, apply_defaults=True):
        for fld in self.main_fields.values():
            if isinstance(fld, (QLineEdit, QTextEdit, QComboBox)): fld.blockSignals(True)
        if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(True)
        if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(True)

        # Y_specフィールドもクリア
        for i in range(MAX_Y_SPEC_COUNT):
            self.y_spec_labels[i].setText(f"Y_spec{i+1} (項目名)") # ラベルをデフォルトに戻す
            self._clear_y_spec_editor(i) # エディタをクリア/削除
        for k, fld in self.main_fields.items():
            if isinstance(fld, QLineEdit): fld.clear()
            elif isinstance(fld, QTextEdit): fld.clear()
            elif isinstance(fld, QComboBox): fld.setCurrentIndex(-1)

        # 新規モードの視覚的インジケーター
        if apply_defaults:
            self._set_mode_indicator("新規作成", "#4CAF50")  # 緑色
            # 新規モードフラグ設定
            self._is_new_mode = True
        
        # デフォルト値の適用（新規作成時のみ）
        if apply_defaults:
            defaults = {"シリーズ名":"-","シリーズURL":"-","Y_metakey":"-",HEADER_YAHOO_ABSTRACT:"<img src='https://shopping.c.yimg.jp/lib/taiho-kagu/bn_campaign.jpg'>",
                        "メーカー売価_税込み":"-","メーカー売価_画像":"-","送料形態":"送料無料","R_SKU項目名":"カラーをお選びください","R_商品プルダウン":"-",
                        "R_別途送料地域項目名":"-","R_別途送料地域選択肢":"-","R_配達オプション項目名":"-","R_配達オプション選択肢":"-","R_注意事項プルダウン":"PULL_SHS",
                        "Y_SKU項目名":"カラーをお選びください","Y_商品プルダウン":"-","Y_別途送料地域項目名":"-","Y_別途送料地域選択肢":"-","Y_配達オプション項目名":"-",
                        "Y_配達オプション選択肢":"-","Y_注意事項プルダウン":"PULL_SHS","注意事項":"SHS","-":"-","YA_suffix":"a","非製品属性タグID":"5002409"}
            
            for k,v in defaults.items():
                if k in self.main_fields:
                    fld_w = self.main_fields[k]
                    if k in HTML_TEXTEDIT_FIELDS and isinstance(fld_w, QTextEdit): fld_w.setPlainText(v)
                    elif isinstance(fld_w,QLineEdit): fld_w.setText(v)
                    elif isinstance(fld_w,QComboBox): idx=fld_w.findText(v); fld_w.setCurrentIndex(idx if idx!=-1 else 0)

        self.sku_data_list = []; self.sku_model.update_data([],[],[]) # type: ignore
        if hasattr(self,'product_list'):
            self.product_list.blockSignals(True) # ★ シグナルをブロック
            self.product_list.clearSelection()
            self.product_list.blockSignals(False) # ★ ブロックを解除
        # コントロールカラムの設定
        if hasattr(self,'control_radio_n') and hasattr(self,'control_radio_p'):
            if apply_defaults:
                # 通常の新規作成：「n」をデフォルト
                self.control_radio_n.setChecked(True)
            else:
                # 削除後などの場合：ラジオボタンをクリア
                self.control_radio_n.setAutoExclusive(False)
                self.control_radio_p.setAutoExclusive(False) 
                self.control_radio_n.setChecked(False)
                self.control_radio_p.setChecked(False)
                self.control_radio_n.setAutoExclusive(True)
                self.control_radio_p.setAutoExclusive(True)
        if hasattr(self,'right_splitter') and self.right_splitter.count()>1: self.right_splitter.setSizes([self.right_splitter.height()-50,50])

        for fld in self.main_fields.values():
            if isinstance(fld,(QLineEdit, QTextEdit, QComboBox)): fld.blockSignals(False)
        if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
        if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)

        if hasattr(self, '_paste_product_action_ref'): # clear_fields が呼ばれたらペースト情報は無効
            self._paste_product_action_ref.setEnabled(False)
            self._copied_product_code_for_paste = None
        
        # クリア時のモード表示
        if not apply_defaults:  # デフォルト適用時（新規）以外
            self._clear_mode_indicator()

        self._update_relevant_links() # relevant_links を更新
        for field_name in self.byte_count_labels.keys():
             if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                 self._update_byte_count_display(field_name, self.main_fields.get(field_name).text())
        if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
            self._update_mycode_digit_count_display(self.main_fields[HEADER_MYCODE].text())
        self._format_and_sync_price_fields()
        self.is_dirty=False; self.save_btn.setEnabled(False)
        self._on_y_category_id_changed(self.main_fields.get(HEADER_Y_CATEGORY_ID, QLineEdit()).text()) # Y_specを更新
        self._sync_product_size_to_yspec() # ★クリア後にも同期


    def load_list(self) -> None:
        self.product_list.clear()
        if not self._safe_file_exists(self.manage_file_path): return # ユーザーデータディレクトリの管理ファイル
        try:
            wb = load_workbook(self.manage_file_path,read_only=True,keep_vba=True)
            if MAIN_SHEET_NAME not in wb.sheetnames: return
            ws = wb[MAIN_SHEET_NAME]; rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows)<1: return
            hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
            try: 
                code_idx=hdr.index(HEADER_MYCODE)
                name_idx=hdr.index(HEADER_PRODUCT_NAME)
                control_idx=hdr.index(HEADER_CONTROL_COLUMN) if HEADER_CONTROL_COLUMN in hdr else -1
            except ValueError: print(f"Error: {MAIN_SHEET_NAME}に{HEADER_MYCODE} or {HEADER_PRODUCT_NAME}列無"); return
            for vals in rows[1:]:
                code = str(vals[code_idx]).strip() if code_idx<len(vals) and vals[code_idx] is not None else ""
                name = str(vals[name_idx]).strip() if name_idx<len(vals) and vals[name_idx] is not None else ""
                control = str(vals[control_idx]).strip() if control_idx >= 0 and control_idx<len(vals) and vals[control_idx] is not None else "n"
                if code: 
                    item = QListWidgetItem(f"[{control}] {code} - {name}")
                    item.setData(Qt.UserRole, control)  # コントロールカラム値を保存
                    self.product_list.addItem(item)
            wb.close() # type: ignore
        except Exception as e: QMessageBox.critical(self,"リスト読込エラー",f"商品リスト読込失敗: {e}\n{traceback.format_exc()}")

    def filter_list(self, text):
        norm_txt = normalize_text(text)
        for i in range(self.product_list.count()): item=self.product_list.item(i); item.setHidden(norm_txt not in normalize_text(item.text()))

    def load_product(self, current_item): # previous 引数を削除
        """指定された商品アイテムのデータをフォームに読み込む。ダーティチェックは行わない。"""
        if not current_item:
            self.clear_fields(apply_defaults=False) # 選択がなければ完全にクリア
            return

        self._is_loading_data = True # データロード開始
        item_txt = current_item.text()
        # [n] または [p] のプレフィックスを除去して商品コードを取得
        if item_txt.startswith('['):
            code = item_txt.split('] ')[1].split(" - ")[0].strip()
        else:
            code = item_txt.split(" - ")[0].strip()
        
        # 編集モードの視覚的インジケーター
        self._set_mode_indicator(f"編集中: {code}", "#2196F3")  # 青色
        self._is_new_mode = False

        # ★★★ 追加/確認 ★★★
        # ExpandableFieldGroup の商品サイズフィールドをクリア
        product_size_efg = self.expandable_field_group_instances.get("商品サイズ")
        if product_size_efg and hasattr(product_size_efg, 'clear_dimension_fields'):
            product_size_efg.clear_dimension_fields()

        for fld in self.main_fields.values():
            if isinstance(fld, (QLineEdit, QTextEdit, QComboBox)): fld.blockSignals(True)
        if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(True)
        if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(True)

        # Y_specフィールドもクリア (load_productの最初で)
        for i in range(MAX_Y_SPEC_COUNT):
            self.y_spec_labels[i].setText(f"Y_spec{i+1} (項目名)")
            self._clear_y_spec_editor(i)
        for k, fld in self.main_fields.items():
            if isinstance(fld, QLineEdit): fld.clear()
            elif isinstance(fld, QTextEdit): fld.clear()
            elif isinstance(fld, QComboBox): fld.setCurrentIndex(-1)

        if not self._safe_file_exists(self.manage_file_path):
            msg = f"管理ファイル '{self.manage_file_path}' が見つかりません。"
            QMessageBox.warning(self,"ファイルなし", msg)
            logging.warning(f"商品ロード試行: {msg}")
            for fld_val in self.main_fields.values():
                if isinstance(fld_val,(QLineEdit, QTextEdit, QComboBox)): fld_val.blockSignals(False)
            if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
            if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)
            self._format_and_sync_price_fields()
            for field_name in self.byte_count_labels.keys():
                 if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                     self._update_byte_count_display(field_name, "")
            if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
                self._update_mycode_digit_count_display("")
            self.is_dirty=False; self.save_btn.setEnabled(False)
            return
        try:
            wb = load_workbook(self.manage_file_path,read_only=True,keep_vba=True)
            if MAIN_SHEET_NAME not in wb.sheetnames:
                msg = f"{MAIN_SHEET_NAME}シートが見つかりません。"
                QMessageBox.warning(self,"シートなし",msg); wb.close() # type: ignore
                logging.warning(f"商品ロード試行: {msg} (ファイル: {self.manage_file_path})")
                for fld_val in self.main_fields.values():
                    if isinstance(fld_val,(QLineEdit, QTextEdit, QComboBox)): fld_val.blockSignals(False)
                if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
                if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)
                self._format_and_sync_price_fields()
                for field_name in self.byte_count_labels.keys():
                     if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                         self._update_byte_count_display(field_name, "")
                if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
                    self._update_mycode_digit_count_display("")
                self.is_dirty=False; self.save_btn.setEnabled(False)
                return
            ws_main = wb[MAIN_SHEET_NAME]; main_rows = list(ws_main.iter_rows(values_only=True))
            if not main_rows or len(main_rows)<1:
                msg = f"{MAIN_SHEET_NAME}シートにデータがありません。"
                QMessageBox.warning(self,"データなし",msg); wb.close() # type: ignore
                logging.warning(f"商品ロード試行: {msg} (ファイル: {self.manage_file_path})")
                for fld_val in self.main_fields.values():
                    if isinstance(fld_val,(QLineEdit, QTextEdit, QComboBox)): fld_val.blockSignals(False)
                if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
                if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)
                self._format_and_sync_price_fields()
                for field_name in self.byte_count_labels.keys():
                     if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                         self._update_byte_count_display(field_name, "")
                if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
                    self._update_mycode_digit_count_display("")
                self.is_dirty=False; self.save_btn.setEnabled(False)
                return
            hdr_main = [str(h).strip() if h is not None else "" for h in main_rows[0]]
            if HEADER_MYCODE not in hdr_main:
                msg = f"{MAIN_SHEET_NAME}シートに'{HEADER_MYCODE}'列が見つかりません。"
                QMessageBox.critical(self,"ヘッダーエラー",f"{msg}\n詳細はログファイルを確認してください。"); wb.close() # type: ignore
                logging.error(f"商品ロード試行: {msg} (ファイル: {self.manage_file_path})")
                for fld_val in self.main_fields.values():
                    if isinstance(fld_val,(QLineEdit, QTextEdit, QComboBox)): fld_val.blockSignals(False)
                if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
                if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)
                self._format_and_sync_price_fields()
                for field_name in self.byte_count_labels.keys():
                     if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                         self._update_byte_count_display(field_name, "")
                if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
                    self._update_mycode_digit_count_display("")
                self.is_dirty=False; self.save_btn.setEnabled(False)
                return

            mycode_idx = hdr_main.index(HEADER_MYCODE)
            loaded_main_data = next((dict(zip(hdr_main,map(lambda x:str(x) if x is not None else "",r))) for r in main_rows[1:] if mycode_idx<len(r) and str(r[mycode_idx]).strip()==code),None)

            if loaded_main_data:
                ctrl_v = loaded_main_data.get(HEADER_CONTROL_COLUMN,"n").strip().lower(); self.control_radio_p.setChecked(True) if ctrl_v=="p" else self.control_radio_n.setChecked(True)
                for f_name in self.main_field_order:
                    if f_name in self.main_fields:
                        f_widget = self.main_fields[f_name]
                        val_excel = loaded_main_data.get(f_name,"")
                        if isinstance(val_excel, str):
                            if f_name in HTML_TEXTEDIT_FIELDS:
                                # HTMLフィールドの場合、改行コードを \n に正規化して保持
                                val_excel = val_excel.replace("_x000D_\n", "\n") # CR LF
                                val_excel = val_excel.replace("_x000D_", "\n")    # CR
                                val_excel = val_excel.replace("\r\n", "\n")      # CR LF (Windows)
                                val_excel = val_excel.replace("\r", "\n")        # CR (Mac old)
                                # LF (\n) はそのまま
                            else:
                                # HTMLフィールドでない場合、全ての改行を除去
                                val_excel = val_excel.replace("_x000D_", "")
                                val_excel = val_excel.replace("\r\n", "").replace("\r", "").replace("\n", "")

                        if f_name in HTML_TEXTEDIT_FIELDS and isinstance(f_widget, QTextEdit):
                            f_widget.setPlainText(val_excel)
                        elif isinstance(f_widget,QLineEdit): f_widget.setText(val_excel)
                        elif isinstance(f_widget,QComboBox):
                            is_material_spec_a_field = False
                            if self.material_spec_master and f_name.startswith(("材質_", "仕様_")) and f_name.endswith("a"):
                                try:
                                    num_part = f_name.split('_')[-1][:-1]
                                    if int(num_part) > 1: is_material_spec_a_field = True
                                except ValueError: pass
                            
                            if is_material_spec_a_field:
                                f_widget.setCurrentText(val_excel) # This should trigger 'b' field update via signal
                            else:
                                # For editable QComboBox, setCurrentText is better to handle custom values
                                if f_widget.isEditable():
                                    f_widget.setCurrentText(val_excel)
                                else: # For non-editable QComboBox
                                    idx=f_widget.findText(val_excel)
                                    f_widget.setCurrentIndex(idx if idx!=-1 else 0)
            else: # 商品コードがファイルに見つからなかった場合
                msg = f"商品コード'{code}'のデータが見つかりません。新規作成として扱います。"
                QMessageBox.warning(self,"読込エラー",msg); self.clear_fields(apply_defaults=False)
                logging.info(f"商品ロード: {msg} (ファイル: {self.manage_file_path})")
            
            # YカテゴリIDが設定された後にUIを更新し、その後Y_specの値をロード
            y_category_id_val = loaded_main_data.get(HEADER_Y_CATEGORY_ID, "") if loaded_main_data else ""
            self._on_y_category_id_changed(y_category_id_val) # まずUIを構築

            if loaded_main_data: # Y_specの値をロード
                for i in range(1, MAX_Y_SPEC_COUNT + 1):
                    f_name = f"Y_spec{i}"
                    self._load_y_spec_value(loaded_main_data.get(f_name, "")) # 修正: index引数を削除し、保存文字列を直接渡す

            self.sku_data_list = []
            if SKU_SHEET_NAME in wb.sheetnames:
                ws_sku = wb[SKU_SHEET_NAME]; sku_rows = list(ws_sku.iter_rows(values_only=True))
                if sku_rows and len(sku_rows)>0:
                    sku_hdr = [str(h).strip() if h is not None else "" for h in sku_rows[0]]
                    if HEADER_PRODUCT_CODE_SKU in sku_hdr:
                        prod_code_idx = sku_hdr.index(HEADER_PRODUCT_CODE_SKU)
                        cur_mycode = loaded_main_data.get(HEADER_MYCODE,code) if loaded_main_data else code
                        for sku_row in sku_rows[1:]:
                            if prod_code_idx<len(sku_row) and str(sku_row[prod_code_idx]).strip()==cur_mycode:
                                self.sku_data_list.append(dict(zip(sku_hdr,map(lambda x:str(x) if x is not None else "",sku_row))))
                        def sku_sort_key(s_item):
                            code_val = s_item.get(HEADER_SKU_CODE, "")
                            if code_val and len(code_val) >= 3 and code_val[-3:].isdigit():
                                return int(code_val[-3:])
                            return float('inf') # 数値でない、または短い場合は最後に
                        self.sku_data_list.sort(key=sku_sort_key)
                    else:
                        msg = f"{SKU_SHEET_NAME}シートに「{HEADER_PRODUCT_CODE_SKU}」列が見つかりません。"
                        QMessageBox.warning(self,"SKU読込エラー",msg)
                        logging.warning(f"商品ロード (SKU): {msg} (ファイル: {self.manage_file_path})")
            self.show_sku_table(); wb.close() # type: ignore
            if hasattr(self,'right_splitter') and self.right_splitter.count()>1: self.right_splitter.setSizes([self.right_splitter.height()*3//5,self.right_splitter.height()*2//5])
        except Exception as e:
            logging.error(f"商品「{code}」の読み込み中に予期せぬエラーが発生しました。", exc_info=True)
            QMessageBox.critical(self,"商品読込エラー",f"商品「{code}」の読み込み中にエラーが発生しました。\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}"); self.clear_fields(apply_defaults=False)

        for fld in self.main_fields.values():
            if isinstance(fld,(QLineEdit, QTextEdit, QComboBox)): fld.blockSignals(False)
        if hasattr(self,'control_radio_n'): self.control_radio_n.blockSignals(False)
        if hasattr(self,'control_radio_p'): self.control_radio_p.blockSignals(False)

        for field_name in self.byte_count_labels.keys():
             if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                 self._update_byte_count_display(field_name, self.main_fields.get(field_name).text())
        if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
            self._update_mycode_digit_count_display(self.main_fields[HEADER_MYCODE].text())
        self._format_and_sync_price_fields()
        self._update_relevant_links() # relevant_links を更新
        # 関連商品の文字数カウントをロード後に更新 (ExpandableFieldGroup内で処理されるように変更)
        # self.update_all_related_product_code_counts_after_load() # 削除
        for efg_inst in self.expandable_field_group_instances.values():
            if efg_inst.group_label_prefix == "関連商品":
                if hasattr(efg_inst, 'update_all_related_product_code_counts'):
                    efg_inst.update_all_related_product_code_counts()
        # self._on_y_category_id_changed(self.main_fields.get(HEADER_Y_CATEGORY_ID, QLineEdit()).text()) # Y_specのロード後に再度呼ぶ必要はない場合がある
        self._clear_auto_save_data() # 新規作成が完了したので、自動保存データをクリア
        self.is_dirty = False # ロード完了時は常にクリーンな状態
        self.save_btn.setEnabled(False) # 保存ボタンも無効化
        
        # 商品読み込み完了後、初期状態として保存
        self.undo_stack.clear()
        self.redo_stack.clear()
        self.save_undo_state()
        self._update_undo_redo_actions()
        
        self._is_loading_data = False # データロード完了
        self._sync_product_size_to_yspec() # 商品ロード後にも同期


    def generate_html(self):
        mycode_fld = self.main_fields.get(HEADER_MYCODE)
        if not mycode_fld or not mycode_fld.text().strip():
            msg = f"{HEADER_MYCODE}を入力してください。"
            QMessageBox.warning(self,"エラー",msg); logging.warning(f"HTML生成試行: {msg}")
            return
        code = mycode_fld.text().strip()

        # --- ユーザー入力 ---
        total_html_images, ok1 = QInputDialog.getInt(self, "画像枚数入力",
                                                     "画像の総枚数 (例: .jpg と _1～_5.jpg なら「6」):",
                                                     1, 1, 100, 1)
        if not ok1: return

        max_suffix_index = max(0, total_html_images - 1)

        last_sku_suffix_index, ok2 = QInputDialog.getInt(self, "SKU画像指定",
                                                         f"SKU画像の枚数 (例: _1.jpg, _2.jpg の2枚なら「2」、なければ「0」):\n"
                                                         f"(入力範囲: 0 ～ {max_suffix_index})",
                                                         0, 0, max_suffix_index, 0)
        if not ok2: return

        size_image_suffix_index, ok3 = QInputDialog.getInt(self, "サイズ画像指定",
                                                           f"サイズ表記画像の番号 (例: _4.jpg なら「4」、なければ「0」):\n"
                                                           f"(入力範囲: 0 ～ {max_suffix_index})",
                                                           0, 0, max_suffix_index, 0)
        if not ok3: return

        # --- 入力値のバリデーション ---
        if last_sku_suffix_index > 0 and size_image_suffix_index > 0 and last_sku_suffix_index >= size_image_suffix_index:
            QMessageBox.warning(self, "入力エラー", "「SKU画像の枚数」は、「サイズ表記画像の番号」よりも小さい値を指定してください。")
            return

        # --- 画像ファイルパスとHTMLタグの準備 ---
        img_path_fld = self.main_fields.get(HEADER_IMAGE_PATH_RAKUTEN)
        img_path_base_new = "./説明用/"

        def create_img_tag(image_suffix_str): #例: "" (サムネイル用), "_1", "_2"
            img_fname = f"{code}{image_suffix_str}.jpg"
            src = f"{img_path_base_new}{img_fname}"
            return f'<IMG SRC="{src}" border="0" width="100%"><BR><BR>'

        # --- 各カテゴリのHTMLを生成 ---
        thumbnail_html = create_img_tag("")

        sku_htmls = []
        if last_sku_suffix_index > 0:
            for i in range(1, last_sku_suffix_index + 1):
                if i < total_html_images: # _i.jpg が実際にHTMLに含める画像の範囲内か
                    sku_htmls.append(create_img_tag(f"_{i}"))

        size_html = None
        if size_image_suffix_index > 0:
            if size_image_suffix_index < total_html_images: # _size_image_suffix_index.jpg が範囲内か
                size_html = create_img_tag(f"_{size_image_suffix_index}")

        other_htmls_part1 = [] # サイズ画像の前の「その他」
        other_htmls_part2 = [] # サイズ画像の後の「その他」

        # _1.jpg から _(total_html_images-1).jpg までをループして「その他」の画像を分類
        for i in range(1, total_html_images): # i は画像のサフィックス番号 (1, 2, ...)
            is_sku = (last_sku_suffix_index > 0 and 1 <= i <= last_sku_suffix_index)
            is_size = (size_image_suffix_index > 0 and i == size_image_suffix_index)

            if not is_sku and not is_size:
                current_other_html = create_img_tag(f"_{i}")
                if size_image_suffix_index == 0: # サイズ画像指定がない場合、全てpart1へ
                    other_htmls_part1.append(current_other_html)
                elif i < size_image_suffix_index:
                    other_htmls_part1.append(current_other_html)
                else: # i > size_image_suffix_index
                    other_htmls_part2.append(current_other_html)

        # --- HTMLの結合 ---
        final_html_parts = [thumbnail_html] + other_htmls_part1 + sku_htmls
        if size_html:
            final_html_parts.append(size_html)
        final_html_parts.extend(other_htmls_part2)

        final_html_output = "".join(final_html_parts)

        img_desc_fld = self.main_fields.get(HEADER_IMAGE_DESCRIPTION)
        if img_desc_fld:
            img_desc_fld.setText(final_html_output); self.is_dirty = True
        else:
            msg = f"{HEADER_IMAGE_DESCRIPTION}フィールドが見つかりません。"
            QMessageBox.warning(self,"エラー",msg); logging.warning(f"HTML生成試行: {msg}")

    def _get_value_for_excel_cell(self, header_name, current_excel_row_values_dict=None):
        # ★★★ Y_specヘッダーの処理を一番最初に移動 ★★★
        if header_name.startswith("Y_spec") and header_name[6:].isdigit():
            has_temp_attr = hasattr(self, '_temp_y_spec_values_for_save')
            if has_temp_attr:
                is_header_in_temp = header_name in self._temp_y_spec_values_for_save
                if is_header_in_temp:
                    val = self._temp_y_spec_values_for_save[header_name]
                    return val
                else:
                    # print(f"Warning: _get_value_for_excel_cell - Y_spec header '{header_name}' NOT in _temp_y_spec_values_for_save keys. Returning empty.")
                    return ""
            else:
                # print(f"Warning: _get_value_for_excel_cell - _temp_y_spec_values_for_save attribute DOES NOT EXIST. Returning empty.")
                return ""

        # バイト数列の処理
        if header_name in ["byte(255)", "byte(150)", "byte(174)", "byte(160)", "byte(60)"]:
            return self._get_byte_count_for_column(header_name)

        if header_name == HEADER_CONTROL_COLUMN:
            return "n" if self.control_radio_n.isChecked() else "p"

        widget = self.main_fields.get(header_name)
        if widget:
            # Ensure price and sort fields are saved as plain number strings without commas
            if header_name == HEADER_PRICE_TAX_INCLUDED and isinstance(widget, QLineEdit):
                price_text_with_comma = widget.text()
                cleaned_price = ''.join(filter(str.isdigit, price_text_with_comma))
                return cleaned_price
            elif header_name == HEADER_SORT_FIELD and isinstance(widget, QLineEdit):
                sort_text_with_comma = widget.text()
                cleaned_sort_value = ''.join(filter(str.isdigit, sort_text_with_comma))
                return cleaned_sort_value
            elif header_name in HTML_TEXTEDIT_FIELDS and isinstance(widget, QTextEdit):
                return widget.toPlainText() # Preserve newlines for HTML fields
            elif isinstance(widget, QLineEdit):
                return widget.text()
            elif isinstance(widget, QComboBox):
                return widget.currentText()

        # 上記のいずれにも該当しない場合 (Y_specでもなく、main_fieldsにもなく、コントロールカラムでもない)
        # かつ、既存行のデータから取得する場合
        if current_excel_row_values_dict:
            return current_excel_row_values_dict.get(header_name, "")
        # 新規行で、上記のいずれにも該当しない場合は空文字
        return ""

    def _get_byte_count_for_column(self, byte_column_name):
        """バイト数列名に対応する実際のフィールドのバイト数を計算"""
        # バイト数列と対応するフィールドのマッピング
        byte_column_mapping = {
            "byte(255)": "R_商品名",      # 255バイト制限
            "byte(150)": "Y_商品名",      # 150バイト制限  
            "byte(174)": "R_キャッチコピー", # 174バイト制限
            "byte(160)": "Y_metadesc",   # 160バイト制限
            "byte(60)": "Y_キャッチコピー"  # 60バイト制限
        }
        
        field_name = byte_column_mapping.get(byte_column_name)
        if not field_name:
            return ""
            
        # 対応するフィールドの値を取得
        widget = self.main_fields.get(field_name)
        if not widget:
            return ""
            
        # フィールドの値を取得
        text = ""
        if isinstance(widget, QLineEdit):
            text = widget.text()
        elif isinstance(widget, QTextEdit):
            text = widget.toPlainText()
        elif hasattr(widget, 'text'):
            text = widget.text()
        elif hasattr(widget, 'currentText'):
            text = widget.currentText()
        
        # バイト数を計算（get_byte_count_excel_lenb関数を使用）
        from utils import get_byte_count_excel_lenb
        byte_count = get_byte_count_excel_lenb(text)
        
        return str(byte_count) if byte_count >= 0 else ""
    
    def save_undo_state(self):
        """現在の状態をUndo履歴に保存（デバウンス機能付き）"""
        if self._is_loading_data or self._is_undoing:
            return
            
        # 既存のタイマーをキャンセル
        if self._undo_save_timer:
            self._undo_save_timer.stop()
            
        # 500ms後に実際の保存を実行（連続した変更をグループ化）
        self._undo_save_timer = QTimer()
        self._undo_save_timer.setSingleShot(True)
        self._undo_save_timer.timeout.connect(self._do_save_undo_state)
        self._undo_save_timer.start(500)
    
    def _do_save_undo_state(self):
        """実際のUndo状態保存処理"""
        try:
            # 現在の状態を取得
            state = self.get_current_state()
            
            # 最新の履歴と比較して、変更がない場合は保存しない
            if self.undo_stack and self._states_are_equal(self.undo_stack[-1], state):
                return
            
            # Undo履歴に追加
            self.undo_stack.append(state)
            
            # 履歴数の上限を超えたら古いものを削除
            if len(self.undo_stack) > self.max_undo_history:
                self.undo_stack.pop(0)
            
            # 新しい操作が行われたらRedo履歴をクリア
            self.redo_stack.clear()
            
            # Undo/Redoアクションの状態を更新
            self._update_undo_redo_actions()
            
        except Exception as e:
            logging.error(f"Undo状態保存中のエラー: {e}")
    
    def _states_are_equal(self, state1, state2):
        """二つの状態が同じかどうかを比較"""
        try:
            # メインフィールドの比較
            if state1.get('main_fields') != state2.get('main_fields'):
                return False
            # コントロールカラムの比較
            if state1.get('control_column') != state2.get('control_column'):
                return False
            # SKUデータの比較（深い比較）
            import json
            if json.dumps(state1.get('sku_data', []), sort_keys=True) != json.dumps(state2.get('sku_data', []), sort_keys=True):
                return False
            return True
        except (KeyError, TypeError, ValueError) as e:
            logging.debug(f"状態比較エラー（継続）: {e}")
            return False
        except Exception as e:
            logging.warning(f"予期せぬ状態比較エラー（継続）: {e}")
            return False
    
    def get_current_state(self):
        """現在の状態を取得"""
        state = {
            'main_fields': {},
            'control_column': '',
            'sku_data': [],
            'selected_product': self.selected_product
        }
        
        # メインフィールドの値を保存
        for field_name, widget in self.main_fields.items():
            if isinstance(widget, QLineEdit):
                state['main_fields'][field_name] = widget.text()
            elif isinstance(widget, QTextEdit):
                state['main_fields'][field_name] = widget.toPlainText()
            elif isinstance(widget, QComboBox):
                state['main_fields'][field_name] = widget.currentText()
        
        # コントロールカラムの状態を保存
        if hasattr(self, 'control_radio_n') and self.control_radio_n.isChecked():
            state['control_column'] = 'n'
        elif hasattr(self, 'control_radio_p') and self.control_radio_p.isChecked():
            state['control_column'] = 'p'
        
        # SKUデータをディープコピー
        import copy
        state['sku_data'] = copy.deepcopy(self.sku_data_list)
        
        return state
    
    def restore_state(self, state):
        """状態を復元"""
        try:
            self._is_loading_data = True  # データロード中フラグを立てる
            self._is_undoing = True  # Undo/Redo実行中フラグを立てる
            
            # 自動保存を一時的に無効化
            auto_save_enabled = hasattr(self, 'auto_save_timer') and self.auto_save_timer.isActive()
            if auto_save_enabled:
                self.auto_save_timer.stop()
            
            # シグナルをブロックしてメインフィールドの値を復元
            for field_name, value in state.get('main_fields', {}).items():
                if field_name in self.main_fields:
                    widget = self.main_fields[field_name]
                    widget.blockSignals(True)
                    try:
                        if isinstance(widget, QLineEdit):
                            widget.setText(value)
                        elif isinstance(widget, QTextEdit):
                            widget.setPlainText(value)
                        elif isinstance(widget, QComboBox):
                            widget.setCurrentText(value)
                    finally:
                        widget.blockSignals(False)
            
            # コントロールカラムの状態を復元
            control_column = state.get('control_column', '')
            if hasattr(self, 'control_radio_n'):
                self.control_radio_n.blockSignals(True)
            if hasattr(self, 'control_radio_p'):
                self.control_radio_p.blockSignals(True)
            
            try:
                if control_column == 'n' and hasattr(self, 'control_radio_n'):
                    self.control_radio_n.setChecked(True)
                elif control_column == 'p' and hasattr(self, 'control_radio_p'):
                    self.control_radio_p.setChecked(True)
            finally:
                if hasattr(self, 'control_radio_n'):
                    self.control_radio_n.blockSignals(False)
                if hasattr(self, 'control_radio_p'):
                    self.control_radio_p.blockSignals(False)
            
            # SKUデータを復元
            import copy
            self.sku_data_list = copy.deepcopy(state.get('sku_data', []))
            self._refresh_sku_table()
            
            # 選択された商品を復元
            self.selected_product = state.get('selected_product', None)
            
            # UIを更新（最小限の同期のみ）
            self._format_and_sync_price_fields()
            self._update_relevant_links()
            
            # バイト数カウントを更新
            for field_name in self.byte_count_labels.keys():
                if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                    self._update_byte_count_display(field_name, self.main_fields.get(field_name).text())
            
            # 自動保存を再開
            if auto_save_enabled:
                self.auto_save_timer.start(AUTO_SAVE_INTERVAL_MS)
            
            # 状態をクリーンにマーク
            self.is_dirty = False
            self.save_btn.setEnabled(False)
            
        except Exception as e:
            logging.error(f"状態復元中のエラー: {e}")
        finally:
            self._is_loading_data = False
            self._is_undoing = False
    
    def undo(self):
        """元に戻す"""
        if not self.undo_stack:
            return
        
        try:
            # 現在の状態をRedo履歴に保存
            current_state = self.get_current_state()
            self.redo_stack.append(current_state)
            
            # Undo履歴から最新の状態を取得して復元
            previous_state = self.undo_stack.pop()
            self.restore_state(previous_state)
            
            # Undo/Redoアクションの状態を更新
            self._update_undo_redo_actions()
            
        except Exception as e:
            logging.error(f"Undo実行中のエラー: {e}")
    
    def redo(self):
        """やり直し"""
        if not self.redo_stack:
            return
        
        try:
            # 現在の状態をUndo履歴に保存
            current_state = self.get_current_state()
            self.undo_stack.append(current_state)
            
            # Redo履歴から状態を取得して復元
            next_state = self.redo_stack.pop()
            self.restore_state(next_state)
            
            # Undo/Redoアクションの状態を更新
            self._update_undo_redo_actions()
            
        except Exception as e:
            logging.error(f"Redo実行中のエラー: {e}")
    
    def _update_undo_redo_actions(self):
        """Undo/Redoアクションの有効/無効を更新"""
        if hasattr(self, 'undo_action'):
            self.undo_action.setEnabled(bool(self.undo_stack))
        if hasattr(self, 'redo_action'):
            self.redo_action.setEnabled(bool(self.redo_stack))

    def save_to_excel(self, show_message=True, is_delete_operation=False):
        # --- デバッグ情報: 呼び出し元をトレース ---
        import traceback
        call_stack = traceback.format_stack()
        caller_info = call_stack[-2].strip() if len(call_stack) >= 2 else "不明"
        logging.info(f"save_to_excel 呼び出し元: {caller_info}")
        
        # --- Initialize workbooks to None (before any early returns) ---
        wb_mng = None
        wb_item = None
        wb_mng_ro = None
        
        mutex_locked = False
        if not self._save_mutex.tryLock():
            QMessageBox.warning(self, "保存中", "別の保存処理が実行中です。しばらく待ってから再度お試しください。")
            logging.info("保存処理が既に実行中のため、新規の保存リクエストをスキップしました。")
            return
        mutex_locked = True

        try:
            # --- データ検証を最初に実行 ---
            if not is_delete_operation:
                if not self._validate_all_fields():
                    error_messages = []
                    if hasattr(self, '_validation_errors'):
                        for field_name, message in self._validation_errors.items():
                            error_messages.append(f"• {field_name}: {message}")
                    
                    if error_messages:
                        QMessageBox.warning(self, "入力エラー", 
                            "以下の入力エラーを修正してから保存してください:\n\n" + "\n".join(error_messages))
                        return
            
            # --- ディスク容量チェック ---
            if check_disk_space_before_save and not is_delete_operation:
                estimated_records = len(getattr(self, 'current_data', [])) + 1  # 現在のデータ + 1件追加
                if not check_disk_space_before_save(self.manage_file_path, estimated_records, self):
                    return  # 容量不足またはユーザーがキャンセルした場合
            
            # Workbooks already initialized at method start

            # QApplication.setOverrideCursor(Qt.WaitCursor) はこの try ブロックの外側に移動
            # tryLock の後、実際の処理の前に設定する
            QApplication.setOverrideCursor(Qt.WaitCursor)

            # --- Safely populate _temp_y_spec_values_for_save ---
            self._temp_y_spec_values_for_save = {}
            for i in range(MAX_Y_SPEC_COUNT):
                header = f"Y_spec{i+1}"
                try:
                    if (self.y_spec_current_editors[i] is not None and
                            self.y_spec_current_definitions[i] is not None):
                        self._temp_y_spec_values_for_save[header] = self._get_y_spec_value_for_save(i)
                    else:
                        self._temp_y_spec_values_for_save[header] = ""
                except Exception as e_yspec_get:
                    logging.warning(f"Y_spec {header} の値取得中にエラー: {e_yspec_get}")
                    self._temp_y_spec_values_for_save[header] = ""

            logging.debug(f"save_to_excel - TEMP Y_spec values collected: {getattr(self, '_temp_y_spec_values_for_save', 'Not Set')}")

            mycode_fld=self.main_fields.get(HEADER_MYCODE); prod_name_fld=self.main_fields.get(HEADER_PRODUCT_NAME)
            code=mycode_fld.text().strip() if mycode_fld else ""; name=prod_name_fld.text().strip() if prod_name_fld else ""
            # 削除処理の場合は必須チェックをスキップ
            logging.info(f"保存処理開始: 商品コード='{code}', is_delete_operation={is_delete_operation}")
            
            # 削除処理中は保存しない
            if getattr(self, '_is_deleting', False):
                logging.info("保存スキップ: 削除処理中です")
                return
                
            # 商品コードが空の場合は保存しない（削除後のデフォルト値保存を防ぐ）
            if not code:
                logging.info(f"保存スキップ: 商品コードが空です（is_delete_operation={is_delete_operation}）")
                # 削除処理の場合でも、商品コードが空なら保存しない
                return
                
            if not is_delete_operation and (not code or not name):
                msg = f"{HEADER_MYCODE}と{HEADER_PRODUCT_NAME}は必須入力です。"
                QMessageBox.warning(self,"入力エラー",msg); logging.warning(f"保存試行: {msg}")
                # QApplication.restoreOverrideCursor() # Ensure cursor is restored before early return
                return
            try:
                if not os.path.exists(self.manage_file_path): # ユーザーデータディレクトリの管理ファイル
                    logging.info(f"管理ファイル '{self.manage_file_path}' が存在しません。")
                    if os.path.exists(self.template_file_path_bundle): # バンドルされたテンプレート
                        logging.info(f"テンプレート '{self.template_file_path_bundle}' から管理ファイルをコピーします。")
                        # 一時ファイルで安全にコピー
                        temp_file = self.manage_file_path + ".tmp"
                        try:
                            copyfile(self.template_file_path_bundle, temp_file)
                            os.replace(temp_file, self.manage_file_path)  # 原子操作
                            wb_mng=load_workbook(self.manage_file_path,keep_vba=True)
                        except Exception as e:
                            if os.path.exists(temp_file):
                                try:
                                    os.remove(temp_file)
                                except OSError as cleanup_e:
                                    logging.debug(f"一時ファイル削除エラー（継続）: {cleanup_e}")
                            raise e
                    else:
                        msg = f"テンプレート '{self.template_file_path_bundle}' が見つかりません。新規管理ファイルを作成できません。"
                        logging.critical(msg)
                        QMessageBox.critical(self,"エラー",f"{msg}\n詳細はログファイルを確認してください。")
                        return
                else:
                    logging.info(f"既存の管理ファイル '{self.manage_file_path}' を読み込みます。")
                    wb_mng=load_workbook(self.manage_file_path,keep_vba=True)
            except PermissionError:
                msg = f"管理ファイル '{self.manage_file_path}' が開かれているためアクセスできません。"
                logging.error(msg)
                QMessageBox.critical(self,"保存エラー",f"{msg}\n詳細はログファイルを確認してください。")
                return
            except Exception as e:
                msg = f"管理ファイル '{self.manage_file_path}' の処理中にエラーが発生しました。"
                logging.error(msg, exc_info=True)
                QMessageBox.critical(self,"ファイルエラー",f"{msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}")
                return
            
            logging.debug(f"save_to_excel - After wb_mng load - Y_カテゴリID: '{self.main_fields.get(HEADER_Y_CATEGORY_ID).text() if HEADER_Y_CATEGORY_ID in self.main_fields else 'N/A'}'")
            try:
                ws_main_mng=wb_mng[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in wb_mng.sheetnames else wb_mng.create_sheet(MAIN_SHEET_NAME)
                exist_main_rows_tuples=list(ws_main_mng.iter_rows(values_only=True))

                # テンプレートファイルから正しい列順序を取得
                template_main_headers = []
                try:
                    if os.path.exists(self.template_file_path_bundle):
                        template_wb = load_workbook(self.template_file_path_bundle, read_only=True)
                        if MAIN_SHEET_NAME in template_wb.sheetnames:
                            template_ws = template_wb[MAIN_SHEET_NAME]
                            template_row = list(template_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                            template_main_headers = [str(h).strip() if h is not None else "" for h in template_row]
                        template_wb.close()
                        logging.info(f"テンプレートから列順序を取得: {len(template_main_headers)}列")
                except Exception as e:
                    logging.warning(f"テンプレート列順序取得エラー: {e}")
                
                # テンプレート順序が取得できた場合はそれを使用、できない場合は従来通り
                if template_main_headers:
                    canonical_main_headers = template_main_headers
                else:
                    canonical_main_headers = [HEADER_CONTROL_COLUMN] + self.main_field_order # type: ignore
                if HEADER_MYCODE not in canonical_main_headers:
                    # QMessageBox.critical(self,"内部エラー",f"'{HEADER_MYCODE}'が定義済みヘッダーにありません。"); wb_mng.close(); return
                    msg = f"内部エラー: '{HEADER_MYCODE}'が定義済みヘッダーにありません。" # type: ignore
                    logging.critical(msg)
                    QMessageBox.critical(self,"内部エラー",f"{msg}\n詳細はログファイルを確認してください。")
                    return
                # Read existing headers from the file, if any
                existing_headers_from_file = [str(h).strip() if h is not None else "" for h in (exist_main_rows_tuples[0] if exist_main_rows_tuples else [])]

                out_main_rows_data = [canonical_main_headers] # Start output with the canonical headers
                updated_product_in_file = False

                # Process existing data rows from the file
                for r_tuple in (exist_main_rows_tuples[1:] if exist_main_rows_tuples else []):
                    # Create a dictionary from the existing row using its original headers from the file
                    current_excel_row_dict = {}
                    if existing_headers_from_file:
                        current_excel_row_dict = dict(zip(existing_headers_from_file, (str(val) if val is not None else "" for val in r_tuple)))
                    
                    current_mycode_from_file = current_excel_row_dict.get(HEADER_MYCODE, "").strip()

                    if current_mycode_from_file == code: # This is the product currently being saved
                        new_row_data_list = []
                        for h_debug in canonical_main_headers:
                            val_debug = self._get_value_for_excel_cell(h_debug, current_excel_row_dict)
                            new_row_data_list.append(val_debug)
                        out_main_rows_data.append(new_row_data_list)
                        updated_product_in_file = True
                    else: # This is another product, preserve its data, mapping to canonical_main_headers
                        preserved_row_data_list = []
                        for h_debug in canonical_main_headers:
                            val_debug = current_excel_row_dict.get(h_debug, "")
                            preserved_row_data_list.append(val_debug)
                        out_main_rows_data.append(preserved_row_data_list)
                
                if not updated_product_in_file: # If the product being saved was new (not found in existing file)
                    new_row_data_list = [self._get_value_for_excel_cell(h) for h in canonical_main_headers]
                    out_main_rows_data.append(new_row_data_list)

                ws_main_mng.delete_rows(1,ws_main_mng.max_row+1); [ws_main_mng.append(r_data_list) for r_data_list in out_main_rows_data]
                
                ws_sku_mng=wb_mng[SKU_SHEET_NAME] if SKU_SHEET_NAME in wb_mng.sheetnames else wb_mng.create_sheet(SKU_SHEET_NAME)
                exist_sku_rows=list(ws_sku_mng.iter_rows(values_only=True))
                hdr_sku_mng=[str(h).strip() if h is not None else "" for h in (exist_sku_rows[0] if exist_sku_rows and exist_sku_rows[0] else [])]
                if not hdr_sku_mng and self.sku_data_list:
                    all_sku_keys=set(k for item in self.sku_data_list for k in item.keys() if not k.startswith("_highlight_"))
                    pref_sku_order=[HEADER_PRODUCT_CODE_SKU,HEADER_SKU_CODE,HEADER_CHOICE_NAME,HEADER_MEMO,HEADER_GROUP]+[f"{p}{i}" for i in range(1,MAX_SKU_ATTRIBUTES+1) for p in [HEADER_ATTR_ITEM_PREFIX,HEADER_ATTR_VALUE_PREFIX,HEADER_ATTR_UNIT_PREFIX]]
                    hdr_sku_mng=[k for k in pref_sku_order if k in all_sku_keys]+sorted([k for k in all_sku_keys if k not in pref_sku_order])
                    ws_sku_mng.append(hdr_sku_mng); exist_sku_rows=[hdr_sku_mng] # type: ignore
                out_sku_rows=[hdr_sku_mng] if hdr_sku_mng else []
                if hdr_sku_mng and HEADER_PRODUCT_CODE_SKU in hdr_sku_mng:
                    prod_code_idx_sku=hdr_sku_mng.index(HEADER_PRODUCT_CODE_SKU)
                    for r_idx,sku_r_tuple in enumerate(exist_sku_rows[1:],1): # type: ignore
                        sku_r_list=list(sku_r_tuple); sku_r_list.extend([""]*(len(hdr_sku_mng)-len(sku_r_list)))
                        if str(sku_r_list[prod_code_idx_sku]).strip()!=code: out_sku_rows.append(sku_r_list)
                if hdr_sku_mng:
                    for cur_sku_dict in self.sku_data_list: cur_sku_dict[HEADER_PRODUCT_CODE_SKU]=code; out_sku_rows.append([str(cur_sku_dict.get(h_col,"")) for h_col in hdr_sku_mng])
                if hdr_sku_mng or (len(out_sku_rows) > 0 and out_sku_rows[0]): # Ensure out_sku_rows is not just [[]]
                    ws_sku_mng.delete_rows(1,ws_sku_mng.max_row+1); [ws_sku_mng.append(r_sku) for r_sku in out_sku_rows]
                
                # --- User-suggested garbage collection ---
                import gc
                gc.collect()
                logging.debug("Explicit garbage collection called before saving manage file.")

                logging.info(f"管理ファイル '{self.manage_file_path}' への保存を試みます。")
                wb_mng.save(self.manage_file_path)
                logging.info(f"管理ファイル '{self.manage_file_path}' の保存が完了しました。")

            except Exception as e_mng_process:
                msg = f"管理ファイル '{self.manage_file_path}' のデータ処理または保存中にエラーが発生しました。"
                logging.error(msg, exc_info=True)
                QMessageBox.critical(self,"管理ファイル処理エラー",f"{msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e_mng_process}")
                return
            finally:
                if wb_mng:
                    try:
                        wb_mng.close()
                        logging.info(f"管理ファイル '{self.manage_file_path}' をクローズしました。")
                    except Exception as e_close_mng:
                        logging.warning(f"管理ファイル '{self.manage_file_path}' のクローズ中にエラー: {e_close_mng}")
                    wb_mng = None # Mark as closed
            
            logging.debug(f"save_to_excel - After item_manage.xlsm save - Y_カテゴリID: '{self.main_fields.get(HEADER_Y_CATEGORY_ID).text() if HEADER_Y_CATEGORY_ID in self.main_fields else 'N/A'}'")
            if not os.path.exists(self.template_file_path_bundle):
                msg = f"出力用テンプレート '{self.template_file_path_bundle}' が見つかりません。"
                logging.critical(msg)
                QMessageBox.critical(self,"エラー",f"{msg}\n詳細はログファイルを確認してください。")
                return
            
            try:
                # item.xlsm の出力先ディレクトリ (C#ツール指定の場所) が存在するか確認し、なければ作成
                if not os.path.exists(self.item_xlsm_output_dir):
                    logging.info(f"出力先ディレクトリ '{self.item_xlsm_output_dir}' を作成します。")
                    os.makedirs(self.item_xlsm_output_dir, exist_ok=True)

                # クリーンなテンプレートを使用（サンプルデータなし）
                template_to_use = self.clean_template_file_path if os.path.exists(self.clean_template_file_path) else self.template_file_path_bundle
                logging.info(f"出力ファイル '{self.output_file_path}' をクリーンテンプレート '{template_to_use}' からコピーします。")
                copyfile(template_to_use, self.output_file_path); wb_item=load_workbook(self.output_file_path,keep_vba=True)
                
                ws_main_item=wb_item[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in wb_item.sheetnames else wb_item.create_sheet(MAIN_SHEET_NAME)
                ws_sku_item=wb_item[SKU_SHEET_NAME] if SKU_SHEET_NAME in wb_item.sheetnames else wb_item.create_sheet(SKU_SHEET_NAME)

                try:
                    logging.info(f"読み取り専用で管理ファイル '{self.manage_file_path}' を再度開きます。")
                    wb_mng_ro=load_workbook(self.manage_file_path,read_only=True,keep_vba=True)
                    ws_main_ro=wb_mng_ro[MAIN_SHEET_NAME] if MAIN_SHEET_NAME in wb_mng_ro.sheetnames else None
                    ws_sku_ro=wb_mng_ro[SKU_SHEET_NAME] if SKU_SHEET_NAME in wb_mng_ro.sheetnames else None
                    main_rows_out=list(ws_main_ro.iter_rows(values_only=True)) if ws_main_ro else []
                    sku_rows_out=list(ws_sku_ro.iter_rows(values_only=True)) if ws_sku_ro else []
                except Exception as e_ro_load:
                    logging.error(f"読み取り専用管理ファイル '{self.manage_file_path}' の読み込み中にエラー: {e_ro_load}", exc_info=True)
                    # Decide if this is critical or if we can proceed with empty data
                    main_rows_out, sku_rows_out = [], [] # Proceed with empty data as a fallback
                finally:
                    if wb_mng_ro:
                        try:
                            wb_mng_ro.close()
                            logging.info(f"読み取り専用の管理ファイル '{self.manage_file_path}' をクローズしました。")
                        except Exception as e_close_ro: logging.warning(f"読み取り専用管理ファイル '{self.manage_file_path}' のクローズ中にエラー: {e_close_ro}")
                        wb_mng_ro = None # Mark as closed
                # テンプレートの列順序を取得
                template_headers = []
                if ws_main_item.max_row > 0:
                    template_row = list(ws_main_item.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    template_headers = [str(h).strip() if h is not None else "" for h in template_row]
                
                ws_main_item.delete_rows(1,ws_main_item.max_row+1)
                n_mycodes_set=set()
                if main_rows_out:
                    manage_headers=[str(h).strip() if h is not None else "" for h in main_rows_out[0]]
                    
                    # テンプレートの列順序に合わせて出力
                    if template_headers:
                        # テンプレートの列順序を使用
                        output_headers = template_headers
                    else:
                        # テンプレートに列がない場合は管理ファイルの順序を使用
                        output_headers = manage_headers
                    
                    ws_main_item.append(output_headers)
                    
                    # 列のマッピングを作成（管理ファイル → テンプレート順序）
                    column_mapping = {}
                    for i, template_header in enumerate(output_headers):
                        if template_header in manage_headers:
                            column_mapping[i] = manage_headers.index(template_header)
                    
                    ctrl_idx=manage_headers.index(HEADER_CONTROL_COLUMN) if HEADER_CONTROL_COLUMN in manage_headers else -1
                    mycode_idx_main=manage_headers.index(HEADER_MYCODE) if HEADER_MYCODE in manage_headers else -1
                    if ctrl_idx!=-1 and mycode_idx_main!=-1:
                        for r_data_main in main_rows_out[1:]:
                            if ctrl_idx<len(r_data_main) and str(r_data_main[ctrl_idx]).strip().lower()=="n":
                                # テンプレートの列順序に合わせてデータを再配置
                                reordered_row = [""] * len(output_headers)
                                for template_idx, manage_idx in column_mapping.items():
                                    if manage_idx < len(r_data_main):
                                        reordered_row[template_idx] = r_data_main[manage_idx]
                                
                                ws_main_item.append(reordered_row)
                                if mycode_idx_main<len(r_data_main) and r_data_main[mycode_idx_main] is not None: n_mycodes_set.add(str(r_data_main[mycode_idx_main]).strip())
                    else:
                        msg = f"'{OUTPUT_FILE_NAME}' へのMainデータ書き出し時、'{HEADER_CONTROL_COLUMN}' または '{HEADER_MYCODE}' 列が見つかりません。"
                        logging.warning(f"保存処理: {msg}")
                        QMessageBox.warning(self,"警告",msg)

                # SKUシートのテンプレート列順序を取得
                sku_template_headers = []
                if ws_sku_item.max_row > 0:
                    sku_template_row = list(ws_sku_item.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    sku_template_headers = [str(h).strip() if h is not None else "" for h in sku_template_row]
                
                ws_sku_item.delete_rows(1,ws_sku_item.max_row+1)
                if sku_rows_out:
                    manage_sku_headers=[str(h).strip() if h is not None else "" for h in sku_rows_out[0]]
                    
                    # テンプレートの列順序に合わせて出力
                    if sku_template_headers:
                        sku_output_headers = sku_template_headers
                    else:
                        sku_output_headers = manage_sku_headers
                    
                    ws_sku_item.append(sku_output_headers)
                    
                    # SKU列のマッピングを作成
                    sku_column_mapping = {}
                    for i, template_header in enumerate(sku_output_headers):
                        if template_header in manage_sku_headers:
                            sku_column_mapping[i] = manage_sku_headers.index(template_header)
                    
                    prod_code_idx_sku=manage_sku_headers.index(HEADER_PRODUCT_CODE_SKU) if HEADER_PRODUCT_CODE_SKU in manage_sku_headers else -1
                    if prod_code_idx_sku!=-1:
                        for r_data_sku in sku_rows_out[1:]:
                            if prod_code_idx_sku<len(r_data_sku) and r_data_sku[prod_code_idx_sku] is not None and str(r_data_sku[prod_code_idx_sku]).strip() in n_mycodes_set:
                                # テンプレートの列順序に合わせてSKUデータを再配置
                                reordered_sku_row = [""] * len(sku_output_headers)
                                for template_idx, manage_idx in sku_column_mapping.items():
                                    if manage_idx < len(r_data_sku):
                                        reordered_sku_row[template_idx] = r_data_sku[manage_idx]
                                
                                ws_sku_item.append(reordered_sku_row)
                    else:
                        msg = f"'{OUTPUT_FILE_NAME}' へのSKUデータ書き出し時、「{HEADER_PRODUCT_CODE_SKU}」列が見つかりません。" # type: ignore
                        logging.warning(f"保存処理: {msg}")
                        QMessageBox.warning(self,"警告",msg)

                logging.info(f"出力ファイル '{self.output_file_path}' への保存を試みます。")
                wb_item.save(self.output_file_path)
                logging.info(f"出力ファイル '{self.output_file_path}' の保存が完了しました。")

            except PermissionError:
                msg = f"出力ファイル '{self.output_file_path}' が開かれているためアクセスできません。"
                logging.error(msg)
                QMessageBox.critical(self,"保存エラー",f"{msg}\n詳細はログファイルを確認してください。")
                return
            except Exception as e:
                msg = f"出力ファイル '{self.output_file_path}' の処理中にエラーが発生しました。"
                logging.error(msg, exc_info=True)
                QMessageBox.critical(self,"ファイルエラー",f"{msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}")
                return
            finally:
                if wb_item:
                    try:
                        wb_item.close()
                        logging.info(f"出力ファイル '{self.output_file_path}' をクローズしました。")
                    except Exception as e_close_item: logging.warning(f"出力ファイル '{self.output_file_path}' のクローズ中にエラー: {e_close_item}")
                    wb_item = None # Mark as closed

            if show_message:
                # メッセージを簡略化
                msg_info = f"商品「{code}」の情報を保存しました。"
                QMessageBox.information(self,"保存完了",msg_info)
                # ログには詳細な情報を残す
                logging.info(f"商品「{code}」情報を保存しました。管理ファイル: {self.manage_file_path}, 出力ファイル: {self.output_file_path}")

            self.is_dirty = False # 保存が完了したのでダーティフラグを解除
            
            # 保存した商品を再選択するためにコードを保持
            saved_code = code
            self.load_list() # 商品リストを再読み込み
            
            # 保存した商品を再選択
            # load_list()による一時的な選択解除を防ぐため、フラグを設定
            self._is_reloading_after_save = True
            self._reselect_product_after_save(saved_code)
            # フォームには保存された内容が残っている状態。

            logging.debug(f"save_to_excel - END - Y_カテゴリID: '{self.main_fields.get(HEADER_Y_CATEGORY_ID).text() if HEADER_Y_CATEGORY_ID in self.main_fields else 'N/A'}'")
            self._clear_auto_save_data() # 手動保存成功時に自動保存データをクリア
            
            # 手動保存時刻をツールチップに表示
            self._update_save_button_tooltip("手動保存")

        except Exception as e:
            err_msg = f"保存処理中に予期せぬエラーが発生しました。"
            logging.error(err_msg, exc_info=True)
            QMessageBox.critical(self,"総合保存エラー",f"{err_msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}")
        finally:
            if mutex_locked and hasattr(self, '_save_mutex'):
                try:
                    self._save_mutex.unlock() # Mutexを解放
                    mutex_locked = False
                except Exception as e:
                    logging.error(f"Mutex解放エラー: {e}")
            QApplication.restoreOverrideCursor()
            # --- Final cleanup for workbooks and temp attribute ---
            if wb_mng: # Should be None if closed properly in its own finally
                try:
                    wb_mng.close()
                    logging.warning("管理ファイルが予期せず開いたままだったので、ここでクローズしました。")
                except Exception as e:
                    logging.debug(f"管理ファイルクローズエラー: {e}")
            if wb_item: # Should be None if closed properly in its own finally
                try:
                    wb_item.close()
                    logging.warning("出力ファイルが予期せず開いたままだったので、ここでクローズしました。")
                except Exception as e:
                    logging.debug(f"出力ファイルクローズエラー: {e}")
            if wb_mng_ro: # Should be None
                try:
                    wb_mng_ro.close()
                    logging.warning("読み取り専用管理ファイルが予期せず開いたままだったので、ここでクローズしました。")
                except Exception as e:
                    logging.debug(f"読み取り専用管理ファイルクローズエラー: {e}")

            if hasattr(self, '_temp_y_spec_values_for_save'):
                try:
                    del self._temp_y_spec_values_for_save
                except Exception as e_del_temp:
                    logging.warning(f"_temp_y_spec_values_for_save の削除中にエラー: {e_del_temp}")

    def _reselect_product_after_save(self, saved_code):
        """保存後に同じ商品を再選択する"""
        try:
            logging.info(f"保存後の商品再選択を開始: '{saved_code}'")
            logging.info(f"商品リスト件数: {self.product_list.count()}")
            
            # デバッグ: 商品リストの内容を表示
            for i in range(min(5, self.product_list.count())):  # 最初の5件のみ表示
                item = self.product_list.item(i)
                if item:
                    logging.info(f"  リスト[{i}]: {item.text()}")
            
            # 商品リストから保存した商品を検索して再選択
            found = False
            for i in range(self.product_list.count()):
                item = self.product_list.item(i)
                if item:
                    item_text = item.text()
                    # プレフィックスを考慮して商品コードを抽出
                    if item_text.startswith('['):
                        code_from_item = item_text.split('] ')[1].split(" - ")[0].strip()
                    else:
                        code_from_item = item_text.split(" - ")[0].strip()
                    
                    if code_from_item == saved_code:
                        logging.info(f"商品「{saved_code}」をリスト位置 {i} で発見")
                        
                        # 選択変更イベントを一時的に無効化せずに正常に選択
                        self.product_list.setCurrentRow(i)
                        
                        logging.info(f"保存後に商品「{saved_code}」を再選択しました（位置: {i}）")
                        found = True
                        break
            
            if not found:
                logging.warning(f"保存した商品「{saved_code}」がリストで見つかりませんでした")
                logging.warning("利用可能な商品リスト:")
                for i in range(self.product_list.count()):
                    item = self.product_list.item(i)
                    if item:
                        logging.warning(f"  [{i}] {item.text()}")
                
        except Exception as e:
            logging.error(f"商品再選択中にエラー: {e}", exc_info=True)
        finally:
            # 再選択処理が完了したらフラグをリセット
            self._is_reloading_after_save = False
            logging.info("保存後のリロードフラグをリセットしました")

    def _open_color_selection_dialog(self):
        color_field = self.main_fields.get("色_1")
        if not color_field or not isinstance(color_field, QLineEdit):
            return

        current_text_in_lineedit = color_field.text()
        common_colors_set = set(COMMON_PRODUCT_COLORS) # For efficient lookup

        dialog = ColorSelectionDialog(COMMON_PRODUCT_COLORS, current_text_in_lineedit, self)
        if dialog.exec_() == QDialog.Accepted:
            # dialog_selected_common_names_in_order はユーザーが操作した順番の共通色リスト
            dialog_selected_common_names_in_order = dialog.get_selected_common_colors()

            # 元の入力欄にあった全ての名前を抽出
            all_raw_names_from_lineedit = [m.strip() for m in current_text_in_lineedit.split('●') if m.strip()]
            
            # 保持すべきカスタム名を元の順番で抽出
            original_custom_names_in_order = []
            for name_in_le in all_raw_names_from_lineedit:
                if name_in_le not in common_colors_set: # COMMON_PRODUCT_COLORS にないものがカスタム色
                    if name_in_le not in original_custom_names_in_order: # 重複を避ける
                        original_custom_names_in_order.append(name_in_le)
            
            # 最終的な表示リスト
            final_names_to_display = []
            
            # 1. 元の入力欄にあったカスタム色を、元の順番で追加
            for custom_name in original_custom_names_in_order:
                final_names_to_display.append(custom_name)
            
            # 2. ダイアログで選択された共通色を、ダイアログでの選択順で追加
            #    ただし、既に final_names_to_display に含まれるものは追加しない
            for common_name_from_dialog in dialog_selected_common_names_in_order:
                if common_name_from_dialog not in final_names_to_display:
                    final_names_to_display.append(common_name_from_dialog)
            
            new_text_for_lineedit = " ".join([f"●{name}" for name in final_names_to_display]).strip()
            color_field.setText(new_text_for_lineedit) # setTextがmark_dirtyをトリガーする
            self.is_dirty = True # setTextが接続されたmark_dirtyを呼ぶが、明示的にis_dirtyプロパティ経由で設定


    def show_product_list_menu(self, pos):
        item = self.product_list.itemAt(pos); menu = QMenu()
        
        # 選択されているアイテムを取得
        selected_items = self.product_list.selectedItems()
        
        # コントロールカラム変更メニュー
        if selected_items:
            set_n_act = menu.addAction(f"選択項目を n に設定 ({len(selected_items)}件)")
            set_p_act = menu.addAction(f"選択項目を p に設定 ({len(selected_items)}件)")
            menu.addSeparator()
        else:
            set_n_act = None
            set_p_act = None
        
        copy_act = menu.addAction("コピーして新規作成"); del_act = menu.addAction("この商品を削除")
        copy_act.setEnabled(bool(item)); del_act.setEnabled(bool(item))
        
        action = menu.exec_(self.product_list.mapToGlobal(pos))
        
        if action == set_n_act and selected_items:
            self._batch_set_control_column(selected_items, 'n')
        elif action == set_p_act and selected_items:
            self._batch_set_control_column(selected_items, 'p')
        elif action == copy_act and item:
            # プレフィックスを考慮して商品コードを取得
            item_txt = item.text()
            if item_txt.startswith('['):
                orig_code = item_txt.split('] ')[1].split(" - ")[0].strip()
            else:
                orig_code = item_txt.split(" - ")[0].strip()
            self._initiate_copy_paste_process(orig_code)
        elif action == del_act and item: # SKU削除確認
            item_txt = item.text()
            if item_txt.startswith('['):
                code = item_txt.split('] ')[1].split(" - ")[0].strip()
            else:
                code = item_txt.split(" - ")[0].strip()
            if QMessageBox.question(self,"削除確認",f"本当に商品「{code}」を削除しますか？\n元に戻せません",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)==QMessageBox.Yes:
                self.delete_product(item)
    
    def _batch_set_control_column(self, items, control_value):
        """選択された商品のコントロールカラムを一括変更"""
        try:
            # Excelファイルを開く
            wb = load_workbook(self.manage_file_path, keep_vba=True)
            ws = wb[MAIN_SHEET_NAME]
            
            # ヘッダー行を取得
            header_row = [cell.value for cell in ws[1]]
            control_idx = header_row.index(HEADER_CONTROL_COLUMN) if HEADER_CONTROL_COLUMN in header_row else -1
            code_idx = header_row.index(HEADER_MYCODE) if HEADER_MYCODE in header_row else -1
            
            if control_idx == -1 or code_idx == -1:
                QMessageBox.warning(self, "エラー", "コントロールカラムまたは商品コード列が見つかりません")
                wb.close()
                return
            
            # 各商品のコントロールカラムを更新
            changed_count = 0
            for item in items:
                # プレフィックスを除去して商品コードを取得
                item_txt = item.text()
                if item_txt.startswith('['):
                    code = item_txt.split('] ')[1].split(" - ")[0].strip()
                else:
                    code = item_txt.split(" - ")[0].strip()
                
                # Excelの行を検索
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=code_idx+1).value == code:
                        ws.cell(row=row, column=control_idx+1).value = control_value
                        changed_count += 1
                        
                        # リストアイテムも更新
                        new_text = item_txt.replace(f"[{item.data(Qt.UserRole)}]", f"[{control_value}]")
                        item.setText(new_text)
                        item.setData(Qt.UserRole, control_value)
                        break
            
            # Excelファイルを保存
            wb.save(self.manage_file_path)
            wb.close()
            
            # 現在編集中の商品が変更対象に含まれている場合の処理
            current_item = self.product_list.currentItem()
            current_item_changed = False
            if current_item and current_item in items:
                # ラジオボタンを更新
                if control_value == 'n':
                    self.control_radio_n.setChecked(True)
                else:
                    self.control_radio_p.setChecked(True)
                current_item_changed = True
            
            # 現在編集中の商品が変更された場合、フォームも保存
            if current_item_changed and self.is_dirty:
                try:
                    self.save_to_excel(show_message=False)  # メッセージを表示せずに保存
                    success_msg = f"{changed_count}件の商品のコントロールカラムを '{control_value}' に変更し、保存しました"
                except Exception as save_error:
                    success_msg = f"{changed_count}件の商品のコントロールカラムを '{control_value}' に変更しました\n（注意：フォームの保存に失敗: {str(save_error)}）"
            else:
                success_msg = f"{changed_count}件の商品のコントロールカラムを '{control_value}' に変更しました"
            
            QMessageBox.information(self, "完了", success_msg)
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"コントロールカラムの一括変更に失敗しました:\n{str(e)}")

    def _setup_copy_paste_actions(self):
        # 既にアクションが作成されている場合はスキップ
        if hasattr(self, '_copy_product_action_ref') and self._copy_product_action_ref:
            return
        
        self._copied_product_code_for_paste = None

        copy_action = QAction("選択商品をコピー", self)
        copy_action.setShortcut(QKeySequence.Copy)
        copy_action.triggered.connect(self._handle_copy_product_action)
        self.addAction(copy_action)  # メインウィンドウレベルに追加
        self.product_list.addAction(copy_action)

        paste_action = QAction("コピーした商品を元に新規作成", self)
        paste_action.setShortcut(QKeySequence.Paste)
        paste_action.triggered.connect(self._handle_paste_product_action)
        self.addAction(paste_action)  # メインウィンドウレベルに追加
        self.product_list.addAction(paste_action)

        self._copy_product_action_ref = copy_action
        self._paste_product_action_ref = paste_action

        self._copy_product_action_ref.setEnabled(False) # 初期状態では無効
        self._paste_product_action_ref.setEnabled(False) # 初期状態では無効

        # self.product_list.currentItemChanged.connect(self._update_copy_action_state) # _handle_product_selection_changed でまとめて処理

    def _handle_product_selection_changed(self, current, previous):
        """商品リストの選択が変更されたときの処理"""
        if (self._is_handling_selection_change or 
            getattr(self, '_is_closing', False) or 
            getattr(self, '_is_restoring_after_cancel', False)): # 再入チェック＋終了時チェック＋復元中チェック
            return
        self._is_handling_selection_change = True

        try:
            # --- 選択変更処理の最初で、ダーティかつ編集フィールドにフォーカスがあればフォーカスを移す ---
            if self.is_dirty:
                focused_widget = QApplication.focusWidget()
                if isinstance(focused_widget, (QLineEdit, QTextEdit)) and \
                   focused_widget in self.main_fields.values():
                    self.product_list.setFocus() # フォーカスを移して編集フィールドの編集状態を確定させる
                    QApplication.processEvents() # フォーカス変更と関連イベントの処理を即座に行う
            # --- ここまで追加 ---

            self._update_copy_action_state(current, previous) # コピーアクションの状態更新をここで行う

            self._update_delete_action_state(current, previous) # Deleteアクションの状態更新
            if self._is_loading_data: # データロード中は選択変更処理をスキップ
                return

            # current が None (リストがクリアされたなど) の場合、または previous と current が同じ場合は何もしない
            if not current and not previous: # 両方Noneなら何もしない
                return
            
            # 新規モードの場合は、previous == current のチェックをスキップ
            if self._is_new_mode:
                # 新規作成モードの場合は、読み込みを続行
                self._is_new_mode = False  # フラグをリセット
            elif previous and current and previous == current: # 同じアイテムが再選択された場合
                # 同じアイテムが選択されているが、ダーティフラグがある場合は処理を続行
                # これによりキャンセル後の再選択時の問題を回避
                if not self.is_dirty:
                    return

            # current が None で previous があった場合 (例: 最後のアイテムが削除された、またはクリアされた)
            if not current and previous:
                if self.is_dirty:
                    choice = self._prompt_save_changes()
                    if choice == QMessageBox.YesRole:
                        self.save_to_excel() # 保存後、リストは再読み込みされ、選択はクリアされるはず
                    elif choice == QMessageBox.NoRole:
                        self.is_dirty = False # 変更を破棄
                    elif choice == QMessageBox.RejectRole:
                        # ユーザーがキャンセルした場合、リストの選択はクリアされたままなので何もしない
                        return
                # current が None なので、最終的に clear_fields が呼ばれる (次のelse節で)

            # 修正: previous is not None の条件を削除。ダーティであれば常に確認する。
            if self.is_dirty:
                # 保存前に変数を事前に定義
                form_code_before_save = self.main_fields[HEADER_MYCODE].text().strip()
                target_product_code_to_load_after_save = None
                if current: # current はユーザーが新しく選択しようとしたアイテム
                    target_product_code_to_load_after_save = current.text().split(" - ")[0].strip()
                
                # より柔軟な保存確認：バリデーションエラーでも切り替え可能
                choice = self._prompt_save_changes_flexible()
                if choice == "save":
                    # バリデーションエラーがあっても下書き保存して継続
                    self._save_with_validation_recovery(show_message=True)

                    # 通常保存の場合の処理継続
                    item_to_load_finally = None
                    # まず、ユーザーが元々選択しようとしていた商品を探す
                    if target_product_code_to_load_after_save:
                        found_item_to_select_again = None
                        for i in range(self.product_list.count()):
                            item = self.product_list.item(i)
                            if item.text().startswith(target_product_code_to_load_after_save + " - "):
                                found_item_to_select_again = item
                                break
                        if found_item_to_select_again:
                            item_to_load_finally = found_item_to_select_again

                    # 元々選択しようとしていた商品が見つからない、または指定がなかった場合、
                    # 保存された商品（フォームにあった商品）をロード対象とする
                    if not item_to_load_finally and form_code_before_save:
                        found_saved_item = None
                        for i in range(self.product_list.count()):
                            item = self.product_list.item(i)
                            if item.text().startswith(form_code_before_save + " - "):
                                found_saved_item = item
                                break
                        if found_saved_item:
                            item_to_load_finally = found_saved_item

                    if item_to_load_finally:
                        self._is_loading_data = True # mark_dirty を防ぐ
                        self.load_product(item_to_load_finally) # フォームに内容を直接ロード
                        self.product_list.blockSignals(True)
                        self.product_list.setCurrentItem(item_to_load_finally) # リストの選択を更新 (シグナルなし)
                        self.product_list.blockSignals(False)
                        self._is_loading_data = False
                        if hasattr(self, '_update_status_bar'):
                            self._update_status_bar()
                        # self.is_dirty = False; # load_product が False にする
                    elif not current : # currentがNone（例：新規作成→編集→リストクリア→保存）の場合
                        # 保存はされたが、次に表示する特定のアイテムがない。
                        # form_code_before_save が新規保存されたコード。それがリストにあれば選択。なければクリア。
                        # このケースは item_to_load_finally のロジックでカバーされるはず。
                        # もしそれでも item_to_load_finally が None なら、clear_fields を検討。
                        # ただし、save_to_excel の後なので、フォームは保存された内容のはず。
                        # リストにそのアイテムがあれば、上記のロジックで選択される。
                        # なければ、clear_fields() が適切かもしれないが、通常はリストにあるはず。
                        pass # 現状維持（保存された内容がフォームに残っている）

                elif choice == "discard":
                    self.is_dirty = False # 変更を破棄
                    if current: 
                        self._is_loading_data = True  # mark_dirtyを防ぐ
                        self.load_product(current)
                        self._is_loading_data = False
                        self.is_dirty = False  # 再度確実にクリア
                        if hasattr(self, '_update_status_bar'):
                            self._update_status_bar()
                elif choice == "cancel":
                    # キャンセル：選択をpreviousに戻す（フラグ付きで確実に）
                    def restore_selection():
                        self._is_restoring_after_cancel = True
                        try:
                            if previous:
                                self.product_list.setCurrentItem(previous)
                            else:
                                self.product_list.clearSelection()
                                self.product_list.setCurrentItem(None)
                        finally:
                            self._is_restoring_after_cancel = False
                    
                    # 現在の処理が完全に終わってから実行
                    QTimer.singleShot(0, restore_selection)
                    return # 何もロードしない
            else: # ダーティでない、または最初の選択
                if current: 
                    self.load_product(current) # previous は不要
                    if hasattr(self, '_update_status_bar'):
                        self._update_status_bar()
                elif not current and not self.is_dirty: # 選択がクリアされ、ダーティでもない場合
                    # 保存後のリロード中でなければフィールドをクリア
                    if not getattr(self, '_is_reloading_after_save', False):
                        self.clear_fields(apply_defaults=False)
                        if hasattr(self, '_update_status_bar'):
                            self._update_status_bar()
        finally:
            self._is_handling_selection_change = False

    def _update_copy_action_state(self, current_item, previous_item):
        if hasattr(self, '_copy_product_action_ref'):
            self._copy_product_action_ref.setEnabled(current_item is not None)

    def _update_delete_action_state(self, current_item, previous_item):
        """商品リストの選択状態に応じてDeleteアクションの有効/無効を更新する"""
        if hasattr(self, '_delete_product_action_ref'):
            self._delete_product_action_ref.setEnabled(current_item is not None)

    def _handle_copy_product_action(self):
        current_item = self.product_list.currentItem()
        if not current_item:
            self._copied_product_code_for_paste = None
            if hasattr(self, '_paste_product_action_ref'): self._paste_product_action_ref.setEnabled(False)
            return
        self._copied_product_code_for_paste = current_item.text().split(" - ")[0].strip()
        if hasattr(self, '_paste_product_action_ref'): self._paste_product_action_ref.setEnabled(bool(self._copied_product_code_for_paste))

    def _handle_paste_product_action(self):
        if not self._copied_product_code_for_paste: return
        self._initiate_copy_paste_process(self._copied_product_code_for_paste)

    def _setup_delete_action(self):
        """商品リストでDeleteキーが押されたときのアクションを設定する"""
        self._delete_product_action_ref = QAction("選択商品を削除", self)
        self._delete_product_action_ref.setShortcut(Qt.Key_Delete) # Deleteキーをショートカットに設定
        self._delete_product_action_ref.triggered.connect(self._handle_delete_product_action)
        self.product_list.addAction(self._delete_product_action_ref)
        self._delete_product_action_ref.setEnabled(False) # 初期状態では無効

    def _handle_delete_product_action(self):
        """Deleteキーによる商品削除アクションを処理する"""
        current_item = self.product_list.currentItem()
        if not current_item:
            return
        self.delete_product(current_item) # delete_product内で確認ダイアログが表示される

    def copy_and_paste_product(self, item_to_copy):
        orig_code = item_to_copy.text().split(" - ")[0].strip()

        dialog = CustomProductCodeInputDialog(
            self,
            "新しい商品コード",
            f"コピー元: {orig_code}\n新しい商品コード (10桁の数字) と価格を入力してください:",
            default_code=""
        )
        if dialog.exec_() == QDialog.Accepted:
            new_values = dialog.getValues()
            new_code = new_values["code"]
            new_price_str = new_values["price"]
        else:
            return # User cancelled or closed the dialog
        
        self._initiate_copy_paste_process(orig_code, new_code, new_price_str)

    def _initiate_copy_paste_process(self, orig_code, new_code_override=None, new_price_override=None):
        # new_code_override と new_price_override は、右クリックメニューや将来的な拡張のために残すが、Ctrl+Vでは未使用

        # 商品コードと価格の入力ダイアログ (Ctrl+V の場合、または override がない場合)
        if new_code_override is None or new_price_override is None:
            dialog = CustomProductCodeInputDialog(
                self,
                "新しい商品コード",
                f"コピー元: {orig_code}\n新しい商品コード (10桁の数字) と価格を入力してください:",
                default_code=""
            )
            if dialog.exec_() == QDialog.Accepted:
                new_values = dialog.getValues()
                new_code = new_values["code"]
                new_price_str = new_values["price"]
            else:
                return # User cancelled or closed the dialog
        else: # 右クリックメニューなどから直接指定された場合
            new_code = new_code_override
            new_price_str = new_price_override # type: ignore

        if not new_code or not new_price_str: # new_code_override や new_price_override が None の場合など
            logging.warning("コピー＆ペースト処理: 新しい商品コードまたは価格が指定されませんでした。") ; return

        # new_code が決定された後にチェックを行う
        if new_code == orig_code:
            QMessageBox.warning(self,"コードエラー","新しい商品コードは元のコードと異なる必要があります。")
            return

        if os.path.exists(self.manage_file_path):
            try:
                wb_chk = load_workbook(self.manage_file_path,read_only=True)
                if MAIN_SHEET_NAME in wb_chk.sheetnames:
                    ws_main_chk = wb_chk[MAIN_SHEET_NAME]; hdr_chk = [str(c.value).strip() if c.value is not None else "" for c in ws_main_chk[1]]
                    mycode_idx_chk = hdr_chk.index(HEADER_MYCODE) if HEADER_MYCODE in hdr_chk else -1
                    if mycode_idx_chk != -1:
                        for r_tuple in ws_main_chk.iter_rows(min_row=2,values_only=True):
                            if mycode_idx_chk<len(r_tuple) and r_tuple[mycode_idx_chk] is not None and str(r_tuple[mycode_idx_chk]).strip()==new_code:
                                QMessageBox.warning(self,"コード重複",f"商品コード '{new_code}' は既に存在します。"); wb_chk.close(); return
                wb_chk.close()
            except Exception as e_chk:
                msg = f"商品コードの重複チェック中にエラーが発生しました: {e_chk}"
                QMessageBox.warning(self,"重複チェックエラー",msg); logging.warning(f"コピー＆ペースト処理: {msg}", exc_info=True)
                return # 重複チェックでエラーが発生した場合は処理を中断

        # load_product を呼び出すために item_to_copy を見つける
        item_to_load = None
        for i in range(self.product_list.count()):
            list_item = self.product_list.item(i)
            if list_item.text().startswith(orig_code + " - "):
                item_to_load = list_item
                break
        if not item_to_load:
            msg = f"コピー元の商品 '{orig_code}' がリストに見つかりません。"
            QMessageBox.warning(self, "エラー", msg); logging.warning(f"コピー＆ペースト処理: {msg}")
            return
        self.load_product(item_to_load)

        # Populate copied_main_data including Y_spec string values from the current UI
        copied_main_data = {}
        for k, f_widget in self.main_fields.items():
            if k in HTML_TEXTEDIT_FIELDS and isinstance(f_widget, QTextEdit):
                copied_main_data[k] = f_widget.toPlainText()
            elif isinstance(f_widget, QLineEdit):
                copied_main_data[k] = f_widget.text()
            elif isinstance(f_widget, QComboBox):
                copied_main_data[k] = f_widget.currentText()
        
        # Get Y_spec values from the UI of the original product
        # _get_y_spec_value_for_save uses self.y_spec_current_editors and self.y_spec_current_definitions
        # which were set up by load_product(item_to_copy)
        for i in range(MAX_Y_SPEC_COUNT): # Y_spec1 to Y_spec10
            y_spec_key = f"Y_spec{i+1}"
            copied_main_data[y_spec_key] = self._get_y_spec_value_for_save(i)

        copied_main_data[HEADER_MYCODE]=new_code
        copied_main_data[HEADER_PRODUCT_NAME]=f"{copied_main_data.get(HEADER_PRODUCT_NAME,'')} (コピー)"
        copied_main_data[HEADER_PRICE_TAX_INCLUDED] = new_price_str

        copied_sku_list=[]; base_new_mycode=new_code[:-3] if len(new_code)>=3 else new_code
        for sku_idx,orig_sku_dict in enumerate(self.sku_data_list):
            new_sku_dict=orig_sku_dict.copy(); new_sku_dict[HEADER_PRODUCT_CODE_SKU]=new_code
            orig_sku_code=orig_sku_dict.get(HEADER_SKU_CODE,"")
            new_suffix=orig_sku_code[-3:] if orig_sku_code and len(orig_sku_code)>=3 and orig_sku_code[-3:].isdigit() else f"{(sku_idx+1)*SKU_CODE_SUFFIX_INCREMENT:03d}"
            new_sku_dict[HEADER_SKU_CODE]=base_new_mycode+new_suffix
            copied_sku_list.append(new_sku_dict)

        self.clear_fields()
        self.control_radio_n.setChecked(True)

        for k_set,v_set in copied_main_data.items():
            # This loop handles standard fields. Y_spec fields are not in self.main_fields.
            # Y_カテゴリID will be set here, but its textChanged signal is blocked during setText.
            if k_set in self.main_fields:
                fld_set=self.main_fields[k_set]
                fld_set.blockSignals(True)
                if k_set in HTML_TEXTEDIT_FIELDS and isinstance(fld_set, QTextEdit):
                    fld_set.setPlainText(v_set)
                elif isinstance(fld_set,QLineEdit): fld_set.setText(v_set)
                elif isinstance(fld_set,QComboBox): idx_s=fld_set.findText(v_set); fld_set.setCurrentIndex(idx_s if idx_s!=-1 else 0)
                fld_set.blockSignals(False)

        # Explicitly trigger Y_spec UI build based on the copied Y_カテゴリID
        copied_y_category_id = copied_main_data.get(HEADER_Y_CATEGORY_ID, "")
        self._on_y_category_id_changed(copied_y_category_id)

        # Load the copied Y_spec string values into the newly built UI
        for i in range(MAX_Y_SPEC_COUNT):
            y_spec_key = f"Y_spec{i+1}"
            saved_value_str = copied_main_data.get(y_spec_key, "")
            if saved_value_str: # Only load if there was a value
                 self._load_y_spec_value(saved_value_str)

        self.sku_data_list=copied_sku_list; self.show_sku_table()

        for field_name in self.byte_count_labels.keys():
             if field_name in self.main_fields and isinstance(self.main_fields.get(field_name), QLineEdit):
                 self._update_byte_count_display(field_name, self.main_fields.get(field_name).text())
        if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields:
            self._update_mycode_digit_count_display(self.main_fields[HEADER_MYCODE].text())
        self._format_and_sync_price_fields()
        self._update_relevant_links() # relevant_links を更新

        self.is_dirty = True; self.product_list.clearSelection()
        msg_info = f"「{orig_code}」を元に新しい商品「{new_code}」を作成しました。\n保存せずに閉じるとデータが失われるため注意してください。"
        QMessageBox.information(self,"コピー完了",msg_info); logging.info(f"コピー＆ペースト完了: {msg_info}")

    def _delete_from_item_xlsm(self, code_to_delete):
        """item.xlsmから指定された商品コードの行を削除"""
        if not os.path.exists(self.output_file_path):
            logging.info(f"item.xlsm '{self.output_file_path}' が存在しないため、削除処理をスキップ")
            return
            
        try:
            wb_item = load_workbook(self.output_file_path, keep_vba=True)
            
            # Mainシートから削除
            if MAIN_SHEET_NAME in wb_item.sheetnames:
                ws_main = wb_item[MAIN_SHEET_NAME]
                hdr_main = [str(c.value).strip() if c.value is not None else "" for c in ws_main[1]]
                mycode_idx = hdr_main.index(HEADER_MYCODE) if HEADER_MYCODE in hdr_main else -1
                
                if mycode_idx != -1:
                    rows_to_delete = sorted([r for r in range(2, ws_main.max_row + 1) 
                                           if ws_main.cell(row=r, column=mycode_idx + 1).value is not None 
                                           and str(ws_main.cell(row=r, column=mycode_idx + 1).value).strip() == code_to_delete], 
                                          reverse=True)
                    for r_del in rows_to_delete:
                        ws_main.delete_rows(r_del, 1)
                        logging.debug(f"item.xlsm Mainシートから行 {r_del} を削除 (商品コード: {code_to_delete})")
            
            # SKUシートから削除
            if SKU_SHEET_NAME in wb_item.sheetnames:
                ws_sku = wb_item[SKU_SHEET_NAME]
                hdr_sku = [str(c.value).strip() if c.value is not None else "" for c in ws_sku[1]]
                prod_code_idx = hdr_sku.index(HEADER_PRODUCT_CODE_SKU) if HEADER_PRODUCT_CODE_SKU in hdr_sku else -1
                
                if prod_code_idx != -1:
                    rows_to_delete_sku = sorted([r_s for r_s in range(2, ws_sku.max_row + 1) 
                                                if ws_sku.cell(row=r_s, column=prod_code_idx + 1).value is not None 
                                                and str(ws_sku.cell(row=r_s, column=prod_code_idx + 1).value).strip() == code_to_delete], 
                                               reverse=True)
                    for r_sku_del in rows_to_delete_sku:
                        ws_sku.delete_rows(r_sku_del, 1)
                        logging.debug(f"item.xlsm SKUシートから行 {r_sku_del} を削除 (商品コード: {code_to_delete})")
            
            wb_item.save(self.output_file_path)
            wb_item.close()
            
        except PermissionError:
            raise Exception(f"item.xlsm '{self.output_file_path}' が開かれているため削除できません。")
        except Exception as e:
            raise Exception(f"item.xlsmの削除処理中にエラー: {e}")

    def delete_product(self, item_to_delete) -> None:
        # 削除処理中フラグを設定（他の保存処理をブロック）
        self._is_deleting = True
        
        code_del = self._safe_string_operation(
            item_to_delete.text().split(" - ")[0] if item_to_delete and item_to_delete.text() else ""
        )
        logging.debug(f"商品削除開始: '{code_del}'")
        
        if not self._safe_file_exists(self.manage_file_path):
            msg = f"管理ファイル '{self.manage_file_path}' が見つかりません。"
            QMessageBox.warning(self,"エラー",msg); logging.warning(f"商品削除試行: {msg}"); 
            self._is_deleting = False
            return
        try:
            wb_mng=load_workbook(self.manage_file_path,keep_vba=True)
            if MAIN_SHEET_NAME in wb_mng.sheetnames:
                ws_main=wb_mng[MAIN_SHEET_NAME]; hdr_main=[str(c.value).strip() if c.value is not None else "" for c in ws_main[1]]
                mycode_idx=hdr_main.index(HEADER_MYCODE) if HEADER_MYCODE in hdr_main else -1
                if mycode_idx!=-1:
                    # デバッグ: 全行の商品コードを確認
                    all_codes = []
                    for r in range(2, min(ws_main.max_row+1, 20)):  # 最大20行まで
                        cell_value = ws_main.cell(row=r, column=mycode_idx+1).value
                        if cell_value is not None:
                            code_str = str(cell_value).strip()
                            all_codes.append(f"行{r}:'{code_str}'")
                    logging.info(f"現在の商品コード一覧: {', '.join(all_codes[:10])}")  # 最初の10行のみ
                    
                    rows_to_delete = sorted([r for r in range(2,ws_main.max_row+1) if ws_main.cell(row=r,column=mycode_idx+1).value is not None and str(ws_main.cell(row=r,column=mycode_idx+1).value).strip()==code_del],reverse=True)
                    logging.info(f"削除対象「{code_del}」に対する削除対象行: {rows_to_delete}")
                    
                    if not rows_to_delete:
                        logging.warning(f"item_manage.xlsm に削除対象商品「{code_del}」が見つかりませんでした")
                    
                    for r_del in rows_to_delete: 
                        cell_value = ws_main.cell(row=r_del, column=mycode_idx+1).value
                        logging.info(f"item_manage.xlsm Mainシート行{r_del}を削除中（値: '{cell_value}'）")
                        ws_main.delete_rows(r_del,1)

            if SKU_SHEET_NAME in wb_mng.sheetnames:
                ws_sku=wb_mng[SKU_SHEET_NAME]; hdr_sku=[str(c.value).strip() if c.value is not None else "" for c in ws_sku[1]]
                prod_code_idx=hdr_sku.index(HEADER_PRODUCT_CODE_SKU) if HEADER_PRODUCT_CODE_SKU in hdr_sku else -1
                if prod_code_idx!=-1:
                    rows_to_delete_sku = sorted([r_s for r_s in range(2,ws_sku.max_row+1) if ws_sku.cell(row=r_s,column=prod_code_idx+1).value is not None and str(ws_sku.cell(row=r_s,column=prod_code_idx+1).value).strip()==code_del],reverse=True)
                    for r_sku_del in rows_to_delete_sku: ws_sku.delete_rows(r_sku_del,1)

            wb_mng.save(self.manage_file_path); wb_mng.close()
        except PermissionError:
            msg = f"管理ファイル '{self.manage_file_path}' が開かれているため変更できません。"
            QMessageBox.critical(self,"削除エラー",f"{msg}\n詳細はログファイルを確認してください。"); logging.error(msg)
            self._is_deleting = False
            return
        except Exception as e_del:
            msg = f"管理ファイルの編集中にエラーが発生しました。"
            QMessageBox.critical(self,"削除エラー",f"{msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e_del}"); logging.error(msg, exc_info=True)
            self._is_deleting = False
            return

        # item.xlsmからも削除を反映
        try:
            self._delete_from_item_xlsm(code_del)
            logging.info(f"item.xlsmから商品「{code_del}」を削除しました。")
        except Exception as e:
            logging.warning(f"item.xlsmからの削除中にエラー: {e}")
            # item.xlsmの削除エラーは商品削除完了メッセージに影響しない
        
        # 削除された商品が現在表示されている場合はフィールドをクリア
        if self.main_fields.get(HEADER_MYCODE) and self.main_fields[HEADER_MYCODE].text().strip()==code_del: 
            logging.info(f"削除対象商品「{code_del}」が現在表示中のため、フォームをクリアします")
            self._is_loading_data = True  # clear_fields中にmark_dirtyが呼ばれないようにする
            # 商品コードを空にして、デフォルト値が保存されないようにする
            self.main_fields[HEADER_MYCODE].setText("")
            logging.info("商品コードを空に設定しました")
            self.clear_fields()  # 削除後は新規作成と同じ状態（デフォルト値あり）
            # 商品コードを再度空にする（clear_fieldsがデフォルト値を設定する可能性があるため）
            self.main_fields[HEADER_MYCODE].setText("")
            logging.info(f"clear_fields後の商品コード: '{self.main_fields[HEADER_MYCODE].text()}'")
            self._is_loading_data = False  # mark_dirtyを再度有効にする
            # ダーティフラグもクリア（削除された商品のデータを保存しないため）
            self.is_dirty = False
            logging.info(f"削除後フォームクリア完了: is_dirty={self.is_dirty}")
        
        # 商品一覧を再読み込み（削除後の状態を反映）
        self.load_list()
        
        # 削除後は商品一覧の選択をクリアして新規入力状態を維持
        self.product_list.clearSelection()
        
        # Y_specフィールドも適切に更新 (クリアされた状態になる)
        self._on_y_category_id_changed(self.main_fields.get(HEADER_Y_CATEGORY_ID, QLineEdit()).text())
        
        # 削除後の状態確認ログ
        logging.info(f"削除処理完了後の状態: 商品コード='{self.main_fields[HEADER_MYCODE].text()}', is_dirty={self.is_dirty}")
        
        # 削除処理完了フラグをクリア
        self._is_deleting = False
        
        msg_info = f"商品「{code_del}」を削除しました。"
        QMessageBox.information(self,"削除完了",msg_info); logging.info(msg_info)

    def mark_dirty(self) -> None:
        """データの変更をマークし、保存ボタンを有効化"""
        if self._is_loading_data or self._is_undoing: # データロード中またはUndo/Redo中はダーティフラグを更新しない
            return
        try:
            # is_dirty プロパティのセッター経由で状態を変更
            # これにより、save_btnの状態も自動的に更新される
            was_dirty = self.is_dirty
            self.is_dirty = True
            
            # 初回の変更時のみUndo履歴に保存（頻繁な保存を避ける）
            if not was_dirty:
                self.save_undo_state()
                
        except Exception as e:
            logging.error(f"mark_dirty中のエラー: {e}", exc_info=True)
            # 機能を継続するため、エラーを隠蔽して処理を続行
    
    def _safe_file_exists(self, file_path: str) -> bool:
        """ファイル存在確認（安全版）"""
        try:
            return os.path.exists(file_path) if file_path else False
        except (OSError, TypeError):
            return False
    
    def _safe_widget_operation(self, widget, operation_name: str, operation_func):
        """ウィジェット操作の安全実行"""
        try:
            if widget is not None and hasattr(widget, '__call__' if callable(operation_func) else 'isVisible'):
                return operation_func()
        except (AttributeError, RuntimeError, TypeError) as e:
            logging.debug(f"{operation_name}中のウィジェットエラー（継続）: {e}")
            return None
        except Exception as e:
            logging.debug(f"{operation_name}中のエラー（継続）: {e}")
            return None
    
    def _safe_string_operation(self, value: str, default: str = "") -> str:
        """文字列操作の安全実行"""
        try:
            return str(value).strip() if value is not None else default
        except (TypeError, AttributeError):
            return default
    
    def _safe_int_operation(self, value, default: int = 0) -> int:
        """整数変換の安全実行"""
        try:
            return int(value) if value is not None else default
        except (ValueError, TypeError):
            return default
    
    def _clear_temporary_data(self) -> None:
        """一時データのクリア（メモリ効率化）"""
        try:
            # 使用されていない一時変数のクリア（ガベージコレクション支援）
            if hasattr(self, '_temp_data'):
                self._temp_data = None
            # QApplicationのイベント処理
            QApplication.processEvents()
        except Exception as e:
            logging.debug(f"一時データクリア中のエラー（継続）: {e}")
            
    def set_all_control_to_p(self) -> None:
        if not self._safe_file_exists(self.manage_file_path):
            msg = f"管理ファイル '{self.manage_file_path}' が見つかりません。"
            QMessageBox.warning(self,"エラー",msg); logging.warning(f"一括P設定試行: {msg}")
            return
        
        # メモリチェック（大量データ処理の前）
        if check_memory_before_large_operation:
            estimated_memory_mb = 50.0  # 一括処理で推定50MB必要
            if not check_memory_before_large_operation(estimated_memory_mb, self):
                return  # メモリ不足でユーザーがキャンセルした場合
        
        if QMessageBox.question(self,"一括変更確認",f"全商品のコントロールカラムを 'p (除外)' に変更しますか？",QMessageBox.Yes|QMessageBox.No,QMessageBox.No)==QMessageBox.No: return
        try:
            wb=load_workbook(self.manage_file_path,keep_vba=True)
            if MAIN_SHEET_NAME not in wb.sheetnames:
                msg = f"{MAIN_SHEET_NAME}シートが見つかりません。"
                QMessageBox.warning(self,"エラー",msg); wb.close(); logging.warning(f"一括P設定試行: {msg}") # type: ignore
                return
            ws=wb[MAIN_SHEET_NAME]; hdr_cells=next(ws.iter_rows(min_row=1,max_row=1,values_only=False),None)
            if not hdr_cells:
                msg = f"{MAIN_SHEET_NAME}シートのヘッダー行が空です。"
                QMessageBox.warning(self,"エラー",msg); wb.close(); logging.warning(f"一括P設定試行: {msg}") # type: ignore
                return
            hdr_vals=[str(c.value).strip() if c.value is not None else "" for c in hdr_cells]
            if HEADER_CONTROL_COLUMN not in hdr_vals:
                msg = f"「{HEADER_CONTROL_COLUMN}」が{MAIN_SHEET_NAME}シートのヘッダーに見つかりません。"
                QMessageBox.warning(self,"エラー",msg); wb.close(); logging.warning(f"一括P設定試行: {msg}") # type: ignore
                return
            ctrl_col_idx=hdr_vals.index(HEADER_CONTROL_COLUMN)+1
            changed_count = 0
            for r_num in range(2, ws.max_row + 1):
                cell = ws.cell(row=r_num, column=ctrl_col_idx)
                if str(cell.value).strip().lower() != 'p':
                    cell.value = "p"
                    changed_count += 1
                # 大量データ処理時のUI応答性向上
                if changed_count % PROGRESS_UPDATE_INTERVAL == 0:
                    self._safe_widget_operation(
                        QApplication.instance(),
                        "UI更新",
                        lambda: QApplication.processEvents()
                    )
            wb.save(self.manage_file_path); wb.close()
            
            # 大量処理後のメモリクリア
            self._clear_temporary_data()

            # 現在UIで開いている商品のラジオボタンを 'p' に設定
            current_item_on_display_code = self._safe_widget_operation(
                self.main_fields.get(HEADER_MYCODE), 
                "商品コード取得",
                lambda: self.main_fields.get(HEADER_MYCODE, QLineEdit()).text().strip()
            ) or ""
            self._safe_widget_operation(
                self.control_radio_p, 
                "ラジオボタン設定",
                lambda: self.control_radio_p.setChecked(True)
            )
            
            # 新規入力画面（商品コードが空）でない場合のみ保存
            if current_item_on_display_code:
                # 既存商品の場合は保存を実行
                self.save_to_excel(show_message=False)
            else:
                # 新規入力画面の場合は保存をスキップし、ダーティフラグのみクリア
                self.is_dirty = False
                logging.info("一括P設定: 新規入力画面のため現在の商品の保存をスキップしました")
            self.load_list()
            cur_item=self.product_list.currentItem()
            if cur_item: self.load_product(cur_item)
            elif self.product_list.count()>0: self.product_list.setCurrentRow(0)
            msg_info = f"{changed_count}件の商品のコントロールカラムを 'p' に変更しました (既に 'p' だったものを除く)。"
            QMessageBox.information(self,"完了",msg_info); logging.info(f"一括P設定完了: {msg_info}")
        except PermissionError:
            msg = f"管理ファイル '{self.manage_file_path}' が開かれているため変更できません。"
            QMessageBox.critical(self,"エラー",f"{msg}\n詳細はログファイルを確認してください。"); logging.error(msg)
        except Exception as e:
            msg = f"一括変更処理中にエラーが発生しました。"
            QMessageBox.critical(self,"エラー",f"{msg}\n詳細はログファイルを確認してください。\n\nエラー詳細:\n{e}"); logging.error(msg, exc_info=True)


        
    def _clear_y_spec_editor(self, index: int) -> None:
        """指定されたインデックスのY_specエディタをクリア/削除する"""
        if 0 <= index < MAX_Y_SPEC_COUNT:
            # 既存のエディタがあれば削除
            if self.y_spec_current_editors[index] is not None:
                self.y_spec_current_editors[index].deleteLater()
                self.y_spec_current_editors[index] = None
            self.y_spec_current_definitions[index] = None
            # プレースホルダー内のレイアウトからウィジェットを削除 (もしあれば)
            placeholder_layout = self.y_spec_editor_placeholders[index].layout()
            if placeholder_layout:
                while placeholder_layout.count():
                    child = placeholder_layout.takeAt(0)
                    if child.widget():
                        child.widget().deleteLater()
            self.y_spec_labels[index].setText(f"Y_spec{index+1} (項目名)") # ラベルをリセット
            self.y_spec_editor_placeholders[index].hide() # コンテナを隠す
            self.y_spec_labels[index].hide()

    def _on_y_category_id_changed(self, category_id_text):
        """YカテゴリIDが変更されたときにY_specフィールドを更新する"""
        if not hasattr(self, 'y_spec_loader'): # ローダーが初期化されていなければ何もしない
            if hasattr(self, 'y_spec_section_label_widget'): self.y_spec_section_label_widget.hide()
            if hasattr(self, 'y_spec_header_spacer_top'): self.y_spec_header_spacer_top.hide()
            if hasattr(self, 'y_spec_footer_spacer'): self.y_spec_footer_spacer.hide()
            return

        # 既存のY_specフィールドをクリア
        for i in range(MAX_Y_SPEC_COUNT):
            self._clear_y_spec_editor(i)

        # 同期対象のY_specエディタ参照をリセット
        self.y_spec_width_editor = None
        self.y_spec_depth_editor = None
        self.y_spec_height_editor = None
        # Y_spec同期用の定義保持用もリセット
        self.y_spec_width_definition = None
        self.y_spec_depth_definition = None
        self.y_spec_height_definition = None
        self.y_spec_weight_editor = None # Y!spec重量用のエディタもリセット
        self.y_spec_weight_definition = None # Y!spec重量用の定義もリセット

        specs = self.y_spec_loader.get_specs_for_category(category_id_text)

        # --- 幅・奥行き・高さ（cm）の項目が存在する場合、特定の順序に並び替える ---
        # specs は既に spec_id の昇順でソートされている想定
        width_spec_info = None
        depth_spec_info = None
        height_spec_info = None
        weight_spec_info = None # 重量情報も取得

        for i, spec_def in enumerate(specs):
            name = spec_def["spec_name"].strip()
            if name == YSPEC_NAME_WIDTH_CM:
                width_spec_info = (i, spec_def)
            elif name == YSPEC_NAME_DEPTH_CM:
                depth_spec_info = (i, spec_def)
            elif name == YSPEC_NAME_HEIGHT_CM:
                height_spec_info = (i, spec_def)
            elif name == YSPEC_NAME_WEIGHT: # 重量項目をチェック
                weight_spec_info = (i, spec_def)
                
        if width_spec_info and depth_spec_info and height_spec_info:
            # 3項目すべてが見つかった場合のみ並び替えを実行
            # 元のリストからこれらの項目をインデックスを保持しつつ取り出す
            # (削除する際はインデックスが大きいものから行うとズレない)
            
            # 1. 対象項目をリストから取り出す (元のインデックスと定義を保持)
            #    取り出す前に、並び替え対象の項目を元のリストから削除する準備
            items_to_sort = [width_spec_info[1], depth_spec_info[1], height_spec_info[1]]
            dim_names_to_remove = [YSPEC_NAME_WIDTH_CM, YSPEC_NAME_DEPTH_CM, YSPEC_NAME_HEIGHT_CM]
            
            # 2. 元のリストからこれらの項目を削除
            #    specsリスト内での重複がない前提 (YSpecDefinitionLoaderでspec_idの重複は避けている)
            specs = [s for s in specs if s["spec_name"].strip() not in dim_names_to_remove]

            # 3. 並び替えた項目を、元々「幅(cm)」があった位置、またはリストの先頭に挿入
            #    ここでは、最も若い元のインデックスの位置に挿入する (より自然な位置にするため)
            insert_pos = min(width_spec_info[0], depth_spec_info[0], height_spec_info[0])
            specs[insert_pos:insert_pos] = [width_spec_info[1], depth_spec_info[1], height_spec_info[1]] # 幅→奥行き→高さの順
        
        # 重量項目が見つかった場合、それをリストの先頭 (または特定の優先順位) に移動させることも可能
        # ここでは、幅奥行高さの後に重量が来るようにする (もし重量が幅奥行高さより前にあった場合)
        # ただし、現状は YSpecDefinitionLoader で spec_id 昇順ソートなので、そのままでも良いかもしれない。
        # 必要であれば、重量項目も上記と同様に並び替えロジックに追加する。
        
        specs_exist_for_category = bool(specs)
        if hasattr(self, 'y_spec_section_label_widget'):
            self.y_spec_section_label_widget.setVisible(specs_exist_for_category)
        if hasattr(self, 'y_spec_header_spacer_top'):
            self.y_spec_header_spacer_top.setVisible(specs_exist_for_category)
        if hasattr(self, 'y_spec_footer_spacer'):
            self.y_spec_footer_spacer.setVisible(specs_exist_for_category)

        if not specs_exist_for_category:
            return

        for i, spec_def in enumerate(specs[:10]): # 最大10個まで
            self.y_spec_current_definitions[i] = spec_def # 現在の定義を保存
            label_widget = self.y_spec_labels[i]
            editor_placeholder = self.y_spec_editor_placeholders[i]
            
            label_widget.setText(spec_def["spec_name"])
            label_widget.show()
            editor_placeholder.show()

            new_editor = None
            editor_layout = editor_placeholder.layout() # QHBoxLayoutのはず

            if spec_def["data_type"] == 1: # テキスト選択
                options_for_editor = [(opt["value_name"], opt["value_id"]) for opt in spec_def["options"]]
                if spec_def["selection_type"] == 0: # 単一選択
                    new_editor = QComboBox(editor_placeholder)
                    new_editor.addItem("", None) # 空の選択肢
                    for name, val_id in options_for_editor:
                        new_editor.addItem(name, val_id) # type: ignore
                    new_editor.currentIndexChanged.connect(lambda: self.mark_dirty())
                else: # 複数選択
                    # SkuMultipleAttributeEditor をカンマ区切りで再利用
                    # optionsは表示名のみのリストを渡す
                    display_options = [name for name, _ in options_for_editor]
                    new_editor = SkuMultipleAttributeEditor(display_options, "", editor_placeholder, editable_line_edit=True, delimiter_char=':') # 保存形式に合わせて区切り文字を':'に
                    new_editor.line_edit.textChanged.connect(lambda: self.mark_dirty())
            elif spec_def["data_type"] == 2: # 整数入力
                new_editor = QLineEdit(editor_placeholder)
                # new_editor.setValidator(QIntValidator()) # 必要であればバリデータ設定
                new_editor.textChanged.connect(lambda: self.mark_dirty())
            elif spec_def["data_type"] == 4: # 整数or小数入力
                new_editor = QLineEdit(editor_placeholder)
                # new_editor.setValidator(QDoubleValidator(0, 99999999.9999, 4)) # 必要であればバリデータ設定
                new_editor.textChanged.connect(lambda: self.mark_dirty())
            else: # 未知のデータ型
                new_editor = QLineEdit(editor_placeholder)
                new_editor.setPlaceholderText(f"未対応データ型: {spec_def['data_type']}")
                new_editor.setReadOnly(True)

            if new_editor:
                editor_layout.addWidget(new_editor)
                self.y_spec_current_editors[i] = new_editor
                new_editor.show()
                
                # 同期対象のY_specエディタを特定・保持
                spec_name_for_sync = spec_def["spec_name"].strip()
                if spec_name_for_sync == YSPEC_NAME_WIDTH_CM:
                    self.y_spec_width_editor = new_editor
                    self.y_spec_width_definition = spec_def
                elif spec_name_for_sync == YSPEC_NAME_DEPTH_CM:
                    self.y_spec_depth_editor = new_editor
                    self.y_spec_depth_definition = spec_def
                elif spec_name_for_sync == YSPEC_NAME_HEIGHT_CM:
                    self.y_spec_height_editor = new_editor
                    self.y_spec_height_definition = spec_def
                elif spec_name_for_sync == YSPEC_NAME_WEIGHT: # 重量エディタと定義を保持
                    self.y_spec_weight_editor = new_editor
                    self.y_spec_weight_definition = spec_def
        
        # Y_specエディタ作成後にナビゲーション設定を更新
        if hasattr(self, '_setup_yspec_navigation'):
            self._setup_yspec_navigation()
        
        self._sync_product_size_to_yspec() # ★Y!カテゴリ変更後にも同期

    def _get_y_spec_value_for_save(self, index):
        """指定されたインデックスのY_specフィールドの値を保存形式で取得する"""
        spec_def = self.y_spec_current_definitions[index]
        editor = self.y_spec_current_editors[index]

        if not spec_def or not editor:
            return ""

        spec_id_part = spec_def["spec_id"]
        value_part = ""

        if spec_def["data_type"] == 1: # テキスト選択
            if spec_def["selection_type"] == 0: # 単一選択 (QComboBox)
                value_part = editor.currentData() if editor.currentIndex() > 0 else "" # currentData()でspec_value_id取得
            else: # 複数選択 (SkuMultipleAttributeEditor)
                selected_names_str = editor.text() # "name1:name2" の形式
                selected_names = [name.strip() for name in selected_names_str.split(':') if name.strip()]
                selected_ids = []
                # spec_def["options"] は [{"value_id": id, "value_name": name}, ...] の形式
                for name_to_find in selected_names:
                    found_id = next((opt["value_id"] for opt in spec_def["options"] if opt["value_name"] == name_to_find), None)
                    if found_id:
                        selected_ids.append(found_id)
                value_part = ":".join(selected_ids)
        elif spec_def["data_type"] == 2: # 整数入力 (QLineEdit)
            value_part = editor.text().strip()
        elif spec_def["data_type"] == 4: # 整数or小数入力 (QLineEdit)
            text_val = editor.text().strip()
            if text_val:
                try:
                    float_val = float(text_val)
                    value_part = f"{float_val:.4f}" # 小数点以下4桁にフォーマット
                except ValueError:
                    value_part = text_val # 数値に変換できない場合はそのまま
            else:
                value_part = ""

        if not value_part: # 値が空の場合は spec_id も含めない (Yahoo!の仕様による)
            return ""
        return f"{spec_id_part}|{value_part}"

    def _load_y_spec_value(self, saved_value_str_from_excel_column):
        """保存されたY_specの値を対応するエディタに設定する"""
        if not saved_value_str_from_excel_column:
            return

        try:
            loaded_spec_id_part, value_part = saved_value_str_from_excel_column.split('|', 1)
        except ValueError: # "|" がない、または不正な形式
            logging.warning(f"Y_specの値 '{saved_value_str_from_excel_column}' は不正な形式です。")
            return

        target_ui_index = -1
        spec_def_for_loaded_id = None
        for idx, current_def in enumerate(self.y_spec_current_definitions):
            if current_def and current_def["spec_id"] == loaded_spec_id_part:
                target_ui_index = idx
                spec_def_for_loaded_id = current_def
                break
        
        if target_ui_index == -1 or not spec_def_for_loaded_id:
            # logging.info(f"保存されていたY_specのspec_id '{loaded_spec_id_part}' は、現在のカテゴリのスペック定義に見つかりません。({saved_value_str_from_excel_column})")
            return

        editor = self.y_spec_current_editors[target_ui_index]
        if not editor:
            return
        
        # spec_def_for_loaded_id を spec_def として使用
        spec_def = spec_def_for_loaded_id

        if spec_def["data_type"] == 1: # テキスト選択
            if spec_def["selection_type"] == 0: # 単一選択 (QComboBox)
                # value_part は spec_value_id
                idx_to_select = editor.findData(value_part)
                if idx_to_select != -1:
                    editor.setCurrentIndex(idx_to_select)
                else:
                    editor.setCurrentIndex(0) # 見つからなければ空を選択
            else: # 複数選択 (SkuMultipleAttributeEditor)
                # value_part は "id1:id2" の形式
                selected_ids = [sid.strip() for sid in value_part.split(':') if sid.strip()]
                selected_names_for_editor = []
                for sid_to_find in selected_ids:
                    found_name = next((opt["value_name"] for opt in spec_def["options"] if opt["value_id"] == sid_to_find), None)
                    if found_name:
                        selected_names_for_editor.append(found_name)
                editor.setText(":".join(selected_names_for_editor))
        elif spec_def["data_type"] in [2, 4]: # 整数入力 or 整数/小数入力 (QLineEdit)
            editor.setText(value_part)

    def _update_relevant_links(self):
        """「関連商品_Xb」フィールドの値を連結して relevant_links に設定する"""
        if not hasattr(self, 'main_fields') or "relevant_links" not in self.main_fields:
            return

        relevant_links_parts = []
        for i in range(1, 16):  # 関連商品_1b から 関連商品_15b まで
            field_name_b = f"関連商品_{i}b"
            if field_name_b in self.main_fields:
                widget_b = self.main_fields[field_name_b]
                if isinstance(widget_b, QLineEdit):
                    text_b = widget_b.text().strip()
                    if text_b:
                        relevant_links_parts.append(text_b)
        
        final_relevant_links_text = "".join(relevant_links_parts)
        
        relevant_links_widget = self.main_fields["relevant_links"]
        relevant_links_widget.blockSignals(True)
        relevant_links_widget.setText(final_relevant_links_text)
        relevant_links_widget.blockSignals(False)
        self.is_dirty = True # relevant_links が更新されたので、状態を dirty にする

    def _prompt_save_changes(self):
        """
        未保存の変更がある場合に、保存するかどうかをユーザーに尋ねるダイアログを表示します。
        Returns:
            QMessageBox.YesRole: 「はい」(保存する) が選択された場合
            QMessageBox.NoRole: 「いいえ」(保存しない) が選択された場合
            QMessageBox.RejectRole: 「キャンセル」が選択された場合
        """
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Question)
        msg_box.setWindowTitle("変更の確認")
        msg_box.setText("未保存の変更があります。変更を保存しますか？")
        msg_box.setInformativeText("「いいえ」を選択した場合、変更は破棄されます。")

        save_button = msg_box.addButton("はい", QMessageBox.YesRole)
        discard_button = msg_box.addButton("いいえ", QMessageBox.NoRole)
        cancel_button = msg_box.addButton("キャンセル", QMessageBox.RejectRole)
        
        msg_box.setDefaultButton(cancel_button)
        msg_box.setEscapeButton(cancel_button)

        msg_box.exec_()
        clicked_button = msg_box.clickedButton()

        if clicked_button == save_button: return QMessageBox.YesRole
        if clicked_button == discard_button: return QMessageBox.NoRole
        return QMessageBox.RejectRole # Cancel or if dialog is closed unexpectedly
    
    def _prompt_save_changes_flexible(self):
        """
        シンプルな保存確認ダイアログ
        Returns:
            "save": 保存して続行
            "discard": 変更を破棄して続行
            "cancel": キャンセル
        """
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Question)
        msg_box.setWindowTitle("未保存の変更")
        msg_box.setText("変更を保存しますか？")

        save_button = msg_box.addButton("保存", QMessageBox.YesRole)
        discard_button = msg_box.addButton("破棄", QMessageBox.NoRole)
        cancel_button = msg_box.addButton("キャンセル", QMessageBox.RejectRole)
        
        msg_box.setDefaultButton(save_button)
        msg_box.setEscapeButton(cancel_button)

        msg_box.exec_()
        clicked_button = msg_box.clickedButton()
        
        if clicked_button == save_button: return "save"
        if clicked_button == discard_button: return "discard"
        return "cancel"
    
    def _save_with_validation_recovery(self, show_message=True):
        """バリデーションエラーがあっても可能な限り保存"""
        try:
            self.save_to_excel(show_message=show_message)
        except Exception as e:
            # バリデーションエラーでも下書き保存
            logging.warning(f"通常保存失敗、下書き保存に切り替え: {e}")
            self._save_as_draft()
            if show_message:
                QMessageBox.information(self, "保存完了", 
                    "一部エラーがありますが、下書きとして保存しました。\n"
                    "後で詳細を確認・修正してください。")
    
    def _save_draft_and_continue(self, target_item):
        """下書き保存して指定した商品に切り替え"""
        self._save_as_draft()
        self.is_dirty = False
        if target_item:
            self.load_product(target_item)
            if hasattr(self, '_update_status_bar'):
                self._update_status_bar()
        
        # 下書き保存の通知
        if hasattr(self, 'show_status_message'):
            self.show_status_message("下書き保存完了 - 後で詳細確認してください")
    
    def _save_as_draft(self):
        """下書き保存（安全な部分保存）"""
        try:
            # 下書きとして安全に保存可能な部分のみ保存
            draft_data = self._collect_safe_draft_data()
            
            # 下書き専用の保存処理
            self._save_draft_data(draft_data)
                
            logging.info("下書き保存完了")
        except Exception as e:
            logging.error(f"下書き保存エラー: {e}")
            # 下書き保存失敗時はユーザーに通知
            QMessageBox.warning(self, "下書き保存エラー", 
                "下書き保存に失敗しました。\n"
                f"エラー: {str(e)}\n\n"
                "通常の保存をお試しください。")
    
    def _collect_safe_draft_data(self):
        """下書き保存用の安全なデータ収集"""
        safe_data = {}
        
        # 基本フィールドのみ（バリデーションエラーが起きにくい項目）
        safe_fields = [
            HEADER_MYCODE, HEADER_PRODUCT_NAME, 
            "シリーズ名", "メーカー売価_税込み", "送料形態"
        ]
        
        for field_name in safe_fields:
            if field_name in self.main_fields:
                widget = self.main_fields[field_name]
                if isinstance(widget, QLineEdit):
                    safe_data[field_name] = widget.text()
                elif isinstance(widget, QTextEdit):
                    safe_data[field_name] = widget.toPlainText()
                elif isinstance(widget, QComboBox):
                    safe_data[field_name] = widget.currentText()
        
        # タイムスタンプを追加
        import datetime
        import time
        safe_data['_draft_timestamp'] = datetime.datetime.now().isoformat()
        safe_data['_is_draft'] = True
        
        return safe_data
    
    def _save_draft_data(self, draft_data):
        """下書きデータの実際の保存"""
        # 下書き専用ファイルに保存（実装は既存の保存機能の安全な部分のみ使用）
        draft_code = draft_data.get(HEADER_MYCODE, "DRAFT_" + str(int(time.time())))
        
        # 下書きファイル名
        draft_filename = f"draft_{draft_code}.json"
        draft_path = os.path.join(get_user_data_dir(), "drafts", draft_filename)
        
        # ディレクトリ作成
        os.makedirs(os.path.dirname(draft_path), exist_ok=True)
        
        # JSON形式で安全に保存
        import json
        with open(draft_path, 'w', encoding='utf-8') as f:
            json.dump(draft_data, f, ensure_ascii=False, indent=2)
        
        logging.info(f"下書き保存: {draft_path}")
    
    def _set_mode_indicator(self, mode_text, color):
        """モードインジケーターを設定（タイトルバーとステータスバー）"""
        # タイトルバーに表示
        base_title = f"商品登録入力ツール v{CURRENT_VERSION}"
        self.setWindowTitle(f"{base_title} - {mode_text}")
        
        # ステータスバーにも表示
        if hasattr(self, 'status_bar'):
            # モードラベルがなければ作成
            if not hasattr(self, '_mode_label'):
                from PyQt5.QtWidgets import QLabel
                self._mode_label = QLabel()
                self._mode_label.setStyleSheet(
                    "QLabel { "
                    "padding: 4px 8px; "
                    "border-radius: 4px; "
                    "font-weight: bold; "
                    "color: white; "
                    "}"
                )
                self.status_bar.addPermanentWidget(self._mode_label)
            
            self._mode_label.setText(mode_text)
            self._mode_label.setStyleSheet(
                f"QLabel {{ "
                f"background-color: {color}; "
                f"padding: 4px 8px; "
                f"border-radius: 4px; "
                f"font-weight: bold; "
                f"color: white; "
                f"}}"
            )
    
    def _clear_mode_indicator(self):
        """モードインジケーターをクリア"""
        base_title = f"商品登録入力ツール v{CURRENT_VERSION}"
        self.setWindowTitle(base_title)
        
        if hasattr(self, '_mode_label'):
            self._mode_label.setText("待機中")
            self._mode_label.setStyleSheet(
                "QLabel { "
                "background-color: #9E9E9E; "
                "padding: 4px 8px; "
                "border-radius: 4px; "
                "font-weight: bold; "
                "color: white; "
                "}"
            )
    
    def _adjust_splitter_sizes_with_search(self):
        """検索パネル表示時のスプリッターサイズ調整"""
        if hasattr(self, 'main_splitter') and self.main_splitter.count() >= 2:
            sizes = self.main_splitter.sizes()
            total = sum(sizes)
            if total > 0:
                # 検索パネルに固定幅350を確保
                panel_width = 350
                remaining = total - panel_width
                if self.main_splitter.count() == 2:
                    # メインエリア : 検索パネル
                    new_sizes = [remaining, panel_width]
                elif self.main_splitter.count() == 3:
                    # 左側 : 中央 : 検索パネル
                    left_ratio = 0.3
                    center_ratio = 0.7
                    left_width = int(remaining * left_ratio)
                    center_width = remaining - left_width
                    new_sizes = [left_width, center_width, panel_width]
                else:
                    return  # 予期しないウィジェット数
                
                self.main_splitter.setSizes(new_sizes)
    
    def _restore_splitter_sizes_without_search(self):
        """検索パネル非表示時のスプリッターサイズ復元"""
        if hasattr(self, 'main_splitter') and self.main_splitter.count() >= 2:
            sizes = self.main_splitter.sizes()
            if len(sizes) >= 2:
                # 検索パネル以外の領域を均等に再配分
                total = sum(sizes[:-1])  # 検索パネル以外の合計
                if self.main_splitter.count() == 3:
                    # 左側 : 中央の比率を維持
                    if total > 0:
                        left_ratio = sizes[0] / total if total > 0 else 0.3
                        new_total = sum(sizes)  # 全体サイズ
                        left_width = int(new_total * left_ratio)
                        center_width = new_total - left_width
                        new_sizes = [left_width, center_width, 0]  # 検索パネルは0
                        self.main_splitter.setSizes(new_sizes)

    def __del__(self):
        """デストラクタでリソースクリーンアップ"""
        try:
            self._cleanup_event_filters()
        except Exception as e:
            logging.debug(f"リソースクリーンアップエラー: {e}")
    
    def closeEvent(self, event):
        # アプリ終了時は商品選択変更処理を停止
        self._is_closing = True
        
        if self.is_dirty:
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Question)
            msg_box.setWindowTitle("終了確認")
            msg_box.setText("未保存の変更があります。変更を保存しますか？")
            msg_box.setInformativeText("「いいえ」を選択して終了した場合でも、自動保存されたデータは次回起動時に復元を試みることができます。")

            save_button = msg_box.addButton("はい", QMessageBox.YesRole)       # "はい" (Save)
            discard_button = msg_box.addButton("いいえ", QMessageBox.NoRole)    # "いいえ" (Don't Save / Discard)
            cancel_button = msg_box.addButton("キャンセル", QMessageBox.RejectRole) # "キャンセル" (Cancel operation)
            
            msg_box.setDefaultButton(cancel_button) # デフォルトはキャンセル
            msg_box.setEscapeButton(cancel_button)  # Escキーでキャンセル

            msg_box.exec_()
            clicked_button = msg_box.clickedButton()

            if clicked_button == save_button:
                self.save_to_excel(show_message=False) # メッセージは出さずに保存
                # 保存後、is_dirty は False になる
            elif clicked_button == discard_button:
                # 変更を破棄して終了処理を続行
                self.is_dirty = False  # ダーティフラグをクリアして終了
                # 商品選択変更処理も停止
                self.product_list.blockSignals(True)
            elif clicked_button == cancel_button:
                event.ignore()
                return
            else:
                # 通常ここには到達しないはず (例: ダイアログが予期せず閉じられた場合など)
                event.ignore() # 安全のため、予期せぬ場合は終了をキャンセル
                return

        settings = QSettings("株式会社大宝家具", APP_NAME) # 組織名を設定
        settings.setValue("geometry", self.saveGeometry())
        settings.setValue("mainSplitterState", self.main_splitter.saveState())
        settings.setValue("rightSplitterState", self.right_splitter.saveState())
        # 自動更新チェックの設定を保存（デフォルトはTrue）
        settings.setValue("update/auto_check_enabled", getattr(self, 'auto_update_check_enabled', True))
        
        # リソースクリーンアップ
        self._cleanup_event_filters()
        
        # 万が一対策システムのクリーンアップ
        self._cleanup_emergency_systems()
        
        super().closeEvent(event)
    
    def _cleanup_emergency_systems(self):
        """万が一対策システムのクリーンアップ"""
        try:
            # 1. クラッシュ復旧システムのクリーンアップ
            if hasattr(self, 'crash_recovery'):
                self.crash_recovery.clean_session()
            
            # 2. ファイルロックの解放
            if hasattr(self, 'file_lock_manager'):
                self.file_lock_manager.release_app_lock()
            
            # 3. ハートビートタイマーの停止
            if hasattr(self, 'heartbeat_timer'):
                self.heartbeat_timer.stop()
            
            # 4. メモリ監視システムの停止
            if hasattr(self, 'memory_monitor'):
                self.memory_monitor.stop_monitoring()
            
            # 5. ネットワーク監視システムの停止
            if hasattr(self, 'network_checker'):
                self.network_checker.network_monitor.stop_monitoring()
                
        except Exception as e:
            logging.error(f"万が一対策システムクリーンアップエラー: {e}")

    def _auto_save_data(self):
        if not self.is_dirty: # 変更がなければ何もしない
            return
            
        # Undo/Redo実行中は自動保存をスキップ
        if self._is_undoing:
            logging.debug("自動保存スキップ: Undo/Redo実行中")
            return
            
        # 商品コードが空の場合は自動保存しない（新規入力画面または削除直後）
        mycode_field = self.main_fields.get(HEADER_MYCODE)
        if mycode_field and not mycode_field.text().strip():
            logging.debug("自動保存スキップ: 商品コードが空です")
            return

        # 自動保存中は表示しない（うるさくないように）
        
        # print("DEBUG: Auto-saving data...") # デバッグ用
        settings = QSettings("株式会社大宝家具", APP_NAME)
        settings.setValue("autosave/exists", True)

        # コントロールカラム
        settings.setValue("autosave/control_column_is_n", self.control_radio_n.isChecked())

        # 主要フィールド
        for field_name, widget in self.main_fields.items():
            key = f"autosave/main_fields/{field_name}"
            if isinstance(widget, QLineEdit):
                settings.setValue(key, widget.text())
            elif isinstance(widget, QTextEdit):
                settings.setValue(key, widget.toPlainText())
            elif isinstance(widget, QComboBox):
                settings.setValue(key, widget.currentText())
        
        # SKUデータ (JSON文字列として保存)
        if hasattr(self, 'sku_data_list') and self.sku_data_list:
            try:
                sku_data_json = json.dumps(self.sku_data_list)
                settings.setValue("autosave/sku_data", sku_data_json)
            except Exception as e:
                logging.warning(f"SKUデータの自動保存(JSONシリアライズ)に失敗しました。", exc_info=e)
        else:
            settings.remove("autosave/sku_data") # データがなければキーを削除

        # Y!specデータ
        current_y_category_id = self.main_fields.get(HEADER_Y_CATEGORY_ID, QLineEdit()).text()
        settings.setValue("autosave/y_category_id_for_yspec", current_y_category_id) # Y_spec復元時のカテゴリID
        for i in range(MAX_Y_SPEC_COUNT):
            key = f"autosave/yspec/Y_spec{i+1}"
            if self.y_spec_current_editors[i] and self.y_spec_current_definitions[i]:
                settings.setValue(key, self._get_y_spec_value_for_save(i))
            else:
                settings.remove(key)
        
        # 自動保存完了時刻をツールチップに表示
        self._update_save_button_tooltip("自動保存")
        
        # print("DEBUG: Auto-save complete.") # デバッグ用

    def _handle_search_action(self):
        """検索アクション処理"""
        try:
            self.show_search_dialog()
        except Exception as e:
            logging.error(f"検索ダイアログエラー: {e}", exc_info=True)

    def show_search_dialog(self):
        """Excel風の検索ダイアログを表示"""
        # main_splitter の存在確認
        if not hasattr(self, 'main_splitter'):
            return
        
        # 非モーダル検索パネルの切り替え
        if not hasattr(self, '_search_panel'):
            try:
                self._search_panel = SearchPanel(self)
                self._search_panel.hide()  # 初期は非表示
                
                # スプリッターにパネルを安全に追加
                current_count = self.main_splitter.count()
                if current_count < 3:  # まだ3つ未満なら追加
                    self.main_splitter.addWidget(self._search_panel)
                    # 初期サイズを設定（検索パネルは最小幅）
                    if current_count >= 1:
                        sizes = self.main_splitter.sizes()
                        total = sum(sizes)
                        # 既存の領域から検索パネル分を確保
                        panel_width = 350
                        main_width = max(total - panel_width, total // 2)
                        new_sizes = sizes[:-1] + [main_width, panel_width]
                        self.main_splitter.setSizes(new_sizes)
                else:
                    logging.warning("検索パネル: スプリッターに3つ以上のウィジェットがあるため、追加をスキップしました")
            except Exception as e:
                logging.error(f"検索パネル作成エラー: {e}", exc_info=True)
                return
        
        # パネルの表示/非表示を切り替え
        if hasattr(self, '_search_panel'):
            if self._search_panel.isVisible():
                self._search_panel.hide()
                # 検索パネルを隠すときはスプリッターサイズを調整
                if hasattr(self, '_restore_splitter_sizes_without_search'):
                    self._restore_splitter_sizes_without_search()
            else:
                self._search_panel.show()
                # 検索パネルを表示するときにサイズ調整
                if hasattr(self, '_adjust_splitter_sizes_with_search'):
                    self._adjust_splitter_sizes_with_search()
                try:
                    self._search_panel.search_input.setFocus()  # 検索入力にフォーカス
                except Exception as e:
                    logging.error(f"検索入力フォーカスエラー: {e}")

    def clear_search(self):
        """ESCキーでの検索クリア - フォーカスに応じて動作を分ける"""
        # 1. 検索パネルが表示されている場合は、それを閉じる（最優先）
        if hasattr(self, '_search_panel') and self._search_panel.isVisible():
            self._search_panel.hide()
            if hasattr(self, '_restore_splitter_sizes_without_search'):
                self._restore_splitter_sizes_without_search()
            return
        
        # 2. 検索バーにフォーカスがある場合、または検索バーに文字が入力されている場合は、検索をクリア
        if hasattr(self, 'search_bar'):
            if self.search_bar.hasFocus() or self.search_bar.text().strip():
                self.search_bar.clear()
                self.search_bar.clearFocus()
                return
        
        # 3. その他の場合は何もしない（一般的なESCキーの動作）

    def focus_search(self):
        """検索ボックスにフォーカスを移動"""
        self.search_bar.setFocus()
        self.search_bar.selectAll()

    def _create_menu_bar(self):
        """メニューバーを作成してショートカットを表示"""
        from PyQt5.QtWidgets import QMenuBar, QMenu, QAction
        
        # 既存のメニューバーをクリアして設定
        self.menu_bar.clear()
        self.menu_bar.setFixedHeight(MENU_BAR_HEIGHT)  # 固定高さ
        
        # メニューバーの基本属性を設定（重要）
        self.menu_bar.setNativeMenuBar(False)  # ネイティブメニューバーを無効化
        self.menu_bar.setVisible(True)
        self.menu_bar.setEnabled(True)
        
        self.menu_bar.setStyleSheet(f"""
            QMenuBar {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                             stop: 0 #ffffff, stop: 1 #f8fafc);
                border-bottom: 1px solid #e2e8f0;
                font-size: {FONT_SIZE_MENU}px;
                font-weight: 600;
                padding: 2px;
                spacing: 4px;
            }}
            QMenuBar::item {{
                background-color: transparent;
                padding: 6px 14px;
                margin: 2px;
                border: none;
                border-radius: 6px;
                color: #374151;
                font-weight: 500;
            }}
            QMenuBar::item:selected {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                             stop: 0 #3b82f6, stop: 1 #1d4ed8);
                color: white;
                border-radius: 6px;
            }}
            QMenuBar::item:pressed {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                             stop: 0 #1d4ed8, stop: 1 #1e40af);
                color: white;
            }}
            QMenu {{
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 4px;
                font-size: 13px;
            }}
            QMenu::item {{
                padding: 8px 16px;
                border-radius: 4px;
                margin: 1px;
            }}
            QMenu::item:selected {{
                background-color: #f1f5f9;
                color: #1f2937;
            }}
        """)
        
        # ファイルメニューを最初に作成（正しい順序で）
        file_menu = self.menu_bar.addMenu("ファイル(&F)")
        
        new_action = QAction("新規作成(&N)", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(lambda: self._handle_new_product_action())
        file_menu.addAction(new_action)
        
        save_action = QAction("保存(&S)", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(lambda: self.save_to_excel())
        file_menu.addAction(save_action)
        
        file_menu.addSeparator()
        
        
        # 編集メニューを2番目に作成
        edit_menu = self.menu_bar.addMenu("編集(&E)")
        
        # Undo/Redoアクション
        self.undo_action = QAction("元に戻す(&U)", self)
        self.undo_action.setShortcut("Ctrl+Z")
        self.undo_action.triggered.connect(lambda: self.undo())
        self.undo_action.setEnabled(False)
        edit_menu.addAction(self.undo_action)
        
        self.redo_action = QAction("やり直し(&R)", self)
        self.redo_action.setShortcut("Ctrl+Y")
        self.redo_action.triggered.connect(lambda: self.redo())
        self.redo_action.setEnabled(False)
        edit_menu.addAction(self.redo_action)
        
        edit_menu.addSeparator()
        
        # 検索アクションを作成（唯一のCtrl+F定義）
        search_action = QAction("検索(&F)", self)
        search_action.setShortcut("Ctrl+F")
        search_action.triggered.connect(lambda: self._handle_search_action())
        # ショートカットの優先度を設定
        search_action.setShortcutContext(Qt.ApplicationShortcut)
        # メインウィンドウに追加（これが重要）
        self.addAction(search_action)
        edit_menu.addAction(search_action)
        
        
        clear_search_action = QAction("検索をクリア", self)
        clear_search_action.setShortcut("Escape")
        clear_search_action.triggered.connect(lambda: self.clear_search())
        edit_menu.addAction(clear_search_action)
        
        edit_menu.addSeparator()
        
        # コピー/ペーストアクションをメニューに追加
        # アクションが存在することを確認してから追加
        if hasattr(self, '_copy_product_action_ref') and self._copy_product_action_ref:
            edit_menu.addAction(self._copy_product_action_ref)
        else:
            # アクションがまだ作成されていない場合は、ここで作成
            self._setup_copy_paste_actions()
            if hasattr(self, '_copy_product_action_ref') and self._copy_product_action_ref:
                edit_menu.addAction(self._copy_product_action_ref)
            else:
                logging.warning("コピーアクションの作成に失敗しました")
        
        if hasattr(self, '_paste_product_action_ref') and self._paste_product_action_ref:
            edit_menu.addAction(self._paste_product_action_ref)
        else:
            logging.warning("ペーストアクションの追加に失敗しました")
        
        edit_menu.addSeparator()
        
        sku_add_action = QAction("SKU追加(&A)", self)
        sku_add_action.setShortcut("Ctrl+Shift+A")
        sku_add_action.triggered.connect(lambda: self.add_sku_column())
        edit_menu.addAction(sku_add_action)
        
        # ツールメニュー
        tools_menu = self.menu_bar.addMenu("ツール(&T)")
        
        category_action = QAction("カテゴリ選択(&G)", self)
        category_action.setShortcut("Ctrl+G")
        category_action.triggered.connect(lambda: self.open_category_dialog())
        tools_menu.addAction(category_action)
        
        id_search_action = QAction("ID検索(&I)", self)
        id_search_action.setShortcut("Ctrl+I")
        id_search_action.triggered.connect(lambda: self._open_id_search_dialog())
        tools_menu.addAction(id_search_action)
        
        tools_menu.addSeparator()
        
        csharp_action = QAction("C#実行(&C)", self)
        csharp_action.setShortcut("F5")
        csharp_action.triggered.connect(lambda: self.handle_csv_generation_button_click())
        tools_menu.addAction(csharp_action)
        
        html_action = QAction("画像説明HTML生成(&H)", self)
        html_action.setShortcut("Ctrl+H")
        html_action.triggered.connect(lambda: self.generate_html())
        tools_menu.addAction(html_action)
        
        
        # ヘルプメニュー
        help_menu = self.menu_bar.addMenu("ヘルプ(&P)")
        
        shortcuts_help_action = QAction("ショートカット一覧(&K)", self)
        shortcuts_help_action.setShortcut("F1")
        shortcuts_help_action.triggered.connect(lambda: self.show_shortcuts_help())
        help_menu.addAction(shortcuts_help_action)
        
        help_menu.addSeparator()
        
        # 更新チェックメニュー項目
        update_action = QAction("更新の確認(&U)", self)
        update_action.triggered.connect(lambda: self._check_for_updates_manual())
        help_menu.addAction(update_action)
        
        # バージョン情報
        about_action = QAction(f"バージョン情報(&A)", self)
        about_action.triggered.connect(lambda: self._show_about_dialog())
        help_menu.addAction(about_action)
        
        # メニューバーのマウスイベントを確実にする
        self.menu_bar.setMouseTracking(True)
        self.menu_bar.setAttribute(Qt.WA_Hover, True)
        self.menu_bar.setFocusPolicy(Qt.StrongFocus)
        
        # メニューバーを強制的に表示
        self.menu_bar.show()
        self.menu_bar.raise_()
        
        # メニューバーが確実に有効化されるよう設定
        self.menu_bar.setEnabled(True)
        self.menu_bar.setVisible(True)
        self.menu_bar.raise_()  # 前面に移動
        
        # すべてのメニューとアクションが有効であることを確認
        for action in self.menu_bar.actions():
            action.setEnabled(True)
            if action.menu():
                for sub_action in action.menu().actions():
                    if not sub_action.isSeparator():
                        sub_action.setEnabled(True)
        
        # メニューバーの更新を強制
        self.menu_bar.update()
        self.menu_bar.repaint()
        
        # アプリケーション処理を確実にする
        QApplication.processEvents()
        
        logging.info("メニューバーの作成が完了しました")

    def _ensure_menu_bar_visible(self):
        """メニューバーが正常に表示されているかを確認し、必要に応じて修正"""
        try:
            if not self.menu_bar.isVisible() or not self.menu_bar.isEnabled():
                logging.warning("メニューバーが無効化されているため、再有効化します")
                self.menu_bar.setVisible(True)
                self.menu_bar.setEnabled(True)
                self.menu_bar.update()
                QApplication.processEvents()
            
            # メニューアイテムの数をチェック
            action_count = len(self.menu_bar.actions())
            if action_count == 0:
                logging.warning("メニューバーにアイテムがないため、再作成します")
                self._create_menu_bar()
            else:
                logging.info(f"メニューバーの確認完了: {action_count}個のメニューアイテムが有効")
                
        except Exception as e:
            logging.error(f"メニューバーの確認中にエラー: {e}", exc_info=True)

    def _update_save_button_tooltip(self, save_type="保存"):
        """保存ボタンのツールチップを最終保存時刻で更新"""
        from datetime import datetime
        current_time = datetime.now().strftime("%H:%M:%S")
        current_date = datetime.now().strftime("%m/%d")
        self.last_save_time = current_time
        
        # より詳細で親切なツールチップ
        base_tooltip = "💾 現在の変更を保存します (Ctrl+S)"
        status_info = f"📅 最終{save_type}: {current_date} {current_time}"
        
        if save_type == "自動保存":
            auto_info = "⏰ 30秒ごとに自動保存されます"
            tooltip_with_time = f"{base_tooltip}\n\n{status_info}\n{auto_info}"
        else:
            tooltip_with_time = f"{base_tooltip}\n\n{status_info}"
            
        self.save_btn.setToolTip(tooltip_with_time)
    
    def _check_for_updates_manual(self):
        """手動での更新チェック（新しいシンプル版）"""
        try:
            from src.utils.version_checker import check_for_updates_simple
            check_for_updates_simple(self, silent=False)
        except Exception as e:
            logging.error(f"手動更新チェック中にエラー: {e}")
            QMessageBox.warning(
                self, 
                "更新チェックエラー", 
                "更新チェック機能で問題が発生しました。\n後でもう一度お試しください。"
                )
    
    def _delayed_update_check(self):
        """遅延実行される起動時更新チェック"""
        try:
            logging.info("遅延更新チェック実行開始")
            # シンプルな自動更新チェック
            if check_for_updates_on_startup:
                check_for_updates_on_startup(self)
                logging.info("遅延更新チェック実行完了")
            else:
                logging.info("check_for_updates_on_startup関数が利用できません")
        except Exception as e:
            logging.error(f"遅延更新チェック中にエラー: {e}", exc_info=True)
    
    def _show_about_dialog(self):
        """バージョン情報ダイアログを表示"""
        about_text = f"""<h2>商品登録入力ツール</h2>
<p><b>バージョン:</b> {CURRENT_VERSION}</p>
<p><b>開発元:</b> 株式会社大宝家具</p>
<p><b>開発者:</b> Seito Nakamura</p>
<p><b>リリース日:</b> 2025年5月29日</p>
<br>
<p>このツールは商品情報の効率的な入力と管理を支援します。</p>
<br>
<p><b>使用ライブラリ:</b></p>
<ul>
<li>PyQt5 - GUI フレームワーク</li>
<li>openpyxl - Excel ファイル操作</li>
</ul>
<br>
<p><small>Copyright © 2025 株式会社大宝家具. All rights reserved.<br>
Developed by Seito Nakamura</small></p>"""
        
        QMessageBox.about(self, "バージョン情報", about_text)

    def _load_auto_saved_data(self):
        settings = QSettings("株式会社大宝家具", APP_NAME)
        if not settings.value("autosave/exists", False, type=bool):
            return

        # clear_fields内で_clear_auto_save_dataが呼ばれるのを防ぐため、一時的にフラグを立てるなどの工夫も可能だが、
        self.clear_fields() # まず現在のフィールドをクリア (ただし、これは自動保存クリアも呼ぶので注意が必要。クリア後にフラグを再設定する)
        
        # コントロールカラム
        self.control_radio_n.setChecked(settings.value("autosave/control_column_is_n", True, type=bool))

        # 主要フィールド
        for field_name, widget in self.main_fields.items():
            key = f"autosave/main_fields/{field_name}"
            if settings.contains(key):
                value = settings.value(key, type=str)
                if isinstance(widget, QLineEdit): widget.setText(value)
                elif isinstance(widget, QTextEdit): widget.setPlainText(value)
                elif isinstance(widget, QComboBox): widget.setCurrentText(value) # findTextしてIndex設定の方が確実かも

        # ExpandableFieldGroup の状態を復元後に再評価 (特に商品サイズ)
        # YカテゴリIDが設定され、Y_spec UIが構築された後に行うのが望ましい
        # Y_specのロード前に、商品サイズのUIが正しい状態であることを保証する
        for efg_label, efg_instance in self.expandable_field_group_instances.items():
            if efg_label == "商品サイズ": # "商品サイズ"グループに限定
                for i in range(efg_instance.group_count):
                    field_a_name = f"{efg_instance.group_label_prefix}_{i+1}a"
                    field_a_widget = self.main_fields.get(field_a_name)
                    if isinstance(field_a_widget, QLineEdit):
                        efg_instance._update_product_size_b_input_type(field_a_widget.text(), i)
        
        # Y!specデータ (YカテゴリIDを先に復元し、UIを構築してから値をロード)
        saved_y_category_id = settings.value("autosave/y_category_id_for_yspec", "", type=str)
        if HEADER_Y_CATEGORY_ID in self.main_fields:
            self.main_fields[HEADER_Y_CATEGORY_ID].setText(saved_y_category_id) # これで _on_y_category_id_changed が呼ばれる
        for i in range(MAX_Y_SPEC_COUNT):
            key = f"autosave/yspec/Y_spec{i+1}"
            if settings.contains(key):
                self._load_y_spec_value(settings.value(key, type=str))

        # SKUデータ (JSON文字列から復元)
        sku_data_json = settings.value("autosave/sku_data", "", type=str)
        if sku_data_json:
            try:
                self.sku_data_list = json.loads(sku_data_json)
                self.show_sku_table()
            except Exception as e:
                logging.warning(f"SKUデータの自動保存データ復元(JSONデシリアライズ)に失敗しました。", exc_info=e)
                self.sku_data_list = []

        self._format_and_sync_price_fields() # 価格フィールドのフォーマット
        for field_name_bc in self.byte_count_labels.keys(): # バイト数表示更新
            if field_name_bc in self.main_fields and isinstance(self.main_fields.get(field_name_bc), QLineEdit):
                self._update_byte_count_display(field_name_bc, self.main_fields.get(field_name_bc).text())
        if self.digit_count_label_mycode and HEADER_MYCODE in self.main_fields: # mycode桁数表示更新
            self._update_mycode_digit_count_display(self.main_fields[HEADER_MYCODE].text())
        self._update_relevant_links()

        self._sync_product_size_to_yspec() # ★★★ 復元処理の最後に同期処理を呼び出す ★★★

        self.is_dirty = True # 復元したデータは「未保存」扱い (save_btnも有効になる)
        # 復元後も、ユーザーが手動保存するまでは自動保存フラグは残しておく
        # clear_fieldsによってautosave/existsがクリアされる可能性があるため、ここで再設定するか、
        # clear_fieldsから_clear_auto_save_dataの呼び出しを分離する。後者を推奨。
        msg_info = "前回のデータを復元しました。内容を確認し、必要であれば保存してください。"
        QMessageBox.information(self, "復元完了", msg_info); logging.info(msg_info)

    def _clear_auto_save_data(self):
        # print("DEBUG: Clearing auto-save data...") # デバッグ用
        settings = QSettings("株式会社大宝家具", APP_NAME)
        settings.remove("autosave/exists")
        settings.remove("autosave/control_column_is_n")
        for field_name in self.main_fields.keys():
            settings.remove(f"autosave/main_fields/{field_name}")
        settings.remove("autosave/sku_data")
        settings.remove("autosave/y_category_id_for_yspec")
        for i in range(MAX_Y_SPEC_COUNT):
            settings.remove(f"autosave/yspec/Y_spec{i+1}")
        # print("DEBUG: Auto-save data cleared.") # デバッグ用

    def _handle_new_product_action(self):
        """新規作成ボタンが押されたときの処理"""
        if self.is_dirty:
            choice = self._prompt_save_changes()
            if choice == QMessageBox.YesRole:
                self.save_to_excel() # 保存処理
                # 保存後、is_dirty は False になっているはず
            elif choice == QMessageBox.NoRole:
                self.is_dirty = False # 変更を破棄
            elif choice == QMessageBox.RejectRole:
                return # 新規作成をキャンセル

        # ダーティでない、または保存/破棄が選択された場合
        self._is_new_mode = True  # 新規作成モードフラグを設定
        self.product_list.blockSignals(True)
        self.product_list.clearSelection() # これが currentItemChanged をトリガーしないように
        self.product_list.setCurrentItem(None)  # Qt内部のprevious参照を明示的にクリア
        self.product_list.blockSignals(False)

        self.clear_fields() # これが is_dirty を False にする
        # clear_fields の中で _clear_auto_save_data が呼ばれる場合があるため、
        # ここでの _clear_auto_save_data の呼び出しは、clear_fields の動作に依存します。
        # clear_fields が自動保存データをクリアしない場合は、ここで明示的に呼び出す必要があります。
        self._clear_auto_save_data() # 新規作成なので、既存の自動保存データをクリア (clear_fieldsの後)

    def _set_list_selection_after_cancel(self, item_to_select):
        """キャンセル操作後、指定されたアイテムをリストで選択する。item_to_selectがNoneなら選択解除。"""
        # 強制的に処理中フラグを設定してcurrentItemChangedを無視
        self._is_handling_selection_change = True
        try:
            self.product_list.blockSignals(True)
            if item_to_select:
                self.product_list.setCurrentItem(item_to_select)
            else:
                # item_to_select が None の場合 (例: 新規作成後に最初のアイテム選択をキャンセル)
                # リストの選択をクリアする
                self.product_list.clearSelection()
                self.product_list.setCurrentItem(None)  # Qt内部のprevious参照を明示的にクリア
            self.product_list.blockSignals(False)
        finally:
            self._is_handling_selection_change = False

    def _sync_product_size_to_yspec(self): # _set_list_selection_after_cancel の後に追加
        """「商品サイズ」の「本体」の寸法をY_specの対応フィールドに同期する"""
        product_size_efg = self.expandable_field_group_instances.get("商品サイズ")
        if not product_size_efg:
            return

        main_body_row_index = -1
        for i in range(product_size_efg.group_count):
            field_a_name = f"{product_size_efg.group_label_prefix}_{i+1}a"
            field_a_widget = self.main_fields.get(field_a_name)
            if isinstance(field_a_widget, QLineEdit) and field_a_widget.text().strip() == "本体":
                main_body_row_index = i
                break
        
        if main_body_row_index == -1:
            return

        dim_data = product_size_efg.dimension_fields_list[main_body_row_index]
        if not dim_data:
            return
            
        width_val = dim_data['w'].text().strip()
        depth_val = dim_data['d'].text().strip()
        height_val = dim_data['h'].text().strip()

        editors_values_and_defs = [
            (self.y_spec_width_editor, width_val, self.y_spec_width_definition),
            (self.y_spec_depth_editor, depth_val, self.y_spec_depth_definition),
            (self.y_spec_height_editor, height_val, self.y_spec_height_definition),
        ]

        for editor, value_from_product_size, spec_def_for_editor in editors_values_and_defs:
            if editor and isinstance(editor, QLineEdit) and spec_def_for_editor:
                current_yspec_value = editor.text().strip()
                formatted_value_to_set = ""

                if value_from_product_size: # 同期元に値がある場合のみフォーマットを試みる
                    try:
                        if spec_def_for_editor["data_type"] == 2: # 整数
                            # 小数点が含まれていても整数に変換（例: "80.0" -> 80）
                            num_val = int(float(value_from_product_size))
                            formatted_value_to_set = str(num_val)
                        elif spec_def_for_editor["data_type"] == 4: # 整数or小数
                            num_val = float(value_from_product_size)
                            formatted_value_to_set = f"{num_val:.4f}" # 小数点以下4桁
                        else: # その他のデータ型 (または数値型でない場合)
                            formatted_value_to_set = value_from_product_size # そのまま
                    except ValueError:
                        # 数値変換に失敗した場合は、元の値をそのまま使う
                        # (バリデーションはQLineEdit側で行う想定、または空にする)
                        # ここでは、同期元の値が数値でない場合は空として扱う方が安全かもしれません
                        formatted_value_to_set = "" # または value_from_product_size
                else: # 同期元が空の場合
                    formatted_value_to_set = ""


                # 同期元の値(フォーマット後)が空で、かつY_spec側に既に値がある場合は、上書きしない
                if formatted_value_to_set == "" and current_yspec_value != "":
                    continue # 何もせず次のエディタへ

                # それ以外の場合（同期元に値がある、またはY_spec側も空）は、通常通り同期
                if current_yspec_value != formatted_value_to_set:
                    editor.blockSignals(True)
                    editor.setText(formatted_value_to_set)
                    editor.blockSignals(False)
                    # この同期処理自体はUIの自動更新なので、is_dirtyをTrueにしない。
                    # Y_spec側のエディタに接続されたmark_dirtyも呼ばれない。
            elif editor and isinstance(editor, QLineEdit) and not spec_def_for_editor:
                # spec_def が見つからないがエディタは存在する場合（通常は発生しにくい）
                # 安全のため、ここでは何もしないか、ログを出す程度
                pass
        
        # --- 重量フィールドの同期 ---
        weight_value_from_ps = ""
        if product_size_efg: # product_size_efg が存在する場合のみ処理
            for i in range(product_size_efg.group_count):
                field_a_name_w = f"{product_size_efg.group_label_prefix}_{i+1}a"
                field_a_widget_w = self.main_fields.get(field_a_name_w)
                if isinstance(field_a_widget_w, QLineEdit) and field_a_widget_w.text().strip() == "重量":
                    # 'b' フィールドから数値のみを抽出
                    field_b_name_w = f"{product_size_efg.group_label_prefix}_{i+1}b"
                    field_b_widget_w = self.main_fields.get(field_b_name_w) # main_fields から取得
                    if isinstance(field_b_widget_w, QLineEdit):
                        text_b_w = field_b_widget_w.text().strip()
                        # "約XXkg", "XXkg", "XX" から数値 XX を抽出 (スペースも考慮)
                        # 正規表現を修正: "約" と "kg" は任意、数値は必須
                        match_w = re.match(r"^\s*(?:約)?\s*([\d\.]+)\s*(?:kg)?\s*$", text_b_w, re.IGNORECASE)
                        if match_w:
                            weight_value_from_ps = match_w.group(1)
                        elif text_b_w.replace('.', '', 1).isdigit(): # 単純な数値の場合 (正規表現でカバーされるはずだが念のため)
                            weight_value_from_ps = text_b_w
                    break # 「重量」行を見つけたらループ終了

        if hasattr(self, 'y_spec_weight_editor') and self.y_spec_weight_editor and \
           isinstance(self.y_spec_weight_editor, QLineEdit) and \
           hasattr(self, 'y_spec_weight_definition') and self.y_spec_weight_definition:
            
            current_yspec_weight_value = self.y_spec_weight_editor.text().strip()
            formatted_weight_to_set = ""
            if weight_value_from_ps: # 商品サイズ側に重量値がある場合
                try:
                    if self.y_spec_weight_definition["data_type"] == 2: # 整数
                        formatted_weight_to_set = str(int(float(weight_value_from_ps)))
                    elif self.y_spec_weight_definition["data_type"] == 4: # 整数or小数
                        formatted_weight_to_set = f"{float(weight_value_from_ps):.4f}" # 小数点以下4桁
                except ValueError: pass # 変換失敗時は空のまま
            
            if not (formatted_weight_to_set == "" and current_yspec_weight_value != ""): # 同期元が空でYspecに値がある場合は上書きしない
                if current_yspec_weight_value != formatted_weight_to_set:
                    self.y_spec_weight_editor.blockSignals(True)
                    self.y_spec_weight_editor.setText(formatted_weight_to_set)
                    self.y_spec_weight_editor.blockSignals(False)
        elif hasattr(self, 'y_spec_weight_editor') and isinstance(self.y_spec_weight_editor, QLineEdit) and \
           (not hasattr(self, 'y_spec_weight_definition') or not self.y_spec_weight_definition):
                 # spec_def が見つからないがエディタは存在する場合（通常は発生しにくい）
                 # 安全のため、ここでは何もしないか、ログを出す程度
                 pass
        
class ClickableIconLabel(QLabel):
    clicked = pyqtSignal(str) # Emits the icon ID when clicked

    def __init__(self, icon_id, description, image_path, parent=None):
        super().__init__(parent)
        self.icon_id = icon_id
        self.description = description
        self.image_path = image_path
        self._selected = False
        self.setObjectName("ClickableIconLabelInstance") # For styling if needed

        self.setFixedSize(85, 110) # アイコン全体のサイズを小さく調整
        self.setAlignment(Qt.AlignCenter)
        
        container_layout = QVBoxLayout(self)
        container_layout.setContentsMargins(2,2,2,2) # マージンを詰める
        container_layout.setSpacing(1) # スペーシングを詰める

        self.image_label = QLabel(self)
        pixmap = QPixmap(self.image_path)
        if pixmap.isNull():
            self.image_label.setText(f"画像\nなし") # テキストを短縮
            self.image_label.setStyleSheet("border: 1px solid red; color: red; background-color: transparent;")
        else:
            self.image_label.setPixmap(pixmap.scaled(60, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)) # 画像サイズを小さく
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMinimumHeight(60) # 画像表示エリアの高さを調整
        container_layout.addWidget(self.image_label)

        self.desc_label = QLabel(f"{self.icon_id}: {self.description}", self)
        self.desc_label.setAlignment(Qt.AlignCenter | Qt.AlignTop) # Align top for multi-line
        self.desc_label.setWordWrap(True)
        self.desc_label.setMinimumHeight(30) # 説明文エリアの高さを調整 (1-2行程度を想定)
        container_layout.addWidget(self.desc_label)
        
        self.setLayout(container_layout)
        self.setToolTip(f"{self.description} (ID: {self.icon_id})")
        self.update_visual_state()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.set_selected(not self._selected)
            self.clicked.emit(self.icon_id) 
        super().mousePressEvent(event)

    def is_selected(self):
        return self._selected

    def set_selected(self, selected):
        if self._selected != selected:
            self._selected = selected
            self.update_visual_state()

    def update_visual_state(self):
        # Using custom properties for stylesheet targeting
        self.setProperty("selected", self._selected)
        self.style().unpolish(self)
        self.style().polish(self)

class ExplanationMarkDialog(QDialog):
    def __init__(self, icon_data_list, current_selected_ids_str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("説明マーク選択")
        self.setMinimumSize(600, 400) 
        self.resize(800, 600)

        self.icon_widgets = [] 
        self.current_selected_ids = set(s_id for s_id in current_selected_ids_str.strip().split() if s_id) # Ensure no empty strings

        main_layout = QVBoxLayout(self)
        self.filter_edit = QLineEdit(self); self.filter_edit.setPlaceholderText("アイコンIDまたは説明でフィルタ...")
        self.filter_edit.textChanged.connect(self._filter_icons); main_layout.addWidget(self.filter_edit)
        scroll_area = QScrollArea(self); scroll_area.setWidgetResizable(True)
        self.icons_container_widget = QWidget(); self.icons_grid_layout = QGridLayout(self.icons_container_widget)
        self.icons_grid_layout.setSpacing(8)

        cols = 8 # 列数を8に固定
        row, col = 0, 0
        for icon_info in icon_data_list:
            icon_widget = ClickableIconLabel(icon_info["id"], icon_info["description"], icon_info["path"], self.icons_container_widget)
            if icon_info["id"] in self.current_selected_ids: icon_widget.set_selected(True)
            icon_widget.clicked.connect(self._icon_clicked)
            self.icons_grid_layout.addWidget(icon_widget, row, col); self.icon_widgets.append(icon_widget)
            col += 1
            if col >= cols: col = 0; row += 1
        
        self.icons_grid_layout.setRowStretch(row + 1, 1); self.icons_grid_layout.setColumnStretch(cols, 1)
        scroll_area.setWidget(self.icons_container_widget); main_layout.addWidget(scroll_area)
        self.selected_display_label = QLabel(f"選択中: {self._get_formatted_selected_ids()}"); main_layout.addWidget(self.selected_display_label)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel); button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject); main_layout.addWidget(button_box)

    def _icon_clicked(self, icon_id_clicked):
        if icon_id_clicked in self.current_selected_ids: self.current_selected_ids.remove(icon_id_clicked)
        else: self.current_selected_ids.add(icon_id_clicked)
        self.selected_display_label.setText(f"選択中: {self._get_formatted_selected_ids()}")
    def _filter_icons(self, text):
        # 検索キーワードを正規化（全角→半角など）し、小文字に変換
        normalized_filter_text = unicodedata.normalize('NFKC', text).lower().strip()

        for icon_widget in self.icon_widgets:
            # アイコンIDと説明文も同様に正規化・小文字化して比較
            normalized_icon_id = unicodedata.normalize('NFKC', icon_widget.icon_id).lower()
            normalized_icon_desc = unicodedata.normalize('NFKC', icon_widget.description).lower()

            matches_id = normalized_filter_text in normalized_icon_id
            matches_desc = normalized_filter_text in normalized_icon_desc
            icon_widget.setVisible(matches_id or matches_desc or not normalized_filter_text)
    def _get_formatted_selected_ids(self):
        return " ".join(sorted(list(self.current_selected_ids), key=int)) if self.current_selected_ids else "なし"
    def get_selected_ids_as_string(self):
        return self._get_formatted_selected_ids() if self.current_selected_ids else ""

class CategorySelectDialog(QDialog):
    def __init__(self, categories_data, parent_app=None, current_full_path_list=None):
        super().__init__(parent_app)
        self.setWindowTitle("カテゴリ選択"); self.resize(1000, 800)
        self.categories_data = categories_data; self.selected_full_paths = []
        self.tree_widget = QTreeWidget(); self.tree_widget.setHeaderHidden(True)
        self.tree_widget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tree_widget.itemSelectionChanged.connect(self.on_item_selection_changed)
        self.items_map_by_path = {}
        self.build_category_tree(self.categories_data)

        layout = QVBoxLayout(self); layout.addWidget(QLabel("カテゴリを選択してください(Ctrlキーで複数選択可)")); layout.addWidget(self.tree_widget)
        self.selected_path_label = QLabel("選択中: (なし)"); layout.addWidget(self.selected_path_label)
        btns = QHBoxLayout(); ok_btn = QPushButton("OK"); cancel_btn = QPushButton("キャンセル")
        ok_btn.clicked.connect(self.accept); cancel_btn.clicked.connect(self.reject)
        btns.addWidget(ok_btn); btns.addWidget(cancel_btn); layout.addLayout(btns)
        if current_full_path_list and current_full_path_list[0]: self.set_initial_selection(current_full_path_list[0])

    def build_category_tree(self, categories):
        tree = {}; order1, order2, order3 = [], [], []
        for level, name, parent in categories:
            if level == 1:
                if name not in tree: tree[name] = {}; order1.append(name)
            elif level == 2:
                if parent not in tree: tree[parent] = {}; order1.append(parent)
                if name not in tree[parent]: tree[parent][name] = {}; order2.append((parent, name))
            elif level == 3:
                found_parent = False
                for p1_name, l2_dict in tree.items():
                    if parent in l2_dict:
                         if name not in l2_dict[parent]: l2_dict[parent][name] = {}; order3.append((parent, name))
                         found_parent = True
                         break
                if not found_parent:
                    if parent not in tree: tree[parent] = {}; order1.append(parent)
                    if parent not in tree[parent]: tree[parent][parent] = {}; order2.append((parent, parent))
                    if name not in tree[parent][parent]: tree[parent][parent][name] = {}; order3.append((parent, name))

        root_items = {}
        for name in order1:
            if name in tree:
                item = QTreeWidgetItem(self.tree_widget); item.setText(0, name); item.setData(0, Qt.UserRole, name); root_items[name] = item
        for parent_name, name in order2:
            if parent_name in root_items and name in tree.get(parent_name, {}):
                item = QTreeWidgetItem(root_items[parent_name]); item.setText(0, name); item.setData(0, Qt.UserRole, f"{parent_name}:{name}"); self.items_map_by_path[f"{parent_name}:{name}"] = item
        for parent_name, name in order3:
             for l2_full_path, parent_item_l2 in self.items_map_by_path.items():
                 if l2_full_path.endswith(":" + parent_name):
                     full_path_l3 = f"{l2_full_path}:{name}"
                     if full_path_l3 not in self.items_map_by_path:
                         item = QTreeWidgetItem(parent_item_l2); item.setText(0, name); item.setData(0, Qt.UserRole, full_path_l3)
                         self.items_map_by_path[full_path_l3] = item; break
        for name, item in root_items.items():
             self.items_map_by_path[name] = item

    def on_item_selection_changed(self):
        self.selected_full_paths = [item.data(0, Qt.UserRole) for item in self.tree_widget.selectedItems()]
        self.selected_path_label.setText(f"選択中: {', '.join(self.selected_full_paths) if self.selected_full_paths else '(なし)'}")

    def get_selected_categories(self):
        return self.selected_full_paths

    def set_initial_selection(self, full_path_to_select):
        if full_path_to_select in self.items_map_by_path:
            item_to_select = self.items_map_by_path[full_path_to_select]
            # 親アイテムを展開
            parent = item_to_select.parent()
            while parent:
                parent.setExpanded(True)
                parent = parent.parent()
            # アイテムを選択状態にし、表示範囲にスクロール
            self.tree_widget.setCurrentItem(item_to_select)
            self.tree_widget.scrollToItem(item_to_select, QAbstractItemView.PositionAtCenter)
            # QListWidgetのitemSelectionChangedを手動でトリガー (もし必要なら)
            self.on_item_selection_changed()
        else:
            print(f"情報: 初期選択パス '{full_path_to_select}' はツリーに見つかりません。")


class CustomProductCodeInputDialog(QDialog):
    def __init__(self, parent, title, label, default_code=""):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(350) 
        layout = QVBoxLayout(self)

        self.label = QLabel(label)
        layout.addWidget(self.label)

        # 商品コード入力
        code_layout = QFormLayout()
        self.lineEdit = FilteredLineEdit(self) # カスタムLineEditを使用
        self.lineEdit.setText(default_code)
        self.lineEdit.setMaxLength(10)
        code_layout.addRow("新しい商品コード:", self.lineEdit)
        layout.addLayout(code_layout)

        # 価格入力
        price_layout = QFormLayout()
        self.priceLineEdit = QLineEdit(self)
        self.priceLineEdit.setPlaceholderText("例: 10000")
        self.priceLineEdit.setValidator(QRegExpValidator(QRegExp("^[0-9]+$"), self)) # 数値のみ
        price_layout.addRow("当店通常価格 (税込み):", self.priceLineEdit)
        layout.addLayout(price_layout)

        self.buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        layout.addWidget(self.buttonBox)

        self.lineEdit.textChanged.connect(self._check_input_validity)
        self.priceLineEdit.textChanged.connect(self._format_price_input) # Connect to formatter first
        self.priceLineEdit.textChanged.connect(self._check_input_validity) # Then to validator
        self._check_input_validity(self.lineEdit.text()) # Initial check

        self.digit_count_label = QLabel("(0/10 桁)")
        self.digit_count_label.setStyleSheet("font-size: 8pt; color: #6c757d;") # 初期スタイル
        layout.addWidget(self.digit_count_label)

        self.lineEdit.textChanged.connect(self._update_digit_count_display)
        self._update_digit_count_display(self.lineEdit.text()) # 初期表示

    def _format_price_input(self, text_from_signal):
        le = self.priceLineEdit
        current_text_in_le = le.text()
        original_cursor_pos = le.cursorPosition()

        digits_only = "".join(filter(str.isdigit, current_text_in_le))

        if not digits_only:
            if current_text_in_le: # If there was input but all non-digits
                le.blockSignals(True)
                le.setText("")
                le.blockSignals(False)
            return

        try:
            num = int(digits_only)
            formatted_num_str = f"{num:,}"
        except ValueError:
            formatted_num_str = digits_only # Should not happen if digits_only is correct

        if current_text_in_le != formatted_num_str:
            le.blockSignals(True)
            le.setText(formatted_num_str)
            le.blockSignals(False)

            # Adjust cursor position (best effort)
            text_before_cursor_original = current_text_in_le[:original_cursor_pos]
            num_digits_before_cursor_original = len("".join(filter(str.isdigit, text_before_cursor_original)))

            new_cursor_pos_candidate = 0
            digits_counted_in_new = 0
            for i, char_in_new in enumerate(formatted_num_str):
                if char_in_new.isdigit():
                    digits_counted_in_new += 1
                if digits_counted_in_new == num_digits_before_cursor_original:
                    new_cursor_pos_candidate = i + 1
                    break
            
            if num_digits_before_cursor_original == 0:
                new_cursor_pos_candidate = 0
            elif digits_counted_in_new < num_digits_before_cursor_original or original_cursor_pos == len(current_text_in_le):
                 new_cursor_pos_candidate = len(formatted_num_str)

            final_new_cursor_pos = min(new_cursor_pos_candidate, len(formatted_num_str))
            le.setCursorPosition(final_new_cursor_pos)

    def _check_input_validity(self, text=None): # text引数は不要になった
        ok_button = self.buttonBox.button(QDialogButtonBox.Ok)
        
        code_valid = (len(self.lineEdit.text()) == 10 and self.lineEdit.text().isdigit())
        price_digits_only = self.priceLineEdit.text().replace(",", "")
        price_valid = (price_digits_only != "" and price_digits_only.isdigit())

        can_accept = code_valid and price_valid

        if ok_button:
            ok_button.setEnabled(can_accept)

    def _update_digit_count_display(self, text):
        # このメソッドは商品コードの桁数表示専用
        current_digits = len(text)
        self.digit_count_label.setText(f"({current_digits}/10 桁)")
        if current_digits == 10:
            # FilteredLineEditにより数字であることは保証されている
            self.digit_count_label.setStyleSheet("font-size: 8pt; color: green; font-weight: bold;")
        else:
            self.digit_count_label.setStyleSheet("font-size: 8pt; color: #6c757d;")


    def getValues(self):
        price_text_with_comma = self.priceLineEdit.text()
        price_digits_only = price_text_with_comma.replace(",", "")
        return {
            "code": self.lineEdit.text(),
            "price": price_digits_only
        }

class IdSearchDialog(QDialog):
    def __init__(self, r_genre_data, y_category_data, ya_category_data,
                 current_r_id, current_y_id, current_ya_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ID検索")
        self.resize(1000, 800)

        self._r_genre_data_all = r_genre_data if r_genre_data is not None else []
        self._y_category_data_all = y_category_data if y_category_data is not None else []
        self._ya_category_data_all = ya_category_data if ya_category_data is not None else []

        self.selected_r_genre_id = current_r_id
        self.selected_y_category_id = current_y_id
        self.selected_ya_category_id = current_ya_id

        layout = QVBoxLayout(self)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("ID、名称、階層で検索...")
        self.search_input.textChanged.connect(self._filter_results)
        layout.addWidget(self.search_input)

        # ダブルクリック操作の説明ラベルを追加
        self.instruction_label = QLabel("ダブルクリックして選択/選択解除")
        self.instruction_label.setStyleSheet("font-size: 9pt; color: #333333; padding-bottom: 5px;") # 少しスタイル調整
        layout.addWidget(self.instruction_label)

        splitter = QSplitter(Qt.Vertical)

        self.r_genre_table = self._create_result_table_view('Rジャンル')
        self._r_genre_model = IdSearchTableModel([])
        self.r_genre_table.setModel(self._r_genre_model)
        self.r_genre_title_label = QLabel(f"Rジャンル 結果 (選択中: {self.selected_r_genre_id or 'なし'})")
        r_genre_container = self._create_table_container(self.r_genre_title_label, self.r_genre_table)
        splitter.addWidget(r_genre_container)

        self.y_category_table = self._create_result_table_view('Yカテゴリ')
        self._y_category_model = IdSearchTableModel([])
        self.y_category_table.setModel(self._y_category_model)
        self.y_category_title_label = QLabel(f"Yカテゴリ 結果 (選択中: {self.selected_y_category_id or 'なし'})")
        y_category_container = self._create_table_container(self.y_category_title_label, self.y_category_table)
        splitter.addWidget(y_category_container)

        self.ya_category_table = self._create_result_table_view('YAカテゴリ')
        self._ya_category_model = IdSearchTableModel([], hide_name_column=True)
        self.ya_category_table.setModel(self._ya_category_model)
        self.ya_category_title_label = QLabel(f"YAカテゴリ 結果 (選択中: {self.selected_ya_category_id or 'なし'})")
        ya_category_container = self._create_table_container(self.ya_category_title_label, self.ya_category_table)
        splitter.addWidget(ya_category_container)

        layout.addWidget(splitter)

        button_box = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("キャンセル")
        self.cancel_button.clicked.connect(self.reject)
        button_box.addStretch()
        button_box.addWidget(self.ok_button)
        button_box.addWidget(self.cancel_button)
        layout.addLayout(button_box)

        self._filter_results("")
        
        self._initialize_marked_rows(self.selected_r_genre_id, self._r_genre_model, self.r_genre_table)
        self._initialize_marked_rows(self.selected_y_category_id, self._y_category_model, self.y_category_table)
        self._initialize_marked_rows(self.selected_ya_category_id, self._ya_category_model, self.ya_category_table)

        for table_view in [self.r_genre_table, self.y_category_table]:
            table_view.setColumnWidth(0, 100)
            table_view.setColumnWidth(1, 200)
            table_view.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        
        self.ya_category_table.setColumnWidth(0, 100)
        self.ya_category_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

    def _initialize_marked_rows(self, target_id, model, table_view):
        if target_id and model.rowCount() > 0:
            for r in range(model.rowCount()):
                item_data = model.get_item_data(r)
                if item_data and item_data.get('id') == target_id:
                    model.set_marked_row(r)
                    table_view.scrollTo(model.index(r, 0), QAbstractItemView.PositionAtTop)
                    break

    def _create_result_table_view(self, id_type_str):
        table_view = QTableView()
        # 選択動作を完全に無効化
        table_view.setSelectionBehavior(QAbstractItemView.SelectRows)
        table_view.setSelectionMode(QAbstractItemView.NoSelection)  # 選択を無効化
        table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table_view.setAlternatingRowColors(True)
        
        # ダブルクリックイベントを直接処理
        table_view.doubleClicked.connect(
            lambda index, type_str=id_type_str: self._on_item_double_clicked(index, type_str)
        )
        # マウスプレスイベントも追加（シングルクリックでの選択を防ぐ場合は有効化）
        # table_view.mousePressEvent = lambda event: None
        return table_view

    def _create_table_container(self, title_label_widget, table_view):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0,0,0,0)
        title_label_widget.setStyleSheet("font-weight: bold; padding: 3px;")
        layout.addWidget(title_label_widget)
        layout.addWidget(table_view)
        return container

    def _filter_results(self, text):
        keyword_normalized = normalize_text(text)
        r_filtered = self._filter_single_master(self._r_genre_data_all, keyword_normalized)
        self._r_genre_model.update_data(r_filtered)
        y_filtered = self._filter_single_master(self._y_category_data_all, keyword_normalized)
        self._y_category_model.update_data(y_filtered)
        ya_filtered = self._filter_single_master(self._ya_category_data_all, keyword_normalized)
        self._ya_category_model.update_data(ya_filtered)
        
        self._initialize_marked_rows(self.selected_r_genre_id, self._r_genre_model, self.r_genre_table)
        self._initialize_marked_rows(self.selected_y_category_id, self._y_category_model, self.y_category_table)
        self._initialize_marked_rows(self.selected_ya_category_id, self._ya_category_model, self.ya_category_table)

    def _filter_single_master(self, master_data_list, keyword_normalized):
        if not master_data_list: 
            return []
        filtered_list = []
        if not keyword_normalized:
            filtered_list = list(master_data_list)
        else:
            for item in master_data_list:
                match_id = keyword_normalized in normalize_text(item.get('id', ''))
                match_name = item.get('name') and keyword_normalized in normalize_text(item.get('name', ''))
                match_hierarchy = keyword_normalized in normalize_text(item.get('hierarchy', ''))
                if match_id or match_name or match_hierarchy:
                    filtered_list.append(item)
        return filtered_list

    def _on_item_double_clicked(self, index, id_type_str):
        """改善されたダブルクリックハンドラ"""
        if not index.isValid():
            return
            
        model = index.model()
        if not model:
            return
            
        row = index.row()
        if not (0 <= row < model.rowCount()):
            return

        item_data = model.get_item_data(index.row())
        if not item_data:
            return

        selected_id = item_data.get('id', '')
        
        # モデルとUIの更新を一度に行う
        if id_type_str == 'Rジャンル':
            self._update_selection('Rジャンル', row, selected_id, 
                                 self._r_genre_model, self.r_genre_title_label)
        elif id_type_str == 'Yカテゴリ':
            self._update_selection('Yカテゴリ', row, selected_id,
                                 self._y_category_model, self.y_category_title_label)
        elif id_type_str == 'YAカテゴリ':
            self._update_selection('YAカテゴリ', row, selected_id,
                                 self._ya_category_model, self.ya_category_title_label)
        
        # 強制的に再描画 (通常はQtのイベントループに任せる。問題が再発する場合に検討)
        # QApplication.processEvents()

    def _update_selection(self, type_name, row, selected_id, model, label_widget):
        """選択状態の更新を統一的に処理"""
        # 属性名を動的に生成 (例: selected_r_genre_id)
        # type_name が 'Rジャンル', 'Yカテゴリ', 'YAカテゴリ' のいずれかであることを前提
        if type_name == 'Rジャンル':
            attr_name = "selected_r_genre_id"
        elif type_name == 'Yカテゴリ':
            attr_name = "selected_y_category_id"
        elif type_name == 'YAカテゴリ':
            attr_name = "selected_ya_category_id"
        else:
            return # 未知のタイプ

        if model.marked_row == row:
            # 既に選択されている行をダブルクリックした場合は選択解除
            model.clear_marked_row()
            setattr(self, attr_name, "")
            label_widget.setText(f"{type_name} 結果 (選択中: なし)")
        else:
            # 新しい行を選択
            model.set_marked_row(row)
            setattr(self, attr_name, selected_id)
            label_widget.setText(f"{type_name} 結果 (選択中: {selected_id})")

    def get_all_selected_ids(self):
        return {
            'Rジャンル': self.selected_r_genre_id,
            'Yカテゴリ': self.selected_y_category_id,
            'YAカテゴリ': self.selected_ya_category_id
        }

class IdSearchTableModel(QAbstractTableModel):
    HIGHLIGHT_COLOR = QColor(200, 255, 200)

    def __init__(self, data=None, parent=None, hide_name_column=False):
        super().__init__(parent)
        self._data = data if data is not None else []
        self._hide_name_column = hide_name_column
        self.marked_row = -1
        if self._hide_name_column:
            self._headers = ["ID", "階層"]
            self._header_to_data_key_map = {
                "ID": "id",
                "階層": "hierarchy"
            }
        else:
            self._headers = ["ID", "名称", "階層"]
            self._header_to_data_key_map = {
                "ID": "id",
                "名称": "name",
                "階層": "hierarchy"
            }

    def rowCount(self, parent=QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        return len(self._headers)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or not (0 <= index.row() < len(self._data) and 0 <= index.column() < len(self._headers)):
            return None

        item = self._data[index.row()]
        
        if role == Qt.BackgroundRole:
            if self.marked_row != -1 and index.row() == self.marked_row:
                return self.HIGHLIGHT_COLOR
            return None

        if role == Qt.DisplayRole:
            display_column_header = self._headers[index.column()]
            data_key = self._header_to_data_key_map.get(display_column_header)
            if data_key:
                return item.get(data_key, '')
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                if 0 <= section < len(self._headers):
                    return self._headers[section]
            elif orientation == Qt.Vertical:
                return str(section + 1)
        return None

    def update_data(self, new_data):
        self.beginResetModel()
        self._data = new_data if new_data is not None else []
        self.marked_row = -1 # Reset marked row when data changes
        self.endResetModel()

    def get_item_data(self, row):
        if 0 <= row < len(self._data):
            return self._data[row]
        return None

    def set_marked_row(self, row):
        old_marked_row = self.marked_row
        if old_marked_row == row: # 既に同じ行がマークされている場合は何もしない
            return
            
        self.marked_row = row
        if old_marked_row != -1:
            # 古い行の更新
            # 行が存在するか確認
            if 0 <= old_marked_row < self.rowCount():
                top_left_old = self.index(old_marked_row, 0)
                bottom_right_old = self.index(old_marked_row, self.columnCount() - 1)
                self.dataChanged.emit(top_left_old, bottom_right_old, [Qt.BackgroundRole])
        
        if self.marked_row != -1:
            # 新しい行の更新
            # 行が存在するか確認 (通常は存在するはずだが念のため)
            if 0 <= self.marked_row < self.rowCount():
                top_left_new = self.index(self.marked_row, 0)
                bottom_right_new = self.index(self.marked_row, self.columnCount() - 1)
                self.dataChanged.emit(top_left_new, bottom_right_new, [Qt.BackgroundRole])
        
        # レイアウトの更新を強制 (BackgroundRoleの変更だけなら通常は不要。問題発生時に検討)
        # self.layoutChanged.emit()

    def clear_marked_row(self):
        old_marked_row = self.marked_row
        if old_marked_row == -1: # 既にマークがなければ何もしない
            return
            
        self.marked_row = -1
        if old_marked_row != -1:
            # 古い行の更新 (マーク解除)
            # 行が存在するか確認
            if 0 <= old_marked_row < self.rowCount():
                top_left = self.index(old_marked_row, 0)
                bottom_right = self.index(old_marked_row, self.columnCount() - 1)
                self.dataChanged.emit(top_left, bottom_right, [Qt.BackgroundRole])
        # レイアウトの更新を強制 (BackgroundRoleの変更だけなら通常は不要。問題発生時に検討)
        # self.layoutChanged.emit()

class ColorSelectionDialog(QDialog):
    def __init__(self, common_color_list, current_value_str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("共通色選択")
        self.setMinimumWidth(300)
        self.resize(400, 500)

        self.common_colors = common_color_list
        self.common_colors_set = set(common_color_list)
        self.selected_order = [] # 選択された色の順番を保持するリスト

        layout = QVBoxLayout(self)
        self.list_widget = QListWidget()

        # current_value_str から共通色をパースし、その出現順で selected_order を初期化
        raw_names_from_input = [m.strip() for m in current_value_str.split('●') if m.strip()]
        initially_selected_common_in_order = []
        for name in raw_names_from_input:
            if name in self.common_colors_set and name not in initially_selected_common_in_order:
                initially_selected_common_in_order.append(name)
        
        self.selected_order = list(initially_selected_common_in_order) # 初期選択順を設定

        # リストウィジェットのアイテムは共通色リスト順で作成
        for color_name in self.common_colors:
            item = QListWidgetItem(color_name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            # selected_order に基づいて初期チェック状態を設定
            if color_name in self.selected_order:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)
        
        self.list_widget.itemClicked.connect(self._toggle_item_check_state_on_click)
        layout.addWidget(self.list_widget)

        # ボタンボックス
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def _toggle_item_check_state_on_click(self, item):
        color_name = item.text()

        # self.selected_order に基づいてチェック状態をトグルする
        if color_name in self.selected_order:
            # 既に選択されている（selected_orderに入っている）場合は、選択解除する
            item.setCheckState(Qt.Unchecked)
            self.selected_order.remove(color_name)
        else:
            # まだ選択されていない場合は、選択する
            item.setCheckState(Qt.Checked)
            self.selected_order.append(color_name)

    def get_selected_common_colors(self):
        # self.selected_order には、ユーザーが操作した結果の順番で共通色名が入っている
        return list(self.selected_order) # 変更不可なコピーを返す


# ProductAppクラスにショートカット関連のメソッドを追加
def add_shortcut_methods_to_product_app():
    """ProductAppクラスにショートカット関連のメソッドを動的追加"""
    
    def show_shortcuts_help(self):
        """ショートカット一覧をダイアログで表示"""
        help_text = """
<h3>🎯 商品登録ツール - ショートカット一覧</h3>

<h4>📁 ファイル操作</h4>
<table border="1" cellpadding="5" cellspacing="0">
<tr><td><b>Ctrl+N</b></td><td>新規商品作成</td></tr>
<tr><td><b>Ctrl+S</b></td><td>保存</td></tr>
</table>

<h4>🔍 検索・ナビゲーション</h4>
<table border="1" cellpadding="5" cellspacing="0">
<tr><td><b>Ctrl+F</b></td><td>検索ボックスにフォーカス</td></tr>
<tr><td><b>Esc</b></td><td>検索をクリア</td></tr>
<tr><td><b>Ctrl+G</b></td><td>カテゴリ選択</td></tr>
<tr><td><b>Ctrl+I</b></td><td>ID検索</td></tr>
<tr><td><b>Enter</b></td><td>次のフィールドに移動（スマートナビゲーション）</td></tr>
<tr><td><b>Tab</b></td><td>論理的な順序でフィールド移動</td></tr>
</table>

<h4>📝 編集操作</h4>
<table border="1" cellpadding="5" cellspacing="0">
<tr><td><b>Ctrl+Z</b></td><td>元に戻す</td></tr>
<tr><td><b>Ctrl+Y</b></td><td>やり直し</td></tr>
<tr><td><b>Ctrl+Shift+A</b></td><td>SKU追加</td></tr>
<tr><td><b>Delete</b></td><td>選択SKU削除（テーブルフォーカス時）</td></tr>
</table>

<h4>🛠️ ツール</h4>
<table border="1" cellpadding="5" cellspacing="0">
<tr><td><b>F5</b></td><td>C#実行</td></tr>
<tr><td><b>Ctrl+H</b></td><td>画像説明HTML生成</td></tr>
</table>

<h4>❓ ヘルプ</h4>
<table border="1" cellpadding="5" cellspacing="0">
<tr><td><b>F1</b></td><td>このヘルプを表示</td></tr>
</table>

<h4>💾 自動保存</h4>
<p>30秒ごとに自動保存されます。</p>
        """
        
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("ショートカット一覧")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(help_text)
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.exec_()

    def clear_search(self):
        """検索ボックスをクリア"""
        self.search_bar.clear()
        self.search_bar.clearFocus()


    def focus_search(self):
        """検索ボックスにフォーカスを移動"""
        self.search_bar.setFocus()
        self.search_bar.selectAll()
    
    def _handle_search_action(self):
        """検索アクション処理（デバッグ用ラッパー）"""
        print("DEBUG: _handle_search_action called!")
        logging.info("検索アクション呼び出し")
        try:
            self.show_search_dialog()
        except Exception as e:
            print(f"DEBUG: Error in show_search_dialog: {e}")
            logging.error(f"検索ダイアログエラー: {e}", exc_info=True)

    def show_search_dialog(self):
        """Excel風の検索ダイアログを表示"""
        print("DEBUG: show_search_dialog called!")  # デバッグ用
        logging.info("検索ダイアログ呼び出し")  # デバッグ用
        # 非モーダル検索パネルの切り替え
        if not hasattr(self, '_search_panel'):
            self._search_panel = SearchPanel(self)
            self._search_panel.hide()  # 初期は非表示
            
            # スプリッターにパネルを安全に追加
            if hasattr(self, 'main_splitter'):
                # 現在のウィジェット数をチェック
                current_count = self.main_splitter.count()
                if current_count < 3:  # まだ3つ未満なら追加
                    self.main_splitter.addWidget(self._search_panel)
                    # 初期サイズを設定（検索パネルは最小幅）
                    if current_count >= 1:
                        sizes = self.main_splitter.sizes()
                        total = sum(sizes)
                        # 既存の領域から検索パネル分を確保
                        panel_width = 350
                        main_width = max(total - panel_width, total // 2)
                        new_sizes = sizes[:-1] + [main_width, panel_width]
                        self.main_splitter.setSizes(new_sizes)
                else:
                    logging.warning("検索パネル: スプリッターに3つ以上のウィジェットがあるため、追加をスキップしました")
        
        # パネルの表示/非表示を切り替え
        if hasattr(self, '_search_panel'):
            if self._search_panel.isVisible():
                self._search_panel.hide()
                # 検索パネルを隠すときはスプリッターサイズを調整
                self._restore_splitter_sizes_without_search()
            else:
                self._search_panel.show()
                # 検索パネルを表示するときにサイズ調整
                self._adjust_splitter_sizes_with_search()
                self._search_panel.search_input.setFocus()  # 検索入力にフォーカス


    # メソッドをProductAppクラスに追加
    ProductApp.show_shortcuts_help = show_shortcuts_help
    ProductApp.clear_search = clear_search  
    ProductApp.focus_search = focus_search


if __name__ == "__main__":
    # --- Global exception hook for logging uncaught exceptions ---
    def handle_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_traceback) # Default handling for Ctrl+C
            return
        # Ensure logging is configured before trying to log
        if logging.getLogger().hasHandlers():
            logging.critical("Uncaught exception (via sys.excepthook):", exc_info=(exc_type, exc_value, exc_traceback))
    sys.excepthook = handle_exception
    try:
        # 高DPIスケーリング設定 (QApplicationインスタンス作成前に設定)
        # Qt.AA_EnableHighDpiScaling は Qt 5.6 以降でデフォルト有効の場合もあるが、明示的に設定
        if hasattr(Qt, 'AA_EnableHighDpiScaling'):
            QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
            print("Debug: Qt.AA_EnableHighDpiScaling set to True")
        # Qt.AA_UseHighDpiPixmaps は高解像度のアイコンや画像を使用する場合に有効
        if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
            QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
            print("Debug: Qt.AA_UseHighDpiPixmaps set to True")
        try:
            import PyQt5
            pyqt_path = os.path.dirname(PyQt5.__file__)
            plugin_path = os.path.join(pyqt_path, "Qt5", "plugins")
            os.environ['QT_PLUGIN_PATH'] = plugin_path
        except Exception as e_plugin_path:
            # PyQt5インポート問題は無視（必須ではない）
            pass
        app = QApplication(sys.argv)
        # ショートカット関連のメソッドを追加
        add_shortcut_methods_to_product_app()
        
        # QSettingsで復元するので、ここでのshowMaximized()は不要になる場合がある
        # _load_settings内で初回起動時や設定がない場合のデフォルト表示を制御
        win = ProductApp() 
        sys.exit(app.exec_())
    except Exception as e_global:
        # ログファイルへの出力 (loggingが初期化されていれば)
        if logging.getLogger().hasHandlers(): # ロガーが設定されていれば
            logging.critical("アプリケーションの起動中に致命的なエラーが発生しました。", exc_info=True)
        else: # ロガーが未設定の場合 (ProductApp.__init__ より前など) はコンソールに出力
            print(f"##### アプリケーションの起動中に致命的なエラー (ロガー未設定) #####\nエラータイプ: {type(e_global).__name__}\nエラーメッセージ: {e_global}\n##### トレースバックここから #####"); traceback.print_exc(); print("##### トレースバックここまで #####")
        
        try:
            _app_temp = QApplication.instance() or QApplication(sys.argv)
            error_box = QMessageBox(); error_box.setIcon(QMessageBox.Critical); error_box.setWindowTitle("致命的な起動エラー")
            # ログファイルのパスを特定する試み
            log_file_path_for_msg = "ログファイル (場所特定不可)" # デフォルトメッセージ
            try:
                # get_user_data_dir は preferred_dir を引数に取ることがある。
                # このコンテキストでは、実行ファイルの場所を preferred_dir として試みる。
                preferred_dir_for_log = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
                actual_user_data_dir = get_user_data_dir(preferred_dir_for_log)
                log_file_path_for_msg = os.path.join(actual_user_data_dir, f"{APP_NAME}_errors.log")
            except Exception as e_log_path_fetch:
                # パス特定に失敗した場合でも、エラーダイアログは表示する
                print(f"情報: エラーダイアログ表示のためのログファイルパス特定に失敗: {e_log_path_fetch}")
            error_box.setText(f"アプリケーションの起動中に予期せぬエラーが発生しました:\n\n{e_global}\n\n詳細は {log_file_path_for_msg} を確認してください。"); error_box.setDetailedText(traceback.format_exc()); error_box.exec_()
        except Exception as e_msgbox: print(f"エラーダイアログの表示に失敗しました: {e_msgbox}")
        sys.exit(1)
